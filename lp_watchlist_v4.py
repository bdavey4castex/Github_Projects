"""
LP Watchlist & Enhanced Monitoring - Automated Monthly Report
=============================================================
Runs on the 1st of each month. Pulls prior month's data, generates
the full enhanced monitoring report, and exports to Excel.

Schedule via Windows Task Scheduler for 4:00 AM Central on the 1st.

v2 Changes:
  - Replaced binary review_status (ALERT FLAGGED / REVIEW REQUIRED / ALERT CLEARED)
    with a continuous risk score (0-100) calculated in Python post-query.
  - Scoring categories: Position & Outcome (25), Front-Running (25),
    Price Impact (20), Aggression Pattern (20).
  - Risk tiers: Low (0-25), Moderate (26-50), Elevated (51-75), High (76-100).
  - Excel tabs now split by tier instead of old alert status.
  - UL tabs removed; underlying asset integration planned for Phase 2
    (will be merged into consolidated alert tab with score breakdown).
  - underlying_directions hook in apply_risk_scores() ready for Yahoo Finance
    integration to apply price impact discounts/penalties.

v4 Changes:
  - P&L NaN fix: guard all None/NaN before arithmetic in calculate_risk_score().
  - P&L log scale recalibrated: log base raised from 2,501 to 50,001 so the
    scale breathes more — $2,500 profit now ~17pts, $10,000 ~21pts, $25,000+ ~25pts.
    Less compression at the low end, meaningful differentiation across the range.
  - Category 5 (Price Pushing) redesigned with two context-aware sub-signals:
      * Aggression Density (max 5): lp_aggressor_count / total_volume * 1000.
        Rewards concern for concentrated aggression in thin markets. 4+ trades
        per 1K contracts = 5pts; 2-4 = 3pts; 1-2 = 1pt.
      * Cluster Impact (max 5): LP aggression window compressed in time AND
        responsible for a large fraction of the day's price move. Computed as:
        (1 - window_fraction_of_day) * abs(lp_impact_pct). Tight window + big
        price move = high cluster score.
    Auto-max (10pts) now requires BOTH density >= 3.0 AND cluster score >= 4,
    replacing the raw 10-trade count threshold.
"""

import os
import sys
import psycopg2
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path

# =============================================================================
# CONFIGURATION
# =============================================================================
DB_CONFIG = {
    "host": os.environ.get("FXDB_HOST", "web-prod-db.cla246ai8gve.us-east-1.rds.amazonaws.com"),
    "port": int(os.environ.get("FXDB_PORT", 5432)),
    "database": os.environ.get("FXDB_NAME", "postgres"),
    "user": os.environ.get("FXDB_USER", "postgres_read_only"),
    "password": os.environ.get("FXDB_PASS", "pgreadonlypw"),
}

LP_ACCOUNTS = ("UIBGIECTC", "UIBGIECTMC")

# Output directory - change to wherever you want the reports saved
OUTPUT_DIR = Path.home() / "Documents" / "LP_Watchlist_Reports"

# =============================================================================
# DATE CALCULATION
# =============================================================================
def get_prior_month_range():
    """Returns (start_date, end_date) for the prior month."""
    today = datetime.now()
    first_of_this_month = today.replace(day=1)
    last_of_prior_month = first_of_this_month - timedelta(days=1)
    first_of_prior_month = last_of_prior_month.replace(day=1)
    return first_of_prior_month.strftime("%Y-%m-%d"), last_of_prior_month.strftime("%Y-%m-%d")


# =============================================================================
# SQL QUERY BUILDER
# =============================================================================
def build_query(start_date, end_date):
    """Builds the full enhanced monitoring query with both fixes applied."""

    return f"""
WITH
-- ==========================================================================
-- STAGE 1: WATCHLIST FLAGGING
-- ==========================================================================

-- Filter 1: Minimum volume (1000 contracts)
min_vol AS (
    SELECT p.trading_date, p.instrument_id
    FROM pairs p
    WHERE trading_date BETWEEN '{start_date}' AND '{end_date}'
    GROUP BY p.trading_date, p.instrument_id
    HAVING SUM(p.quantity) >= 1000
),

-- Filter 2: Market depth (>= 3 customer accounts)
market_depth AS (
    SELECT sub.trading_date, sub.instrument_id
    FROM (
        SELECT instrument_id, trading_date, aggressor_account AS account FROM pairs
        WHERE trading_date BETWEEN '{start_date}' AND '{end_date}'
        UNION
        SELECT instrument_id, trading_date, passive_account AS account FROM pairs
        WHERE trading_date BETWEEN '{start_date}' AND '{end_date}'
    ) sub
    JOIN accounts a ON a.account_id = sub.account
    WHERE a.account_type = 'CUSTOMER'
    GROUP BY sub.trading_date, sub.instrument_id
    HAVING COUNT(DISTINCT sub.account) >= 3
),

-- Filtered pairings (pass Filter 1 + 2) for flagging purposes
account_pairings AS (
    SELECT p.instrument_id, p.trading_date, p.aggressor_account AS account, p.quantity,
        'TRUE' AS aggressor, p.yes_price, p.pair_time,
        p.aggressor_symbol_subtype AS symbol_subtype
    FROM pairs p
    JOIN min_vol mv ON mv.trading_date = p.trading_date AND mv.instrument_id = p.instrument_id
    JOIN market_depth md ON md.trading_date = p.trading_date AND md.instrument_id = p.instrument_id
    UNION ALL
    SELECT p.instrument_id, p.trading_date, p.passive_account AS account, p.quantity,
        'FALSE' AS aggressor, p.yes_price, p.pair_time,
        p.passive_symbol_subtype AS symbol_subtype
    FROM pairs p
    JOIN min_vol mv ON mv.trading_date = p.trading_date AND mv.instrument_id = p.instrument_id
    JOIN market_depth md ON md.trading_date = p.trading_date AND md.instrument_id = p.instrument_id
),

lp_pairings AS (
    SELECT * FROM account_pairings
    WHERE account IN {LP_ACCOUNTS}
),

-- LP activity on filtered product-days (for flagging)
lp_all AS (
    SELECT lp.trading_date, lp.instrument_id, lp.account,
        COUNT(*) AS lp_pairs,
        SUM(lp.quantity) AS lp_vol,
        SUM(lp.quantity) FILTER (WHERE lp.aggressor = 'TRUE') AS lp_aggressor_vol,
        COUNT(*) FILTER (WHERE lp.aggressor = 'TRUE') AS lp_aggressor_count,
        (COALESCE(SUM(lp.quantity) FILTER (WHERE lp.aggressor = 'TRUE'), 0))::numeric
            / NULLIF(SUM(quantity)::numeric, 0) AS lp_aggressor_pct
    FROM lp_pairings lp
    GROUP BY trading_date, instrument_id, account
),

lp_all_with_category AS (
    SELECT la.trading_date, la.instrument_id, pr.category, pr.category_l2,
        CASE WHEN pr.category IN ('Financial Markets', 'Environmental') THEN pr.category_l2 ELSE pr.category END AS stat_group,
        la.account, la.lp_pairs, la.lp_vol, la.lp_aggressor_count, la.lp_aggressor_vol, la.lp_aggressor_pct
    FROM lp_all la
    JOIN instrument_definitions id ON id.instrument_id = la.instrument_id
    JOIN products pr ON pr.product_id = id.product_id
),

-- Filter 3: At least 5 aggressive pairings to be flaggable
lp_flaggable AS (
    SELECT * FROM lp_all_with_category
    WHERE COALESCE(lp_aggressor_count, 0) >= 5
      AND COALESCE(lp_aggressor_vol, 0) > 0
),

-- ==========================================================================
-- FIX 1: BROADENED STATS BASELINE
-- All LP product-days in the subcategory, NO volume/depth filter,
-- INCLUDING 0% aggression days.
-- ==========================================================================
all_lp_activity AS (
    SELECT
        p.trading_date,
        p.instrument_id,
        lp_acc.account,
        pr.category,
        pr.category_l2,
        CASE WHEN pr.category IN ('Financial Markets', 'Environmental') THEN pr.category_l2 ELSE pr.category END AS stat_group,
        COALESCE(SUM(p.quantity) FILTER (WHERE
            (p.aggressor_account = lp_acc.account AND 'TRUE' = 'TRUE')
        ), 0) AS agg_vol,
        SUM(p.quantity) AS total_vol
    FROM pairs p
    CROSS JOIN (SELECT unnest(ARRAY{list(LP_ACCOUNTS)}) AS account) lp_acc
    JOIN instrument_definitions id ON id.instrument_id = p.instrument_id
    JOIN products pr ON pr.product_id = id.product_id
    WHERE p.trading_date BETWEEN '{start_date}' AND '{end_date}'
      AND (p.aggressor_account = lp_acc.account OR p.passive_account = lp_acc.account)
    GROUP BY p.trading_date, p.instrument_id, lp_acc.account, pr.category, pr.category_l2
),

all_lp_with_pct AS (
    SELECT *,
        CASE WHEN total_vol > 0 THEN agg_vol::numeric / total_vol::numeric ELSE 0 END AS lp_aggressor_pct_full
    FROM all_lp_activity
),

category_stats AS (
    SELECT stat_group,
        AVG(lp_aggressor_pct_full) AS group_mean_aggression,
        COALESCE(STDDEV(lp_aggressor_pct_full), 0) AS group_stdev_aggression,
        AVG(lp_aggressor_pct_full) + COALESCE(STDDEV(lp_aggressor_pct_full), 0) AS group_threshold,
        COUNT(*) AS group_product_days
    FROM all_lp_with_pct
    GROUP BY stat_group
),

-- Flagged rows: pass all 4 filters
flagged AS (
    SELECT lf.*
    FROM lp_flaggable lf
    JOIN category_stats cs ON cs.stat_group = lf.stat_group
    WHERE lf.lp_aggressor_pct > cs.group_threshold
),

-- ==========================================================================
-- STAGE 2: ENHANCED MONITORING
-- ==========================================================================

expiration_info AS (
    SELECT
        f.trading_date, f.instrument_id, f.account,
        idef.expiration_date, idef.settlement_price,
        EXTRACT(DAY FROM (idef.expiration_date - f.trading_date))::int AS days_to_expiry,
        CASE
            WHEN EXTRACT(DAY FROM (idef.expiration_date - f.trading_date))::int <= 1 THEN 'LOW PRIORITY - EXPIRATION'
            WHEN EXTRACT(DAY FROM (idef.expiration_date - f.trading_date))::int <= 3 THEN 'MEDIUM - NEAR EXPIRATION'
            ELSE 'REVIEW'
        END AS expiry_priority
    FROM flagged f
    JOIN instrument_definitions idef ON idef.instrument_id = f.instrument_id
),

all_trades AS (
    SELECT p.instrument_id, p.trading_date, p.pair_time, p.quantity, p.yes_price,
        p.aggressor_account, p.passive_account,
        p.aggressor_symbol_subtype, p.passive_symbol_subtype
    FROM pairs p
    WHERE EXISTS (
        SELECT 1 FROM flagged f
        WHERE f.instrument_id = p.instrument_id AND f.trading_date = p.trading_date
    )
),

price_timeline AS (
    SELECT instrument_id, trading_date,
        (ARRAY_AGG(yes_price ORDER BY pair_time ASC))[1] AS open_price,
        (ARRAY_AGG(yes_price ORDER BY pair_time DESC))[1] AS close_price,
        MIN(yes_price) AS low_price, MAX(yes_price) AS high_price,
        AVG(yes_price) AS avg_price,
        COUNT(*) AS total_trades, SUM(quantity) AS total_volume
    FROM all_trades
    GROUP BY instrument_id, trading_date
),

lp_aggressive_trades AS (
    SELECT at.instrument_id, at.trading_date, at.pair_time, at.quantity, at.yes_price,
        at.aggressor_account AS lp_account, at.aggressor_symbol_subtype AS lp_side,
        CASE WHEN at.aggressor_symbol_subtype = 'YES' THEN at.yes_price ELSE (1 - at.yes_price) END AS lp_effective_cost
    FROM all_trades at
    WHERE at.aggressor_account IN {LP_ACCOUNTS}
),

lp_passive_trades AS (
    SELECT at.instrument_id, at.trading_date, at.pair_time, at.quantity, at.yes_price,
        at.passive_account AS lp_account, at.passive_symbol_subtype AS lp_side,
        CASE WHEN at.passive_symbol_subtype = 'YES' THEN at.yes_price ELSE (1 - at.yes_price) END AS lp_effective_cost
    FROM all_trades at
    WHERE at.passive_account IN {LP_ACCOUNTS}
),

lp_agg_summary AS (
    SELECT instrument_id, trading_date, lp_account,
        MIN(pair_time) AS first_agg_time, MAX(pair_time) AS last_agg_time,
        AVG(yes_price) FILTER (WHERE lp_side = 'YES') AS lp_agg_avg_yes_price,
        AVG(yes_price) FILTER (WHERE lp_side = 'NO') AS lp_agg_avg_no_price,
        SUM(quantity) FILTER (WHERE lp_side = 'YES') AS lp_agg_yes_qty,
        SUM(quantity) FILTER (WHERE lp_side = 'NO') AS lp_agg_no_qty,
        AVG(lp_effective_cost) AS lp_agg_avg_cost
    FROM lp_aggressive_trades
    GROUP BY instrument_id, trading_date, lp_account
),

lp_position AS (
    SELECT instrument_id, trading_date, lp_account,
        SUM(yes_qty) AS total_yes_qty, SUM(no_qty) AS total_no_qty,
        SUM(yes_qty) - SUM(no_qty) AS net_yes_position,
        CASE
            WHEN SUM(yes_qty) - SUM(no_qty) > 0 THEN 'NET YES BUYER'
            WHEN SUM(yes_qty) - SUM(no_qty) < 0 THEN 'NET NO BUYER'
            ELSE 'FLAT'
        END AS position_direction
    FROM (
        SELECT instrument_id, trading_date, lp_account,
            SUM(quantity) FILTER (WHERE lp_side = 'YES') AS yes_qty,
            SUM(quantity) FILTER (WHERE lp_side = 'NO') AS no_qty
        FROM lp_aggressive_trades GROUP BY instrument_id, trading_date, lp_account
        UNION ALL
        SELECT instrument_id, trading_date, lp_account,
            SUM(quantity) FILTER (WHERE lp_side = 'YES') AS yes_qty,
            SUM(quantity) FILTER (WHERE lp_side = 'NO') AS no_qty
        FROM lp_passive_trades GROUP BY instrument_id, trading_date, lp_account
    ) combined
    GROUP BY instrument_id, trading_date, lp_account
),

prev_day_price AS (
    SELECT DISTINCT ON (p.instrument_id, f.trading_date)
        f.trading_date, p.instrument_id, p.yes_price AS prev_day_last_price
    FROM flagged f
    JOIN pairs p ON p.instrument_id = f.instrument_id AND p.trading_date < f.trading_date
    ORDER BY p.instrument_id, f.trading_date, p.pair_time DESC
),

price_impact AS (
    SELECT las.instrument_id, las.trading_date, las.lp_account,
        COALESCE(
            (SELECT AVG(at2.yes_price) FROM all_trades at2
             WHERE at2.instrument_id = las.instrument_id AND at2.trading_date = las.trading_date
               AND at2.pair_time < las.first_agg_time),
            pdp.prev_day_last_price
        ) AS price_before_lp,
        COALESCE(
            (SELECT AVG(at2.yes_price) FROM all_trades at2
             WHERE at2.instrument_id = las.instrument_id AND at2.trading_date = las.trading_date
               AND at2.pair_time > las.last_agg_time),
            pt.close_price
        ) AS price_after_lp
    FROM lp_agg_summary las
    LEFT JOIN prev_day_price pdp ON pdp.instrument_id = las.instrument_id AND pdp.trading_date = las.trading_date
    LEFT JOIN price_timeline pt ON pt.instrument_id = las.instrument_id AND pt.trading_date = las.trading_date
),

customer_prices AS (
    SELECT las.instrument_id, las.trading_date, las.lp_account,
        AVG(at.yes_price) AS customer_avg_price_during_lp,
        COUNT(*) AS customer_trades_during_lp
    FROM lp_agg_summary las
    JOIN all_trades at ON at.instrument_id = las.instrument_id
        AND at.trading_date = las.trading_date
        AND at.pair_time BETWEEN las.first_agg_time AND las.last_agg_time
        AND at.aggressor_account NOT IN {LP_ACCOUNTS}
        AND at.passive_account NOT IN {LP_ACCOUNTS}
    GROUP BY las.instrument_id, las.trading_date, las.lp_account
),

front_running AS (
    SELECT lat.instrument_id, lat.trading_date, lat.lp_account,
        AVG(CASE WHEN lat.lp_side = 'YES' THEN ct.yes_price - lat.yes_price
                 WHEN lat.lp_side = 'NO' THEN lat.yes_price - ct.yes_price END
        ) FILTER (WHERE ct.pair_time > lat.pair_time AND ct.pair_time <= lat.pair_time + INTERVAL '10 seconds'
        ) AS customer_disadvantage_10s,
        COUNT(*) FILTER (WHERE ct.pair_time > lat.pair_time AND ct.pair_time <= lat.pair_time + INTERVAL '10 seconds'
        ) AS customer_trades_within_10s,
        AVG(CASE WHEN lat.lp_side = 'YES' THEN ct.yes_price - lat.yes_price
                 WHEN lat.lp_side = 'NO' THEN lat.yes_price - ct.yes_price END
        ) FILTER (WHERE ct.pair_time > lat.pair_time AND ct.pair_time <= lat.pair_time + INTERVAL '30 seconds'
        ) AS customer_disadvantage_30s,
        COUNT(*) FILTER (WHERE ct.pair_time > lat.pair_time AND ct.pair_time <= lat.pair_time + INTERVAL '30 seconds'
        ) AS customer_trades_within_30s,
        AVG(CASE WHEN lat.lp_side = 'YES' THEN ct.yes_price - lat.yes_price
                 WHEN lat.lp_side = 'NO' THEN lat.yes_price - ct.yes_price END
        ) FILTER (WHERE ct.pair_time > lat.pair_time AND ct.pair_time <= lat.pair_time + INTERVAL '60 seconds'
        ) AS customer_disadvantage_60s,
        COUNT(*) FILTER (WHERE ct.pair_time > lat.pair_time AND ct.pair_time <= lat.pair_time + INTERVAL '60 seconds'
        ) AS customer_trades_within_60s
    FROM lp_aggressive_trades lat
    JOIN all_trades ct ON ct.instrument_id = lat.instrument_id
        AND ct.trading_date = lat.trading_date
        AND ct.aggressor_account NOT IN {LP_ACCOUNTS}
        AND ct.passive_account NOT IN {LP_ACCOUNTS}
        AND ct.pair_time > lat.pair_time
        AND ct.pair_time <= lat.pair_time + INTERVAL '60 seconds'
    GROUP BY lat.instrument_id, lat.trading_date, lat.lp_account
),

last_trade AS (
    SELECT DISTINCT ON (p.instrument_id)
        p.instrument_id, p.pair_time AS last_trade_time,
        p.trading_date AS last_trade_date, p.yes_price AS last_trade_price,
        CASE WHEN p.yes_price >= 0.90 THEN 'LIKELY YES'
             WHEN p.yes_price <= 0.10 THEN 'LIKELY NO'
             ELSE 'UNCERTAIN' END AS likely_outcome
    FROM pairs p
    WHERE p.instrument_id IN (SELECT instrument_id FROM flagged)
    ORDER BY p.instrument_id, p.pair_time DESC
),

-- ==========================================================================
-- BOD (BEGINNING OF DAY) POSITION
-- Net position carried into the flagged day from the prior 90 days.
-- Scoped to flagged instruments only for performance.
-- ==========================================================================
bod_position AS (
    SELECT
        f.instrument_id,
        f.trading_date,
        f.account,
        COALESCE(SUM(yes_qty) - SUM(no_qty), 0) AS bod_net_yes_position,
        CASE
            WHEN COALESCE(SUM(yes_qty) - SUM(no_qty), 0) > 0 THEN 'NET YES BUYER'
            WHEN COALESCE(SUM(yes_qty) - SUM(no_qty), 0) < 0 THEN 'NET NO BUYER'
            ELSE 'FLAT'
        END AS bod_position_direction
    FROM flagged f
    LEFT JOIN (
        SELECT
            p.instrument_id,
            p.trading_date AS trade_date,
            lp_acc.account,
            SUM(p.quantity) FILTER (
                WHERE p.aggressor_account = lp_acc.account
                  AND p.aggressor_symbol_subtype = 'YES'
            ) AS yes_qty,
            SUM(p.quantity) FILTER (
                WHERE p.aggressor_account = lp_acc.account
                  AND p.aggressor_symbol_subtype = 'NO'
            ) AS no_qty
        FROM pairs p
        CROSS JOIN (SELECT unnest(ARRAY{list(LP_ACCOUNTS)}) AS account) lp_acc
        WHERE p.trading_date >= (
                SELECT MIN(f2.trading_date) - INTERVAL '90 days' FROM flagged f2
              )
          AND p.instrument_id IN (SELECT instrument_id FROM flagged)
          AND (p.aggressor_account = lp_acc.account OR p.passive_account = lp_acc.account)
        GROUP BY p.instrument_id, p.trading_date, lp_acc.account
    ) prior ON prior.instrument_id = f.instrument_id
           AND prior.account = f.account
           AND prior.trade_date < f.trading_date
           AND prior.trade_date >= f.trading_date - INTERVAL '90 days'
    GROUP BY f.instrument_id, f.trading_date, f.account
),

-- ==========================================================================
-- LP BURST DETECTION
-- For each flagged LP product-day, find the maximum number of LP aggressive
-- trades occurring within any rolling 60s / 300s / 600s window.
-- Strategy: for every LP aggressive trade, count how many OTHER LP aggressive
-- trades on the same instrument-day fall within the window starting at that
-- trade's pair_time. MAX over all anchor trades = tightest burst.
-- Pre-filtered to flagged instruments only; scoped to LP aggressor trades only
-- to keep the self-join small.
-- ==========================================================================
lp_agg_trades_flagged AS (
    SELECT
        p.instrument_id,
        p.trading_date,
        p.aggressor_account AS account,
        p.pair_time
    FROM pairs p
    WHERE p.trading_date BETWEEN '{start_date}' AND '{end_date}'
      AND p.aggressor_account IN {LP_ACCOUNTS}
      AND EXISTS (
          SELECT 1 FROM flagged f
          WHERE f.instrument_id = p.instrument_id
            AND f.trading_date  = p.trading_date
            AND f.account       = p.aggressor_account
      )
),

lp_burst AS (
    SELECT
        a.instrument_id,
        a.trading_date,
        a.account,
        MAX(cnt_60s)  AS max_burst_60s,
        MAX(cnt_300s) AS max_burst_300s,
        MAX(cnt_600s) AS max_burst_600s
    FROM lp_agg_trades_flagged a
    JOIN LATERAL (
        SELECT
            COUNT(*) FILTER (WHERE b.pair_time < a.pair_time + INTERVAL '60 seconds')  AS cnt_60s,
            COUNT(*) FILTER (WHERE b.pair_time < a.pair_time + INTERVAL '300 seconds') AS cnt_300s,
            COUNT(*) FILTER (WHERE b.pair_time < a.pair_time + INTERVAL '600 seconds') AS cnt_600s
        FROM lp_agg_trades_flagged b
        WHERE b.instrument_id = a.instrument_id
          AND b.trading_date  = a.trading_date
          AND b.account       = a.account
          AND b.pair_time    >= a.pair_time  -- anchor: count from this trade forward
    ) burst ON true
    GROUP BY a.instrument_id, a.trading_date, a.account
)

-- ==========================================================================
-- FINAL OUTPUT
-- ==========================================================================
-- outcome_price: 3-tier fallback
--   1) settlement_price (if populated)
--   2) $1.00 or $0.00 (if expired and last trade indicates clear outcome)
--   3) last_trade_price (if not expired / uncertain)
SELECT
    f.trading_date,
    f.instrument_id,
    f.category,
    f.category_l2,
    f.stat_group,
    f.account,
    ROUND(f.lp_aggressor_pct * 100, 2) AS lp_aggressor_pct,

    ROUND(cs.group_mean_aggression * 100, 2) AS group_mean_pct,
    ROUND(cs.group_threshold * 100, 2) AS group_threshold_pct,
    cs.group_product_days,

    ei.expiration_date,
    ei.days_to_expiry,
    ei.expiry_priority,
    ei.settlement_price,

    lt.last_trade_date,
    lt.last_trade_price,
    lt.likely_outcome,

    -- Resolved outcome price
    COALESCE(
        ei.settlement_price,
        CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00
             WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00
             ELSE NULL END,
        lt.last_trade_price
    ) AS outcome_price,
    CASE
        WHEN ei.settlement_price IS NOT NULL THEN 'SETTLEMENT'
        WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 'EXPIRED - YES ($1.00)'
        WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 'EXPIRED - NO ($0.00)'
        WHEN ei.expiration_date < CURRENT_DATE THEN 'EXPIRED - UNCERTAIN'
        ELSE 'LAST TRADE (NOT EXPIRED)'
    END AS outcome_price_source,

    -- LP outcome alignment (uses 3-tier outcome price)
    CASE
        WHEN lp.position_direction = 'NET YES BUYER' AND COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) >= 0.90 THEN 'LP SIDED WITH WINNER'
        WHEN lp.position_direction = 'NET YES BUYER' AND COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) <= 0.10 THEN 'LP SIDED WITH LOSER'
        WHEN lp.position_direction = 'NET NO BUYER' AND COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) <= 0.10 THEN 'LP SIDED WITH WINNER'
        WHEN lp.position_direction = 'NET NO BUYER' AND COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) >= 0.90 THEN 'LP SIDED WITH LOSER'
        ELSE 'INCONCLUSIVE'
    END AS lp_outcome_alignment,

    -- Price timeline
    pt.open_price, pt.close_price,
    ROUND((pt.close_price - pt.open_price)::numeric, 4) AS price_change,
    pt.low_price, pt.high_price,
    ROUND(pt.avg_price::numeric, 4) AS market_avg_price,
    pt.total_trades, pt.total_volume,

    -- LP aggressive trade detail
    las.first_agg_time, las.last_agg_time,
    ROUND(las.lp_agg_avg_yes_price::numeric, 4) AS lp_agg_avg_yes_price,
    COALESCE(las.lp_agg_yes_qty, 0) AS lp_agg_yes_qty,
    ROUND(las.lp_agg_avg_no_price::numeric, 4) AS lp_agg_avg_no_price,
    COALESCE(las.lp_agg_no_qty, 0) AS lp_agg_no_qty,

    -- LP position
    COALESCE(lp.total_yes_qty, 0) AS lp_total_yes_qty,
    COALESCE(lp.total_no_qty, 0) AS lp_total_no_qty,
    COALESCE(lp.net_yes_position, 0) AS lp_net_yes_position,
    lp.position_direction,

    -- Price impact
    ROUND(pi.price_before_lp::numeric, 4) AS price_before_lp_aggression,
    ROUND(pi.price_after_lp::numeric, 4) AS price_after_lp_aggression,
    ROUND((pi.price_after_lp - pi.price_before_lp)::numeric, 4) AS lp_price_impact,
    CASE
        WHEN (pt.close_price - pt.open_price) = 0 THEN NULL
        ELSE ROUND(((pi.price_after_lp - pi.price_before_lp) / (pt.close_price - pt.open_price))::numeric * 100, 2)
    END AS lp_impact_pct_of_market_move,

    -- Customer comparison
    ROUND(cp.customer_avg_price_during_lp::numeric, 4) AS customer_avg_price_during_lp,
    COALESCE(cp.customer_trades_during_lp, 0) AS customer_trades_during_lp,
    ROUND((las.lp_agg_avg_yes_price - cp.customer_avg_price_during_lp)::numeric, 4) AS lp_vs_customer_price_diff,

    -- Front-running indicators
    ROUND(fr.customer_disadvantage_10s::numeric, 4) AS customer_disadvantage_10s,
    COALESCE(fr.customer_trades_within_10s, 0) AS customer_trades_within_10s,
    ROUND(fr.customer_disadvantage_30s::numeric, 4) AS customer_disadvantage_30s,
    COALESCE(fr.customer_trades_within_30s, 0) AS customer_trades_within_30s,
    ROUND(fr.customer_disadvantage_60s::numeric, 4) AS customer_disadvantage_60s,
    COALESCE(fr.customer_trades_within_60s, 0) AS customer_trades_within_60s,

    -- ======================================================================
    -- P&L: REALIZED vs UNREALIZED (using 3-tier outcome price)
    -- Realized = settled (tier 1) or expired with clear outcome (tier 2)
    -- Unrealized = not expired (tier 3)
    -- ======================================================================
    CASE WHEN (ei.settlement_price IS NOT NULL OR (ei.expiration_date < CURRENT_DATE AND (lt.last_trade_price >= 0.90 OR lt.last_trade_price <= 0.10)))
        THEN CASE WHEN lp.net_yes_position > 0
            THEN ROUND((COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END) - las.lp_agg_avg_cost) * lp.net_yes_position, 2)
            WHEN lp.net_yes_position < 0
            THEN ROUND((las.lp_agg_avg_cost - COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END)) * ABS(lp.net_yes_position), 2)
            ELSE NULL END
        ELSE NULL
    END AS realized_pnl,

    CASE WHEN ei.expiration_date >= CURRENT_DATE
        THEN CASE WHEN lp.net_yes_position > 0
            THEN ROUND((lt.last_trade_price - las.lp_agg_avg_cost) * lp.net_yes_position, 2)
            WHEN lp.net_yes_position < 0
            THEN ROUND((las.lp_agg_avg_cost - lt.last_trade_price) * ABS(lp.net_yes_position), 2)
            ELSE NULL END
        ELSE NULL
    END AS unrealized_pnl,

    -- estimated_pnl: whichever is available
    CASE WHEN lp.net_yes_position > 0
        THEN ROUND((COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) - las.lp_agg_avg_cost) * lp.net_yes_position, 2)
        WHEN lp.net_yes_position < 0
        THEN ROUND((las.lp_agg_avg_cost - COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price)) * ABS(lp.net_yes_position), 2)
        ELSE NULL
    END AS estimated_pnl,

    CASE
        WHEN ei.settlement_price IS NOT NULL THEN 'REALIZED (SETTLEMENT)'
        WHEN ei.expiration_date < CURRENT_DATE AND (lt.last_trade_price >= 0.90 OR lt.last_trade_price <= 0.10) THEN 'REALIZED (EXPIRED)'
        WHEN ei.expiration_date < CURRENT_DATE THEN 'EXPIRED - UNCERTAIN OUTCOME'
        ELSE 'UNREALIZED (NOT EXPIRED)'
    END AS pnl_type,

    -- BOD position (carried from prior 90 days)
    COALESCE(bod.bod_net_yes_position, 0) AS bod_net_yes_position,
    COALESCE(bod.bod_position_direction, 'FLAT') AS bod_position_direction,

    -- ======================================================================
    -- CLEARING FACTORS (LP Lost Money only auto-clears on REALIZED losses)
    -- ======================================================================
    CONCAT_WS(', ',
        CASE WHEN ei.days_to_expiry <= 1 THEN 'Expiration Day (' || ei.days_to_expiry || ' days)' END,
        CASE WHEN ei.days_to_expiry BETWEEN 2 AND 3 THEN 'Near Expiration (' || ei.days_to_expiry || ' days)' END,
        CASE WHEN (pi.price_after_lp - pi.price_before_lp) = 0 THEN 'No Price Impact (0.0000)' END,
        CASE WHEN ABS(pi.price_after_lp - pi.price_before_lp) > 0 AND (pt.close_price - pt.open_price) != 0 AND ABS((pi.price_after_lp - pi.price_before_lp) / (pt.close_price - pt.open_price)) < 0.25 THEN 'Low Price Impact (' || ROUND(ABS((pi.price_after_lp - pi.price_before_lp) / (pt.close_price - pt.open_price))::numeric * 100, 1) || '% of market move)' END,
        CASE WHEN COALESCE(fr.customer_disadvantage_10s, 0) <= 0 AND COALESCE(fr.customer_disadvantage_30s, 0) <= 0 AND COALESCE(fr.customer_disadvantage_60s, 0) <= 0 THEN 'No Front-Running Detected' END,
        -- Realized loss: settled or expired with clear outcome
        CASE WHEN (ei.settlement_price IS NOT NULL OR (ei.expiration_date < CURRENT_DATE AND (lt.last_trade_price >= 0.90 OR lt.last_trade_price <= 0.10)))
             AND ((lp.net_yes_position > 0 AND COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END) < las.lp_agg_avg_cost)
               OR (lp.net_yes_position < 0 AND COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END) > las.lp_agg_avg_cost))
             THEN 'LP Lost Money - Realized ($' || ABS(CASE WHEN lp.net_yes_position > 0 THEN ROUND((COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END) - las.lp_agg_avg_cost) * lp.net_yes_position, 2) WHEN lp.net_yes_position < 0 THEN ROUND((las.lp_agg_avg_cost - COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END)) * ABS(lp.net_yes_position), 2) ELSE 0 END) || ')' END,
        -- Unrealized loss: not expired, currently losing (informational, does NOT auto-clear)
        CASE WHEN ei.expiration_date >= CURRENT_DATE
             AND ((lp.net_yes_position > 0 AND lt.last_trade_price < las.lp_agg_avg_cost)
               OR (lp.net_yes_position < 0 AND lt.last_trade_price > las.lp_agg_avg_cost))
             THEN 'LP Losing - Unrealized ($' || ABS(CASE WHEN lp.net_yes_position > 0 THEN ROUND((lt.last_trade_price - las.lp_agg_avg_cost) * lp.net_yes_position, 2) WHEN lp.net_yes_position < 0 THEN ROUND((las.lp_agg_avg_cost - lt.last_trade_price) * ABS(lp.net_yes_position), 2) ELSE 0 END) || ')' END,
        CASE WHEN lp.net_yes_position = 0 THEN 'LP Position Flat' END,
        CASE WHEN (lp.position_direction = 'NET YES BUYER' AND COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) <= 0.10) OR (lp.position_direction = 'NET NO BUYER' AND COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) >= 0.90) THEN 'LP Sided With Loser' END
    ) AS clearing_factors,

    -- ======================================================================
    -- CONCERN FACTORS (distinguishes realized vs unrealized profit)
    -- ======================================================================
    CONCAT_WS(', ',
        CASE WHEN ei.days_to_expiry > 3 THEN 'Expiration ' || ei.days_to_expiry || ' Days Out' END,
        CASE WHEN (lp.position_direction = 'NET YES BUYER' AND COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) >= 0.90) OR (lp.position_direction = 'NET NO BUYER' AND COALESCE(ei.settlement_price, CASE WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price >= 0.90 THEN 1.00 WHEN ei.expiration_date < CURRENT_DATE AND lt.last_trade_price <= 0.10 THEN 0.00 ELSE NULL END, lt.last_trade_price) <= 0.10) THEN 'LP Sided With Winner' END,
        -- Realized large profit
        CASE WHEN (ei.settlement_price IS NOT NULL OR (ei.expiration_date < CURRENT_DATE AND (lt.last_trade_price >= 0.90 OR lt.last_trade_price <= 0.10)))
             AND (CASE WHEN lp.net_yes_position > 0 THEN (COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END) - las.lp_agg_avg_cost) * lp.net_yes_position WHEN lp.net_yes_position < 0 THEN (las.lp_agg_avg_cost - COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END)) * ABS(lp.net_yes_position) ELSE NULL END) > 500
             THEN 'LP Large Realized Profit ($' || ROUND((CASE WHEN lp.net_yes_position > 0 THEN (COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END) - las.lp_agg_avg_cost) * lp.net_yes_position WHEN lp.net_yes_position < 0 THEN (las.lp_agg_avg_cost - COALESCE(ei.settlement_price, CASE WHEN lt.last_trade_price >= 0.90 THEN 1.00 ELSE 0.00 END)) * ABS(lp.net_yes_position) ELSE 0 END)::numeric, 2) || ')' END,
        -- Unrealized large profit
        CASE WHEN ei.expiration_date >= CURRENT_DATE
             AND (CASE WHEN lp.net_yes_position > 0 THEN (lt.last_trade_price - las.lp_agg_avg_cost) * lp.net_yes_position WHEN lp.net_yes_position < 0 THEN (las.lp_agg_avg_cost - lt.last_trade_price) * ABS(lp.net_yes_position) ELSE NULL END) > 500
             THEN 'LP Large Unrealized Profit ($' || ROUND((CASE WHEN lp.net_yes_position > 0 THEN (lt.last_trade_price - las.lp_agg_avg_cost) * lp.net_yes_position WHEN lp.net_yes_position < 0 THEN (las.lp_agg_avg_cost - lt.last_trade_price) * ABS(lp.net_yes_position) ELSE 0 END)::numeric, 2) || ')' END,
        CASE WHEN f.lp_aggressor_pct > 0.90 THEN 'Very High Aggression (' || ROUND(f.lp_aggressor_pct * 100, 1) || '%)' END,
        CASE WHEN (pt.close_price - pt.open_price) != 0 AND ABS((pi.price_after_lp - pi.price_before_lp) / (pt.close_price - pt.open_price)) > 0.50 THEN 'LP Drove Significant Price Movement (' || ROUND(ABS((pi.price_after_lp - pi.price_before_lp) / (pt.close_price - pt.open_price))::numeric * 100, 1) || '% of day)' END,
        CASE WHEN COALESCE(fr.customer_disadvantage_10s, 0) > 0 OR COALESCE(fr.customer_disadvantage_30s, 0) > 0 OR COALESCE(fr.customer_disadvantage_60s, 0) > 0 THEN 'Front-Running Indicators Present (10s: ' || COALESCE(ROUND(fr.customer_disadvantage_10s::numeric, 4)::text, 'N/A') || ', 30s: ' || COALESCE(ROUND(fr.customer_disadvantage_30s::numeric, 4)::text, 'N/A') || ', 60s: ' || COALESCE(ROUND(fr.customer_disadvantage_60s::numeric, 4)::text, 'N/A') || ')' END,
        CASE WHEN COALESCE(las.lp_agg_avg_yes_price - cp.customer_avg_price_during_lp, 0) > 0.02 THEN 'LP Got Better Prices Than Customers (' || ROUND((las.lp_agg_avg_yes_price - cp.customer_avg_price_during_lp)::numeric, 4) || ' diff)' END
    ) AS concern_factors,

    -- ======================================================================
    -- review_status removed: replaced by Python risk scoring (calculate_risk_score)
    -- ======================================================================
    f.lp_aggressor_pct AS lp_aggressor_pct_raw,  -- passthrough for scorer
    f.lp_aggressor_count,                          -- for price pushing score
    COALESCE(lb.max_burst_60s,  0) AS max_burst_60s,
    COALESCE(lb.max_burst_300s, 0) AS max_burst_300s,
    COALESCE(lb.max_burst_600s, 0) AS max_burst_600s

FROM flagged f
LEFT JOIN category_stats cs ON cs.stat_group = f.stat_group
LEFT JOIN expiration_info ei ON ei.instrument_id = f.instrument_id
    AND ei.trading_date = f.trading_date AND ei.account = f.account
LEFT JOIN price_timeline pt ON pt.instrument_id = f.instrument_id AND pt.trading_date = f.trading_date
LEFT JOIN lp_agg_summary las ON las.instrument_id = f.instrument_id
    AND las.trading_date = f.trading_date AND las.lp_account = f.account
LEFT JOIN lp_position lp ON lp.instrument_id = f.instrument_id
    AND lp.trading_date = f.trading_date AND lp.lp_account = f.account
LEFT JOIN price_impact pi ON pi.instrument_id = f.instrument_id
    AND pi.trading_date = f.trading_date AND pi.lp_account = f.account
LEFT JOIN customer_prices cp ON cp.instrument_id = f.instrument_id
    AND cp.trading_date = f.trading_date AND cp.lp_account = f.account
LEFT JOIN front_running fr ON fr.instrument_id = f.instrument_id
    AND fr.trading_date = f.trading_date AND fr.lp_account = f.account
LEFT JOIN last_trade lt ON lt.instrument_id = f.instrument_id
LEFT JOIN bod_position bod ON bod.instrument_id = f.instrument_id
    AND bod.trading_date = f.trading_date AND bod.account = f.account
LEFT JOIN lp_burst lb ON lb.instrument_id = f.instrument_id
    AND lb.trading_date = f.trading_date AND lb.account = f.account
ORDER BY
    CASE WHEN ei.days_to_expiry <= 1 THEN 3 WHEN ei.days_to_expiry <= 3 THEN 2 ELSE 1 END,
    f.stat_group,
    f.lp_aggressor_pct DESC;
"""


# =============================================================================
# YAHOO FINANCE TICKER MAPPING
# =============================================================================
TICKER_MAP = {
    # === Stock Market Indices ===
    "FES=MAN": "^GSPC",           # US 500 (S&P 500)
    "FYM=MAN": "^DJI",            # US Industrial 30 (Dow)
    "FNQ=MAN": "^IXIC",           # US Tech 100 (Nasdaq)
    "FRTY=MAN": "^RUT",           # US Small Cap 2000 (Russell)
    "INDAX=MAN": "^GDAXI",        # DAX 40
    "INMIB=MAN": "FTSEMIB.MI",    # FTSE MIB
    "INHSI=MAN": "^HSI",          # Hang Seng
    "INNKI=MAN": "^N225",         # Nikkei
    "INCAC=MAN": "^FCHI",         # CAC 40
    "INEUR=MAN": "^STOXX50E",     # EURO STOXX 50
    "INFTS=MAN": "^FTSE",         # FTSE 100
    "INBEX=MAN": "^IBEX",         # IBEX 35
    "INBVA=MAN": "^BVSP",         # Ibovespa
    "INKOS=MAN": "^KS11",         # KOSPI
    "INSPM=MAN": "^MID",          # S&P Midcap 400
    "INSTI=MAN": "^STI",          # Straits Times

    # === Cryptocurrency ===
    "CBBTC=MAN": "BTC-USD",       # BTC Price
    "HLBTH=MAN": "BTC-USD",       # Bitcoin Highest Price
    "YXHBT=MAN": "BTC-USD",       # Bitcoin Highest Price
    "HLBTL=MAN": "BTC-USD",       # Bitcoin Lowest Price
    "YXLBT=MAN": "BTC-USD",       # Bitcoin Lowest Price
    "CTB1A=MAN": "BTC-USD",       # Bitcoin to $150k
    "CTB1B=MAN": "BTC-USD",       # Bitcoin to $125k
    "CTB1C=MAN": "BTC-USD",       # Bitcoin to $175k
    "CBETH=MAN": "ETH-USD",       # Ethereum Price
    "HLETH=MAN": "ETH-USD",       # Ethereum Highest Price
    "YXHET=MAN": "ETH-USD",       # Ethereum Highest Price
    "HLETL=MAN": "ETH-USD",       # Ethereum Lowest Price
    "YXLET=MAN": "ETH-USD",       # Ethereum Lowest Price
    "CTE2A=MAN": "ETH-USD",       # Ethereum to $5000
    "CTE2B=MAN": "ETH-USD",       # Ethereum to $5250
    "CTE2C=MAN": "ETH-USD",       # Ethereum to $5500
    "CBSOL=MAN": "SOL-USD",       # Solana Price
    "CFSOL=MAN": "SOL-USD",       # Solana Price
    "YXHSO=MAN": "SOL-USD",       # Solana Highest Price
    "YXLSO=MAN": "SOL-USD",       # Solana Lowest Price
    "CBXRP=MAN": "XRP-USD",       # XRP Price
    "CFXRP=MAN": "XRP-USD",       # XRP Price
    "YXHXR=MAN": "XRP-USD",       # XRP Highest Price
    "YXLXR=MAN": "XRP-USD",       # XRP Lowest Price
    "CBADA=MAN": "ADA-USD",       # Cardano Price
    "CFDOG=MAN": "DOGE-USD",      # Dogecoin Price
    "YXHDG=MAN": "DOGE-USD",      # Dogecoin Highest Price
    "YXLDG=MAN": "DOGE-USD",      # Dogecoin Lowest Price

    # === Foreign Exchange ===
    "USEUR=MAN": "EURUSD=X",      # EUR/USD
    "JPUSD=MAN": "USDJPY=X",      # USD/JPY
    "USJPY=MAN": "JPYUSD=X",      # JPY/USD
    "USCAD=MAN": "USDCAD=X",      # USD/CAD
    "EUJPY=MAN": "EURJPY=X",      # EUR/JPY
    "EUCHF=MAN": "EURCHF=X",      # EUR/CHF
    "USHKD=MAN": "USDHKD=X",      # USD/HKD
    "EUHKD=MAN": "EURHKD=X",      # EUR/HKD
    "USAUD=MAN": "AUDUSD=X",      # AUD/USD
    "USGBP=MAN": "GBPUSD=X",      # GBP/USD
    "USCHF=MAN": "USDCHF=X",      # USD/CHF

    # === Commodities ===
    "METLG=MAN": "GC=F",          # Gold Price
    "CSGLD=MAN": "GC=F",          # Gold Spot Price
    "METLS=MAN": "SI=F",          # Silver Price
    "CSSLV=MAN": "SI=F",          # Silver Spot Price
    "METLP=MAN": "PL=F",          # Platinum Price
    "CSPLT=MAN": "PL=F",          # Platinum Spot Price
    "METLD=MAN": "PA=F",          # Palladium Price
    "COMCO=MAN": "CL=F",          # WTI Crude Oil
    "CRUDW=MAN": "CL=F",          # WTI Crude Oil
    "CRUDB=MAN": "BZ=F",          # Brent Crude Oil
    "COMCP=MAN": "HG=F",          # Copper
    "COMHH=MAN": "NG=F",          # Henry Hub Natural Gas
    "USGP=MAN": "RB=F",           # Gasoline
}


# =============================================================================
# DYNAMIC TICKER RESOLUTION (fallback for unmapped products)
# =============================================================================
def resolve_yahoo_ticker(external_symbol, product_name, category_l2):
    """Try to resolve a Yahoo Finance ticker from the hardcoded map first,
    then fall back to keyword matching on the product name.
    Only called for non-cleared flags."""

    # 1) Hardcoded map (precise)
    if external_symbol in TICKER_MAP:
        return TICKER_MAP[external_symbol]

    # 2) Not a Financial Markets product — no underlying to pull
    if category_l2 not in (
        "Stock Market Indices", "Cryptocurrency", "Foreign Exchange",
        "Commodities", "Oil & Energy", "Corporate Performance",
        "Index Performance",
    ):
        return None

    name = (product_name or "").lower()

    # Crypto — pattern: "{CoinName} Price/Highest/Lowest" -> {SYMBOL}-USD
    crypto_keywords = {
        "bitcoin": "BTC-USD", "btc": "BTC-USD",
        "ethereum": "ETH-USD", "eth": "ETH-USD",
        "solana": "SOL-USD", "sol": "SOL-USD",
        "cardano": "ADA-USD", "ada": "ADA-USD",
        "xrp": "XRP-USD", "ripple": "XRP-USD",
        "dogecoin": "DOGE-USD", "doge": "DOGE-USD",
        "litecoin": "LTC-USD", "ltc": "LTC-USD",
        "polkadot": "DOT-USD", "dot": "DOT-USD",
        "avalanche": "AVAX-USD", "avax": "AVAX-USD",
        "chainlink": "LINK-USD", "link": "LINK-USD",
        "polygon": "MATIC-USD", "matic": "MATIC-USD",
        "shiba": "SHIB-USD",
        "tron": "TRX-USD", "trx": "TRX-USD",
        "uniswap": "UNI-USD",
        "pepe": "PEPE-USD",
    }
    for keyword, ticker in crypto_keywords.items():
        if keyword in name:
            print(f"    AUTO-MAPPED: {external_symbol} ({product_name}) -> {ticker}")
            return ticker

    # Stock indices
    index_keywords = {
        "s&p 500": "^GSPC", "us 500": "^GSPC",
        "dow": "^DJI", "industrial 30": "^DJI",
        "nasdaq": "^IXIC", "tech 100": "^IXIC",
        "russell": "^RUT", "small cap 2000": "^RUT",
        "dax": "^GDAXI", "ftse mib": "FTSEMIB.MI",
        "ftse 100": "^FTSE", "hang seng": "^HSI",
        "nikkei": "^N225", "cac 40": "^FCHI",
        "euro stoxx": "^STOXX50E", "ibex": "^IBEX",
        "ibovespa": "^BVSP", "kospi": "^KS11",
        "midcap 400": "^MID", "straits times": "^STI",
    }
    for keyword, ticker in index_keywords.items():
        if keyword in name:
            print(f"    AUTO-MAPPED: {external_symbol} ({product_name}) -> {ticker}")
            return ticker

    # Commodities
    commodity_keywords = {
        "gold": "GC=F", "silver": "SI=F", "platinum": "PL=F",
        "palladium": "PA=F", "copper": "HG=F",
        "wti crude": "CL=F", "brent crude": "BZ=F",
        "natural gas": "NG=F", "gasoline": "RB=F",
        "crude oil": "CL=F",
    }
    for keyword, ticker in commodity_keywords.items():
        if keyword in name:
            print(f"    AUTO-MAPPED: {external_symbol} ({product_name}) -> {ticker}")
            return ticker

    # FX
    fx_keywords = {
        "euro to us dollar": "EURUSD=X", "us dollar to japanese yen": "USDJPY=X",
        "canadian dollar to us dollar": "USDCAD=X",
        "pound sterling to us dollar": "GBPUSD=X",
        "australian dollar to us dollar": "AUDUSD=X",
        "swiss franc to us dollar": "USDCHF=X",
    }
    for keyword, ticker in fx_keywords.items():
        if keyword in name:
            print(f"    AUTO-MAPPED: {external_symbol} ({product_name}) -> {ticker}")
            return ticker

    # No match found
    print(f"    WARNING: No Yahoo ticker mapping for {external_symbol} ({product_name}). Add to TICKER_MAP if needed.")
    return None


# =============================================================================
# TRADE DETAIL QUERY
# =============================================================================
def build_trade_detail_query(instrument_id, trading_date, lp_accounts):
    """Pull all trades for an instrument-day with flag reason tagging."""
    accts = ", ".join(f"'{a}'" for a in lp_accounts)
    return f"""
    WITH lp_agg_trades AS (
        SELECT pair_time, yes_price, quantity, aggressor_symbol_subtype AS lp_side
        FROM pairs
        WHERE instrument_id = '{instrument_id}'
          AND trading_date = '{trading_date}'
          AND aggressor_account IN ({accts})
    ),
    lp_agg_window AS (
        SELECT MIN(pair_time) AS first_agg, MAX(pair_time) AS last_agg
        FROM lp_agg_trades
    )
    SELECT
        p.pair_time,
        p.instrument_id,
        p.trading_date,
        p.aggressor_account,
        p.passive_account,
        p.quantity,
        p.yes_price,
        p.aggressor_symbol_subtype AS aggressor_side,
        p.passive_symbol_subtype AS passive_side,
        -- Flag reasons
        CONCAT_WS(', ',
            CASE WHEN p.aggressor_account IN ({accts}) THEN 'LP_AGGRESSIVE' END,
            CASE WHEN p.passive_account IN ({accts}) AND p.aggressor_account NOT IN ({accts}) THEN 'LP_PASSIVE' END,
            CASE WHEN p.aggressor_account NOT IN ({accts}) AND p.passive_account NOT IN ({accts}) THEN 'CUSTOMER_ONLY' END,
            -- Front-running: customer trade within 10/30/60s after any LP aggressive trade, customer disadvantaged
            CASE WHEN p.aggressor_account NOT IN ({accts}) AND p.passive_account NOT IN ({accts})
                 AND EXISTS (
                    SELECT 1 FROM lp_agg_trades lat
                    WHERE p.pair_time > lat.pair_time
                      AND p.pair_time <= lat.pair_time + INTERVAL '10 seconds'
                      AND CASE WHEN lat.lp_side = 'YES' THEN p.yes_price > lat.yes_price
                               WHEN lat.lp_side = 'NO' THEN p.yes_price < lat.yes_price
                               ELSE FALSE END
                 ) THEN 'FRONT_RUN_10S' END,
            CASE WHEN p.aggressor_account NOT IN ({accts}) AND p.passive_account NOT IN ({accts})
                 AND EXISTS (
                    SELECT 1 FROM lp_agg_trades lat
                    WHERE p.pair_time > lat.pair_time
                      AND p.pair_time <= lat.pair_time + INTERVAL '30 seconds'
                      AND p.pair_time > lat.pair_time + INTERVAL '10 seconds'
                      AND CASE WHEN lat.lp_side = 'YES' THEN p.yes_price > lat.yes_price
                               WHEN lat.lp_side = 'NO' THEN p.yes_price < lat.yes_price
                               ELSE FALSE END
                 ) THEN 'FRONT_RUN_30S' END,
            CASE WHEN p.aggressor_account NOT IN ({accts}) AND p.passive_account NOT IN ({accts})
                 AND EXISTS (
                    SELECT 1 FROM lp_agg_trades lat
                    WHERE p.pair_time > lat.pair_time
                      AND p.pair_time <= lat.pair_time + INTERVAL '60 seconds'
                      AND p.pair_time > lat.pair_time + INTERVAL '30 seconds'
                      AND CASE WHEN lat.lp_side = 'YES' THEN p.yes_price > lat.yes_price
                               WHEN lat.lp_side = 'NO' THEN p.yes_price < lat.yes_price
                               ELSE FALSE END
                 ) THEN 'FRONT_RUN_60S' END,
            -- Price context
            CASE WHEN p.pair_time < (SELECT first_agg FROM lp_agg_window)
                 THEN 'BEFORE_LP_AGGRESSION' END,
            CASE WHEN p.pair_time > (SELECT last_agg FROM lp_agg_window)
                 THEN 'AFTER_LP_AGGRESSION' END
        ) AS flag_reasons
    FROM pairs p
    WHERE p.instrument_id = '{instrument_id}'
      AND p.trading_date = '{trading_date}'
    ORDER BY p.pair_time;
    """


# =============================================================================
# YAHOO FINANCE UNDERLYING DATA
# =============================================================================
# Tickers that Yahoo returns in UTC (crypto, FX, some commodities).
# Everything else is assumed US/Eastern (NYSE/NASDAQ exchange time).
YAHOO_UTC_TICKERS = {
    "BTC-USD", "ETH-USD", "SOL-USD", "XRP-USD", "ADA-USD",
    "DOGE-USD", "LTC-USD", "DOT-USD", "AVAX-USD", "LINK-USD",
    "MATIC-USD", "SHIB-USD", "TRX-USD", "UNI-USD", "PEPE-USD",
    "EURUSD=X", "USDJPY=X", "JPYUSD=X", "USDCAD=X", "EURJPY=X",
    "EURCHF=X", "USDHKD=X", "EURHKD=X", "AUDUSD=X", "GBPUSD=X",
    "USDCHF=X",
    "GC=F", "SI=F", "PL=F", "PA=F", "CL=F", "BZ=F",
    "HG=F", "NG=F", "RB=F",
}


def fetch_underlying_data(external_symbol, trading_date, interval="15m",
                          return_utc=False):
    """Fetch intraday data from Yahoo Finance for a given date.

    Parameters
    ----------
    external_symbol : str
    trading_date    : str or date  (YYYY-MM-DD)
    interval        : str  '15m' for monthly (60-day window), '60m' for backfill
    return_utc      : bool
        If True, return timestamps as UTC-aware (for comparison with pair_time).
        If False, strip timezone for Excel export.

    Returns
    -------
    DataFrame with columns: timestamp, ticker, Open, High, Low, Close, Volume
    or None on failure.
    """
    ticker = resolve_yahoo_ticker(external_symbol, None, None) if external_symbol else None
    if not ticker:
        ticker = TICKER_MAP.get(external_symbol)
    if not ticker:
        return None

    try:
        import yfinance as yf
    except ImportError:
        print("    yfinance not installed. Skipping underlying data.")
        return None

    try:
        dt = pd.Timestamp(trading_date)
        start = dt.strftime("%Y-%m-%d")
        end = (dt + pd.Timedelta(days=1)).strftime("%Y-%m-%d")

        tk = yf.Ticker(ticker)
        hist = tk.history(start=start, end=end, interval=interval)

        if hist.empty:
            hist = tk.history(start=start, end=end, interval="1d")
            if hist.empty:
                return None

        hist = hist.reset_index()
        if "Datetime" in hist.columns:
            hist = hist.rename(columns={"Datetime": "timestamp"})
        elif "Date" in hist.columns:
            hist = hist.rename(columns={"Date": "timestamp"})

        hist["ticker"] = ticker
        hist["external_symbol"] = external_symbol

        # ── Timezone normalization ──────────────────────────────────────────
        # Yahoo returns crypto/FX/commodities in UTC, equities in US/Eastern.
        # We normalize everything to UTC-aware for comparison with pair_time,
        # then optionally strip for Excel export.
        ts = hist["timestamp"]
        if hasattr(ts.dtype, "tz") and ts.dtype.tz is not None:
            # Already tz-aware — convert to UTC
            hist["timestamp"] = ts.dt.tz_convert("UTC")
        else:
            # Naive — localize based on ticker type
            local_tz = "UTC" if ticker in YAHOO_UTC_TICKERS else "America/New_York"
            hist["timestamp"] = ts.dt.tz_localize(local_tz, ambiguous="infer",
                                                   nonexistent="shift_forward")
            hist["timestamp"] = hist["timestamp"].dt.tz_convert("UTC")

        if not return_utc:
            # Strip for Excel
            hist["timestamp"] = hist["timestamp"].dt.tz_localize(None)

        return hist[["timestamp", "ticker", "Open", "High", "Low", "Close", "Volume"]]

    except Exception as e:
        print(f"    Yahoo Finance error for {ticker}: {e}")
        return None


# =============================================================================
# GET EXTERNAL SYMBOL FOR INSTRUMENT
# =============================================================================
def get_product_info(conn, instrument_id):
    """Look up external_symbol and product name for an instrument."""
    query = f"""
    SELECT pr.external_symbol, pr.name, pr.category, pr.category_l2
    FROM instrument_definitions id
    JOIN products pr ON pr.product_id = id.product_id
    WHERE id.instrument_id = '{instrument_id}'
    LIMIT 1;
    """
    result = pd.read_sql(query, conn)
    if result.empty:
        return None, None, None, None
    row = result.iloc[0]
    return row.get("external_symbol"), row.get("name"), row.get("category"), row.get("category_l2")



# =============================================================================
# RISK SCORING ENGINE
# =============================================================================
# Scoring framework (0-100):
#   Position & Outcome  : max 25  ($1 per $100 profit, realized or unrealized)
#   Front-Running       : max 25
#   Price Impact        : max 20  (reduced by underlying asset movement)
#   Aggression Pattern  : max 20  (bidirectional - can reduce score)
#   Total cap           : 100
#
# Tiers: 0-25 Low | 26-50 Moderate | 51-75 Elevated | 76-100 High
# =============================================================================

SCORE_WEIGHTS = {
    # 1. P&L / Position & Outcome (max 25) — logarithmic, liquidity-adjusted
    # Log base 50,001: $2,500 profit in thin market ~17pts, $10K ~21pts, $25K+ ~25pts.
    # Raises the cap threshold so mid-range profits score meaningfully.
    "pnl_log_scale":         25,
    "pnl_log_base":          50001,   # was 2501 — recalibrated for better spread
    "pnl_liquidity_divisor": 10000,
    "pnl_cap":               25,

    # 2. Front-Running (max 25) — escalating multipliers
    # SQL columns remapped: 10s col = 1s window, 30s col = 10s window, 60s col = 20s window
    "fr_1s_base":            8,
    "fr_10s_base":           5,
    "fr_20s_base":           3,
    "fr_mult_3":             2.0,
    "fr_mult_5":             3.0,
    "fr_mult_10":            25,
    "fr_cap":                25,

    # 3. Price Impact (max 20)
    # Max path: >50% move (+13) + underlying opposite (+7) = 20. Verified.
    "impact_gt50pct":        13,   # was 12
    "impact_25_50pct":       6,
    "underlying_same_dir":   -8,
    "underlying_opp_dir":    7,    # was 5 — bumped so max path reaches 20
    "impact_cap":            20,

    # 4. Aggression Pattern (max 20) — losses handled in P&L, not here
    "aggression_gt90":       10,
    "aggression_75_90":      5,
    "expiry_gt7d":           5,
    "expiry_1_3d":           -5,
    # BOD position: scores only when LP has a large existing position that
    # benefits directionally from today's aggressive trading.
    # Minimum 20,000 contracts to qualify — small residual positions excluded.
    "bod_min_position":      20000,  # abs(bod_net_yes_position) threshold
    "bod_large_position":    50000,  # >= 50K → full 5 pts
    "bod_aligned_large":     5,      # >= 50K contracts aligned with winner
    "bod_aligned_mid":       3,      # 20K–50K contracts aligned with winner
    "aggression_cap":        20,

    # 5. Price Pushing (max 10) — two context-aware sub-signals
    #
    # Sub-signal A — Aggression Density (max 5):
    #   Scores based on volume tier AND density ratio.
    #   Five volume tiers: very thin / thin / mid / liquid / very liquid.
    "push_vol_very_thin":       500,    # total_volume < 500    → very thin
    "push_vol_thin":           2000,    # total_volume < 2K     → thin
    "push_vol_liquid":        10000,    # total_volume > 10K    → liquid
    "push_vol_very_liquid":   50000,    # total_volume > 50K    → very liquid
    #   Points by tier × density band (hi/mid/low):
    "push_density_vthin_full":   5,     # very thin + density >= hi
    "push_density_vthin_half":   4,     # very thin + density >= mid
    "push_density_thin_full":    4,     # thin      + density >= hi
    "push_density_thin_half":    3,     # thin      + density >= mid
    "push_density_mid_full":     3,     # mid       + density >= hi
    "push_density_mid_half":     2,     # mid       + density >= mid
    "push_density_liq_full":     2,     # liquid    + density >= hi
    "push_density_liq_half":     1,     # liquid    + density >= mid
    "push_density_vliq_full":    1,     # very liq  + density >= hi
    "push_density_vliq_half":    0,     # very liq  + density >= mid → no score
    "push_density_any_low":      1,     # any tier, density >= low_thresh (except very liquid)
    "push_density_hi_thresh":    3.0,   # trades per 1K contracts
    "push_density_mid_thresh":   1.0,
    "push_density_low_thresh":   0.5,
    #
    # Sub-signal B — Burst Cluster (max 5):
    #   Uses max_burst_60s / max_burst_300s / max_burst_600s from SQL.
    #   Higher burst in shorter window = more concern.
    #   Thresholds: number of LP aggressive trades within the window.
    "push_burst_60s_hi":        5,      # >= 5 trades in 60s  → 5 pts
    "push_burst_60s_mid":       3,      # >= 3 trades in 60s  → 3 pts
    "push_burst_60s_lo":        2,      # >= 2 trades in 60s  → 1 pt
    "push_burst_300s_hi":       7,      # >= 7 trades in 300s → 3 pts
    "push_burst_300s_mid":      4,      # >= 4 trades in 300s → 2 pts
    "push_burst_600s_hi":      10,      # >= 10 trades in 600s → 2 pts
    "push_burst_600s_mid":      6,      # >= 6 trades in 600s → 1 pt
    "push_burst_cap":           5,      # sub-signal B capped at 5
    #
    # Auto-max: density sub-score == 5 AND burst sub-score >= 4
    "push_automax_density_pts": 5,
    "push_automax_burst_pts":   4,
    "push_cap":                10,
}


def _fr_multiplier(instances, w):
    """Return escalating multiplier. None = auto-max (10+ instances)."""
    if instances >= 10:
        return None
    elif instances >= 5:
        return w["fr_mult_5"]
    elif instances >= 3:
        return w["fr_mult_3"]
    return 1.0


def _pnl_log_score(profit, total_volume, w):
    """Logarithmic P&L score adjusted for market liquidity.

    Log base is pnl_log_base (default 50,001):
      ~$2,500 profit in thin market  → ~17 pts
      ~$10,000 profit in thin market → ~21 pts
      ~$25,000+ profit               → 25 pts (cap)
    Higher base = more spread across the mid range.
    """
    import math
    if profit is None or profit != profit or profit <= 0:  # None / NaN / negative
        return 0
    log_base = w.get("pnl_log_base", 50001)
    base = math.log(profit + 1) / math.log(log_base)
    volume_factor = 1.0
    if total_volume and total_volume > w["pnl_liquidity_divisor"]:
        volume_factor = max(0.5, 1.0 - math.log(total_volume / w["pnl_liquidity_divisor"]) * 0.1)
    return min(int(base * volume_factor * w["pnl_log_scale"]), w["pnl_cap"])


def calculate_risk_score(row, underlying_direction=None):
    """
    Composite risk score (0-100) across 5 categories.
    Auto-flag if any category hits its cap.
    """
    w            = SCORE_WEIGHTS
    detail_parts = []
    auto_flags   = []

    # 1. P&L / POSITION & OUTCOME (max 25)
    def _safe_float(val):
        """Convert to float, returning 0.0 for None/NaN/empty."""
        try:
            f = float(val)
            return 0.0 if f != f else f  # NaN check: NaN != NaN
        except (TypeError, ValueError):
            return 0.0

    realized_pnl   = _safe_float(row.get("realized_pnl"))
    unrealized_pnl = _safe_float(row.get("unrealized_pnl"))
    estimated_pnl  = _safe_float(row.get("estimated_pnl"))
    total_volume   = _safe_float(row.get("total_volume"))

    if realized_pnl > 0:
        pnl_used, pnl_type_used = realized_pnl,   "REALIZED"
    elif unrealized_pnl > 0:
        pnl_used, pnl_type_used = unrealized_pnl, "UNREALIZED"
    elif estimated_pnl > 0:
        pnl_used, pnl_type_used = estimated_pnl,  "ESTIMATED"
    else:
        pnl_used, pnl_type_used = 0.0, "NONE"

    loss_adj = 0
    if realized_pnl < 0:
        loss_adj = max(int(realized_pnl / 100), -10)
    elif unrealized_pnl < 0:
        loss_adj = max(int(unrealized_pnl / 200), -5)

    score_outcome = max(_pnl_log_score(pnl_used, total_volume, w) + loss_adj, 0)
    if score_outcome >= w["pnl_cap"]:
        auto_flags.append(f"P&L cap hit (score={score_outcome})")
    detail_parts.append(
        f"P&L: +{score_outcome} (${pnl_used:,.0f} {pnl_type_used}"
        + (f", loss_adj={loss_adj}" if loss_adj < 0 else "") + ")"
        if pnl_used > 0 else
        f"P&L: +{score_outcome}" + (f" (loss_adj={loss_adj})" if loss_adj < 0 else " (no profit)")
    )

    # 2. FRONT-RUNNING (max 25, escalating multipliers)
    # SQL columns remapped to Graham's windows:
    #   customer_trades_within_10s = 1s window
    #   customer_trades_within_30s = 10s window
    #   customer_trades_within_60s = 20s window
    cd_1s  = float(row.get("customer_disadvantage_10s") or 0)
    cd_10s = float(row.get("customer_disadvantage_30s") or 0)
    cd_20s = float(row.get("customer_disadvantage_60s") or 0)
    ct_1s  = int(row.get("customer_trades_within_10s") or 0)
    ct_10s = int(row.get("customer_trades_within_30s") or 0)
    ct_20s = int(row.get("customer_trades_within_60s") or 0)

    score_fr = 0
    fr_parts = []

    def _score_window(cd, ct, base_pts, label):
        nonlocal score_fr
        if cd <= 0 or ct == 0:
            return
        mult = _fr_multiplier(ct, w)
        if mult is None:
            pts = w["fr_cap"]
            auto_flags.append(f"FR 10+ instances in {label} window (auto-max)")
        else:
            pts = min(int(base_pts * mult), w["fr_cap"])
        score_fr = min(score_fr + pts, w["fr_cap"])
        fr_parts.append(f"FR@{label}:{ct}x{'' if mult is None else f'{mult}x'}(+{pts})")

    _score_window(cd_1s,  ct_1s,  w["fr_1s_base"],  "1s")
    _score_window(cd_10s, ct_10s, w["fr_10s_base"], "10s")
    _score_window(cd_20s, ct_20s, w["fr_20s_base"], "20s")

    score_frontrun = min(score_fr, w["fr_cap"])
    if score_frontrun >= w["fr_cap"] and not any("FR" in f for f in auto_flags):
        auto_flags.append(f"Front-Running cap hit (score={score_frontrun})")
    detail_parts.append(
        f"Front-Run: +{score_frontrun}" + (f" ({', '.join(fr_parts)})" if fr_parts else " (none)")
    )

    # 3. PRICE IMPACT (max 20)
    lp_impact_pct = float(row.get("lp_impact_pct_of_market_move") or 0)
    score_imp, imp_parts = 0, []

    if abs(lp_impact_pct) > 50:
        score_imp += w["impact_gt50pct"]
        imp_parts.append(f">50%move(+{w['impact_gt50pct']})")
    elif abs(lp_impact_pct) > 25:
        score_imp += w["impact_25_50pct"]
        imp_parts.append(f"25-50%move(+{w['impact_25_50pct']})")

    if underlying_direction == "SAME":
        score_imp += w["underlying_same_dir"]
        imp_parts.append(f"UnderlySame({w['underlying_same_dir']})")
    elif underlying_direction == "OPPOSITE":
        score_imp += w["underlying_opp_dir"]
        imp_parts.append(f"UnderlyOpp(+{w['underlying_opp_dir']})")

    score_impact = max(min(score_imp, w["impact_cap"]), 0)
    if score_impact >= w["impact_cap"]:
        auto_flags.append(f"Price Impact cap hit (score={score_impact})")
    detail_parts.append(
        f"Impact: +{score_impact}" + (f" ({', '.join(imp_parts)})" if imp_parts else " (no significant impact)")
    )

    # 4. AGGRESSION PATTERN (max 20) — losses handled in P&L above
    agg_pct     = float(row.get("lp_aggressor_pct") or row.get("lp_aggressor_pct_raw") or 0)
    # lp_aggressor_pct is already *100 from SQL — compare against 90/75 not 0.90/0.75
    days_to_exp = row.get("days_to_expiry")
    bod_net     = float(row.get("bod_net_yes_position") or 0)
    bod_dir     = row.get("bod_position_direction", "FLAT")
    outcome_pr  = row.get("outcome_price")

    score_agg, agg_parts = 0, []

    if agg_pct > 90:
        score_agg += w["aggression_gt90"]
        agg_parts.append(f"Agg>90%(+{w['aggression_gt90']})")
    elif agg_pct > 75:
        score_agg += w["aggression_75_90"]
        agg_parts.append(f"Agg75-90%(+{w['aggression_75_90']})")

    if days_to_exp is not None:
        days_to_exp = int(days_to_exp)
        if days_to_exp > 7:
            score_agg += w["expiry_gt7d"]
            agg_parts.append(f"Exp>7d(+{w['expiry_gt7d']})")
        elif 1 <= days_to_exp <= 3:
            score_agg += w["expiry_1_3d"]
            agg_parts.append(f"Exp1-3d({w['expiry_1_3d']})")

    if bod_net != 0 and outcome_pr is not None:
        outcome_pr  = float(outcome_pr)
        abs_bod     = abs(bod_net)
        # Only score if LP has a meaningful existing position (>= 20K contracts).
        # The aggression must benefit that position directionally.
        if abs_bod >= w["bod_min_position"]:
            if ((bod_dir == "NET YES BUYER" and outcome_pr >= 0.90) or
                    (bod_dir == "NET NO BUYER"  and outcome_pr <= 0.10)):
                if abs_bod >= w["bod_large_position"]:
                    score_agg += w["bod_aligned_large"]
                    agg_parts.append(f"BODBenefits>=50K(+{w['bod_aligned_large']})")
                else:
                    score_agg += w["bod_aligned_mid"]
                    agg_parts.append(f"BODBenefits20-50K(+{w['bod_aligned_mid']})")

    score_aggression = max(min(score_agg, w["aggression_cap"]), 0)
    if score_aggression >= w["aggression_cap"]:
        auto_flags.append(f"Aggression Pattern cap hit (score={score_aggression})")
    detail_parts.append(
        f"Aggression: +{score_aggression}" + (f" ({', '.join(agg_parts)})" if agg_parts else " (none)")
    )

    # 5. PRICE PUSHING (max 10) — density in volume-tiered markets + real burst detection
    #
    # Sub-signal A: Aggression Density (max 5)
    #   Volume tier determines sensitivity: same density is more concerning
    #   in a thin market than a liquid one.
    #
    # Sub-signal B: Burst Cluster (max 5)
    #   Uses max_burst_60s / max_burst_300s / max_burst_600s from SQL CTE.
    #   Tighter window with more trades = stronger burst signal.
    #
    agg_count      = int(row.get("lp_aggressor_count") or 0)
    total_vol      = _safe_float(row.get("total_volume"))
    burst_60s      = int(row.get("max_burst_60s")  or 0)
    burst_300s     = int(row.get("max_burst_300s") or 0)
    burst_600s     = int(row.get("max_burst_600s") or 0)

    score_push, push_parts = 0, []

    # --- Sub-signal A: Aggression Density (volume-tiered) ---
    density = 0.0
    density_pts = 0
    if total_vol > 0 and agg_count > 0:
        density = (agg_count / total_vol) * 1000

        # Determine volume tier (five bands)
        if total_vol < w["push_vol_very_thin"]:
            tier_full  = w["push_density_vthin_full"]
            tier_half  = w["push_density_vthin_half"]
            tier_label = "very_thin"
            low_eligible = True
        elif total_vol < w["push_vol_thin"]:
            tier_full  = w["push_density_thin_full"]
            tier_half  = w["push_density_thin_half"]
            tier_label = "thin"
            low_eligible = True
        elif total_vol <= w["push_vol_liquid"]:
            tier_full  = w["push_density_mid_full"]
            tier_half  = w["push_density_mid_half"]
            tier_label = "mid"
            low_eligible = True
        elif total_vol <= w["push_vol_very_liquid"]:
            tier_full  = w["push_density_liq_full"]
            tier_half  = w["push_density_liq_half"]
            tier_label = "liquid"
            low_eligible = True
        else:
            tier_full  = w["push_density_vliq_full"]
            tier_half  = w["push_density_vliq_half"]
            tier_label = "very_liquid"
            low_eligible = False  # very liquid markets: low density = no score

        if density >= w["push_density_hi_thresh"]:
            density_pts = tier_full
        elif density >= w["push_density_mid_thresh"]:
            density_pts = tier_half
        elif density >= w["push_density_low_thresh"] and low_eligible:
            density_pts = w["push_density_any_low"]

        if density_pts > 0:
            push_parts.append(f"Density:{density:.2f}/1K@{tier_label}(+{density_pts})")

    score_push += density_pts

    # --- Sub-signal B: Burst Cluster ---
    burst_pts = 0
    burst_detail = []

    # 60s window (highest weight — tightest burst)
    if burst_60s >= w["push_burst_60s_hi"]:
        burst_pts = max(burst_pts, 5)
        burst_detail.append(f"60s:{burst_60s}trades(+5)")
    elif burst_60s >= w["push_burst_60s_mid"]:
        burst_pts = max(burst_pts, 3)
        burst_detail.append(f"60s:{burst_60s}trades(+3)")
    elif burst_60s >= w["push_burst_60s_lo"]:
        burst_pts = max(burst_pts, 1)
        burst_detail.append(f"60s:{burst_60s}trades(+1)")

    # 300s window (medium)
    if burst_300s >= w["push_burst_300s_hi"]:
        burst_pts = max(burst_pts, 3)
        burst_detail.append(f"300s:{burst_300s}trades(+3)")
    elif burst_300s >= w["push_burst_300s_mid"]:
        burst_pts = max(burst_pts, 2)
        burst_detail.append(f"300s:{burst_300s}trades(+2)")

    # 600s window (looser — only scores if no tighter burst already maxed)
    if burst_600s >= w["push_burst_600s_hi"]:
        burst_pts = max(burst_pts, 2)
        burst_detail.append(f"600s:{burst_600s}trades(+2)")
    elif burst_600s >= w["push_burst_600s_mid"]:
        burst_pts = max(burst_pts, 1)
        burst_detail.append(f"600s:{burst_600s}trades(+1)")

    burst_pts = min(burst_pts, w["push_burst_cap"])
    if burst_detail:
        push_parts.append(f"Burst({', '.join(burst_detail)})")
    score_push += burst_pts

    # Auto-max: density at full thin-tier points AND burst near-max
    if density_pts >= w["push_automax_density_pts"] and burst_pts >= w["push_automax_burst_pts"]:
        score_push = w["push_cap"]
        auto_flags.append(
            f"Price Pushing auto-max (density={density:.2f}/1K @{tier_label}, "
            f"burst_60s={burst_60s}, burst_300s={burst_300s})"
        )
        push_parts.append("AUTO-MAX")

    score_push = min(score_push, w["push_cap"])
    detail_parts.append(
        f"Pushing: +{score_push}" + (f" ({', '.join(push_parts)})" if push_parts else " (none)")
    )

    # COMPOSITE
    raw_total       = score_outcome + score_frontrun + score_impact + score_aggression + score_push
    risk_score      = min(raw_total, 100)
    auto_flagged    = len(auto_flags) > 0
    auto_flag_reason = "; ".join(auto_flags) if auto_flags else ""

    if auto_flagged and risk_score < 51:
        risk_score = 51

    if   risk_score <= 25: risk_tier = "Low"
    elif risk_score <= 50: risk_tier = "Moderate"
    elif risk_score <= 75: risk_tier = "Elevated"
    else:                  risk_tier = "High"

    score_detail = " | ".join(detail_parts) + f" | TOTAL: {risk_score} ({risk_tier})"
    if auto_flagged:
        score_detail += f" | AUTO-FLAG: {auto_flag_reason}"

    return {
        "risk_score":        risk_score,
        "risk_tier":         risk_tier,
        "score_outcome":     score_outcome,
        "score_frontrun":    score_frontrun,
        "score_impact":      score_impact,
        "score_aggression":  score_aggression,
        "score_push":        score_push,
        "auto_flagged":      auto_flagged,
        "auto_flag_reason":  auto_flag_reason,
        "pnl_used":          pnl_used,
        "pnl_type_used":     pnl_type_used,
        "score_detail":      score_detail,
    }


def apply_risk_scores(df, underlying_directions=None):
    """
    Apply calculate_risk_score() to every row in the dataframe.
    Drops any existing score columns first to prevent duplicate columns
    when called a second time for re-scoring.
    """
    if underlying_directions is None:
        underlying_directions = {}

    score_cols = [
        "risk_score", "risk_tier", "score_outcome", "score_frontrun",
        "score_impact", "score_aggression", "score_push",
        "auto_flagged", "auto_flag_reason",
        "pnl_used", "pnl_type_used", "score_detail",
    ]
    df = df.drop(columns=[c for c in score_cols if c in df.columns], errors="ignore")
    df = df.reset_index(drop=True)

    results = []
    for _, row in df.iterrows():
        key = (row.get("instrument_id"), str(row.get("trading_date", ""))[:10])
        ul_dir = underlying_directions.get(key)
        results.append(calculate_risk_score(row, underlying_direction=ul_dir))

    score_df = pd.DataFrame(results)
    return pd.concat([df, score_df], axis=1)


# =============================================================================
# UNDERLYING ASSET DIRECTION ANALYSIS
# =============================================================================

def build_underlying_directions(df, conn, interval="15m"):
    """
    For each flagged row that has a mapped Yahoo Finance ticker, fetch
    intraday price data and determine whether the underlying asset moved
    in the SAME or OPPOSITE direction as the LP's aggressive trades
    during the LP's aggression window.

    Comparison logic:
      - Find the Yahoo candle closest to first_agg_time (UTC)
      - Find the Yahoo candle closest to last_agg_time (UTC)
      - Underlying direction = sign(close_at_last - close_at_first)
      - LP direction = NET YES BUYER → positive / NET NO BUYER → negative
      - If signs match  → SAME     (underlying explains LP move, -8 pts)
      - If signs differ → OPPOSITE (LP going against market,    +5 pts)
      - If no data      → None     (no adjustment applied)

    Returns
    -------
    dict keyed by (instrument_id, trading_date_str) → 'SAME' | 'OPPOSITE' | None
    """
    import warnings
    directions = {}

    # Only process rows that have a resolvable ticker and aggression timestamps
    needs_ul = df[
        df["first_agg_time"].notna() &
        df["last_agg_time"].notna() &
        df["position_direction"].notna()
    ].copy()

    if needs_ul.empty:
        return directions

    # Fetch product info once per unique instrument to avoid redundant DB calls
    unique_instruments = needs_ul["instrument_id"].unique()
    product_info = {}
    for inst_id in unique_instruments:
        ext_sym, prod_name, cat, cat2 = get_product_info(conn, inst_id)
        ticker = resolve_yahoo_ticker(ext_sym, prod_name, cat2) if ext_sym else None
        product_info[inst_id] = {"ticker": ticker, "ext_sym": ext_sym,
                                  "prod_name": prod_name, "cat2": cat2}

    # Cache Yahoo data per (ticker, trading_date) to avoid re-fetching
    yahoo_cache = {}

    for _, row in needs_ul.iterrows():
        inst_id    = row["instrument_id"]
        tdate      = str(row["trading_date"])[:10]
        key        = (inst_id, tdate)
        pinfo      = product_info.get(inst_id, {})
        ticker     = pinfo.get("ticker")

        if not ticker:
            directions[key] = None
            continue

        cache_key = (ticker, tdate)
        if cache_key not in yahoo_cache:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                ul = fetch_underlying_data(
                    pinfo["ext_sym"], tdate,
                    interval=interval, return_utc=True
                )
            yahoo_cache[cache_key] = ul
        else:
            ul = yahoo_cache[cache_key]

        if ul is None or ul.empty:
            directions[key] = None
            continue

        # Ensure Yahoo timestamps are UTC-aware for comparison
        if ul["timestamp"].dt.tz is None:
            ul = ul.copy()
            ul["timestamp"] = ul["timestamp"].dt.tz_localize("UTC")

        # Ensure pair_time values are UTC-aware
        first_agg = pd.Timestamp(row["first_agg_time"])
        last_agg  = pd.Timestamp(row["last_agg_time"])
        if first_agg.tzinfo is None:
            first_agg = first_agg.tz_localize("UTC")
        else:
            first_agg = first_agg.tz_convert("UTC")
        if last_agg.tzinfo is None:
            last_agg = last_agg.tz_localize("UTC")
        else:
            last_agg = last_agg.tz_convert("UTC")

        # Find closest candle to first and last aggression times
        def closest_close(ts):
            diffs = (ul["timestamp"] - ts).abs()
            idx   = diffs.idxmin()
            # Only use if within 20 minutes — otherwise data gap is too large
            if diffs[idx] > pd.Timedelta(minutes=20):
                return None
            return float(ul.loc[idx, "Close"])

        price_at_first = closest_close(first_agg)
        price_at_last  = closest_close(last_agg)

        if price_at_first is None or price_at_last is None:
            directions[key] = None
            continue

        underlying_move = price_at_last - price_at_first
        lp_direction    = row.get("position_direction", "FLAT")

        # LP YES buyer = positive direction, NO buyer = negative direction
        if abs(underlying_move) < 1e-8:
            # Underlying flat — no directional signal
            directions[key] = None
        elif underlying_move > 0 and lp_direction == "NET YES BUYER":
            directions[key] = "SAME"
        elif underlying_move < 0 and lp_direction == "NET NO BUYER":
            directions[key] = "SAME"
        else:
            directions[key] = "OPPOSITE"

    same_count = sum(1 for v in directions.values() if v == "SAME")
    opp_count  = sum(1 for v in directions.values() if v == "OPPOSITE")
    none_count = sum(1 for v in directions.values() if v is None)
    print(f"  Underlying direction analysis: {int(same_count)} SAME, "
          f"{int(opp_count)} OPPOSITE, {int(none_count)} no data")

    return directions


# =============================================================================
# STRIP TIMEZONES HELPER
# =============================================================================
def strip_tz(df):
    """Strip timezone info from all datetime columns."""
    for col in df.select_dtypes(include=["datetimetz"]).columns:
        df[col] = df[col].dt.tz_convert("UTC").dt.tz_localize(None)
    return df


# =============================================================================
# EXCEL EXPORT (enhanced with trade detail and underlying tabs)
# =============================================================================
def export_to_excel(df, output_path, month_label, conn=None, yf_interval="15m", underlying_directions=None):
    """
    Two-pass Excel export:
      Pass 1 — pandas writes all standard tabs (Summary, All Flags, tier tabs)
      Pass 2 — openpyxl reopens the file and appends consolidated alert tabs

    Keeping the passes separate avoids the openpyxl/pandas writer conflict
    that caused the corrupted file (missing [Content_Types].xml).
    """
    df = strip_tz(df)
    if underlying_directions is None:
        underlying_directions = {}

    high_df     = df[df["risk_tier"] == "High"]
    elevated_df = df[df["risk_tier"] == "Elevated"]
    moderate_df = df[df["risk_tier"] == "Moderate"]
    low_df      = df[df["risk_tier"] == "Low"]

    # ==================================================================
    # PASS 1: pandas writes all standard tabs
    # ==================================================================
    summary_data = {
        "Metric": [
            "Report Period", "Total Flags",
            "High Risk (76-100)", "Elevated Risk (51-75)",
            "Moderate Risk (26-50)", "Low Risk (0-25)", "Avg Risk Score",
        ],
        "Value": [
            month_label, len(df),
            len(high_df), len(elevated_df), len(moderate_df), len(low_df),
            f"{df['risk_score'].mean():.1f}" if len(df) > 0 else "N/A",
        ],
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

        df_sorted = df.sort_values("risk_score", ascending=False)
        df_sorted.to_excel(writer, sheet_name="All Flags", index=False)

        if not high_df.empty:
            high_df.sort_values("risk_score", ascending=False).to_excel(
                writer, sheet_name="High Risk", index=False)
        if not elevated_df.empty:
            elevated_df.sort_values("risk_score", ascending=False).to_excel(
                writer, sheet_name="Elevated Risk", index=False)
        if not moderate_df.empty:
            moderate_df.sort_values("risk_score", ascending=False).to_excel(
                writer, sheet_name="Moderate Risk", index=False)
        if not low_df.empty:
            low_df.sort_values("risk_score", ascending=False).to_excel(
                writer, sheet_name="Low Risk", index=False)

        # Auto-fit columns on standard tabs
        for sheet_name in ["Summary", "All Flags", "High Risk", "Elevated Risk",
                            "Moderate Risk", "Low Risk"]:
            if sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                for col in ws.columns:
                    max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    print(f"  Pass 1 complete: standard tabs written.")

    # ==================================================================
    # PASS 2: openpyxl reopens file and appends consolidated alert tabs
    # ==================================================================
    non_cleared = df.sort_values("risk_score", ascending=False)

    if non_cleared.empty or conn is None:
        print(f"  No consolidated alert tabs to generate.")
        print(f"  Excel saved: {output_path}")
        return

    print(f"  Pass 2: generating consolidated alert tabs for all {len(non_cleared)} flags...")

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(output_path)

    TIER_COLORS = {
        "High":     "C00000",
        "Elevated": "ED7D31",
        "Moderate": "B8860B",
        "Low":      "70AD47",
    }

    def hdr(ws, row, col, value, bg="1F4E79", fg="FFFFFF", bold=True, size=10):
        c = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
        c.font = Font(name="Arial", bold=bold, color=fg, size=size)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        return c

    def dat(ws, row, col, value, bold=False, bg=None):
        # Sanitize value — openpyxl can't handle numpy types or NaT
        if value is None or (hasattr(value, '__class__') and 'NaT' in str(type(value))):
            value = ""
        elif hasattr(value, 'item'):  # numpy scalar
            value = value.item()
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        return c

    for _, row in non_cleared.iterrows():
        inst_id    = row["instrument_id"]
        tdate      = str(row["trading_date"])[:10]
        short_date = tdate.replace("-", "")
        score_val  = int(row.get("risk_score", 0) or 0)
        tab_name   = f"{short_date}_{inst_id}_S{score_val:02d}"[:31]
        tier       = row.get("risk_tier", "")
        tc         = TIER_COLORS.get(tier, "808080")

        try:
            ws = wb.create_sheet(title=tab_name)

            # ── Column widths ─────────────────────────────────────────────
            col_widths = {"A": 32, "B": 20, "C": 20, "D": 18, "E": 18,
                          "F": 18, "G": 16, "H": 16, "I": 16, "J": 16}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            r = 1

            # ==============================================================
            # SECTION 1 — ALERT HEADER
            # ==============================================================
            hdr(ws, r, 1, "ALERT SUMMARY", bg="1F4E79", size=12)
            ws.merge_cells(f"A{r}:J{r}")
            r += 1

            header_fields = [
                ("Instrument",      inst_id),
                ("Trading Date",    tdate),
                ("Account",         row.get("account", "")),
                ("Category",        f"{row.get('category','')} / {row.get('category_l2','')}"),
                ("Risk Score",      row.get("risk_score", "")),
                ("Risk Tier",       tier),
                ("Aggression %",    f"{float(row.get('lp_aggressor_pct', 0) or 0):.1f}%"),
                ("Days to Expiry",  row.get("days_to_expiry", "")),
                ("Expiration Date", str(row.get("expiration_date", ""))[:10]),
                ("Settlement Price",row.get("settlement_price", "N/A")),
                ("BOD Position",    f"{row.get('bod_position_direction','')} ({float(row.get('bod_net_yes_position', 0) or 0):+.0f} net YES)"),
                ("P&L",             f"${float(row.get('estimated_pnl') or 0):,.2f} ({row.get('pnl_type','')})"),
                ("Outcome Price",   f"{row.get('outcome_price','')} ({row.get('outcome_price_source','')})"),
                ("LP Alignment",    row.get("lp_outcome_alignment", "")),
                ("Likely Outcome",  row.get("likely_outcome", "")),
                ("Stat Group",      row.get("stat_group", "")),
                ("Group Mean %",    f"{float(row.get('group_mean_pct') or 0):.2f}%"),
                ("Group Threshold %",f"{float(row.get('group_threshold_pct') or 0):.2f}%"),
                ("Group Product-Days", row.get("group_product_days", "")),
            ]
            # Two-column layout for header: label col A, value col B-E | label col F, value col G-J
            left  = header_fields[:len(header_fields)//2 + len(header_fields)%2]
            right = header_fields[len(header_fields)//2 + len(header_fields)%2:]
            max_rows = max(len(left), len(right))
            for i in range(max_rows):
                if i < len(left):
                    label, val = left[i]
                    hdr(ws, r, 1, label, bg="D9E1F2", fg="000000", bold=True, size=10)
                    c = dat(ws, r, 2, val)
                    ws.merge_cells(f"B{r}:E{r}")
                    if label == "Risk Tier":
                        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                        c.fill = PatternFill("solid", fgColor=tc)
                    elif label == "Risk Score":
                        c.font = Font(name="Arial", bold=True, size=11)
                if i < len(right):
                    label, val = right[i]
                    hdr(ws, r, 6, label, bg="D9E1F2", fg="000000", bold=True, size=10)
                    c2 = dat(ws, r, 7, val)
                    ws.merge_cells(f"G{r}:J{r}")
                r += 1
            r += 1  # spacer

            # ==============================================================
            # SECTION 2 — RISK SCORE BREAKDOWN (factor-level detail)
            # ==============================================================
            hdr(ws, r, 1, "RISK SCORE BREAKDOWN — FACTOR DETAIL", bg="2E75B6", size=11)
            ws.merge_cells(f"A{r}:J{r}")
            r += 1

            score_headers = ["Category", "Factor", "Raw Value", "Threshold", "Points Applied", "Fired?"]
            sh_widths = [1, 2, 3, 4, 5, 6]
            for ci, h in enumerate(score_headers, 1):
                hdr(ws, r, ci, h, bg="BDD7EE", fg="000000", bold=True, size=10)
            ws.merge_cells(f"F{r}:J{r}")
            r += 1

            def score_factor_row(ws, r, category, factor, raw, threshold, pts, fired, cat_bg="F2F2F2"):
                fired_str = "✓" if fired else "—"
                fired_bg  = "E2EFDA" if fired and pts > 0 else ("FCE4D6" if fired and pts < 0 else cat_bg)
                dat(ws, r, 1, category, bold=True, bg=cat_bg)
                dat(ws, r, 2, factor,   bg=cat_bg)
                dat(ws, r, 3, raw,      bg=cat_bg)
                dat(ws, r, 4, threshold,bg=cat_bg)
                pts_c = dat(ws, r, 5, pts, bold=True, bg=cat_bg)
                if pts > 0:
                    pts_c.font = Font(name="Arial", bold=True, color="375623", size=10)
                elif pts < 0:
                    pts_c.font = Font(name="Arial", bold=True, color="C00000", size=10)
                fired_c = dat(ws, r, 6, fired_str, bg=fired_bg)
                ws.merge_cells(f"F{r}:J{r}")
                return r + 1

            # Helper to safely get float
            def sf(key, default=0.0):
                v = row.get(key)
                try: return float(v) if v is not None else default
                except: return default

            pnl_used    = sf("pnl_used")
            agg_pct     = sf("lp_aggressor_pct")  # already multiplied by 100 in SQL
            days_exp    = row.get("days_to_expiry")
            net_pos     = sf("lp_net_yes_position")
            realized_pnl = sf("realized_pnl")
            cd_10s      = sf("customer_disadvantage_10s")
            cd_30s      = sf("customer_disadvantage_30s")
            cd_60s      = sf("customer_disadvantage_60s")
            ct_10s      = int(sf("customer_trades_within_10s"))
            ct_30s      = int(sf("customer_trades_within_30s"))
            ct_60s      = int(sf("customer_trades_within_60s"))
            lp_vs_cust  = sf("lp_vs_customer_price_diff")
            impact_pct  = sf("lp_impact_pct_of_market_move")
            bod_dir     = row.get("bod_position_direction", "FLAT")
            outcome_pr  = sf("outcome_price")
            ul_dir      = underlying_directions.get((inst_id, tdate))
            score_o     = int(row.get("score_outcome", 0) or 0)
            score_fr    = int(row.get("score_frontrun", 0) or 0)
            score_imp   = int(row.get("score_impact", 0) or 0)
            score_agg   = int(row.get("score_aggression", 0) or 0)
            risk_score  = int(row.get("risk_score", 0) or 0)

            # ── POSITION & OUTCOME ────────────────────────────────────────
            cat_bg = "EBF3FB"
            pts_profit = min(int(pnl_used / 100), 25)
            r = score_factor_row(ws, r, "Position & Outcome",
                f"Profit (${pnl_used:,.0f} {row.get('pnl_type_used','')})",
                f"${pnl_used:,.2f}", "$100/pt → max 25", pts_profit, pnl_used > 0, cat_bg)

            # subtotal row
            hdr(ws, r, 1, "Position & Outcome Subtotal", bg="BDD7EE", fg="000000", bold=True, size=10)
            dat(ws, r, 2, "", bg="BDD7EE")
            dat(ws, r, 3, "", bg="BDD7EE")
            dat(ws, r, 4, "", bg="BDD7EE")
            sc = dat(ws, r, 5, score_o, bold=True, bg="BDD7EE")
            sc.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
            dat(ws, r, 6, f"/ 25", bg="BDD7EE")
            ws.merge_cells(f"F{r}:J{r}")
            r += 1
            r += 1  # spacer

            # ── FRONT-RUNNING ─────────────────────────────────────────────
            cat_bg = "FFF2CC"
            fr_10s_fired = cd_10s > 0 and ct_10s > 0
            fr_30s_fired = cd_30s > 0 and ct_30s > 0 and not fr_10s_fired
            fr_60s_fired = cd_60s > 0 and ct_60s > 0 and not fr_30s_fired and not fr_10s_fired
            better_fired = lp_vs_cust > 0.02

            r = score_factor_row(ws, r, "Front-Running",
                "Customer disadvantage within 1s (SQL: 10s column)",
                f"{cd_10s:.4f} avg ({ct_10s} trades)", "> 0 disadvantage",
                15 if fr_10s_fired else 0, fr_10s_fired, cat_bg)
            r = score_factor_row(ws, r, "",
                "Customer disadvantage within 10s (SQL: 30s column)",
                f"{cd_30s:.4f} avg ({ct_30s} trades)", "> 0 disadvantage",
                10 if fr_30s_fired else 0, fr_30s_fired, cat_bg)
            r = score_factor_row(ws, r, "",
                "Customer disadvantage within 20s (SQL: 60s column)",
                f"{cd_60s:.4f} avg ({ct_60s} trades)", "> 0 disadvantage",
                6 if fr_60s_fired else 0, fr_60s_fired, cat_bg)
            r = score_factor_row(ws, r, "",
                "LP obtained better prices than customers",
                f"{lp_vs_cust:.4f} diff", "> 0.02 diff",
                8 if better_fired else 0, better_fired, cat_bg)

            hdr(ws, r, 1, "Front-Running Subtotal", bg="FFE699", fg="000000", bold=True, size=10)
            dat(ws, r, 2, "", bg="FFE699"); dat(ws, r, 3, "", bg="FFE699"); dat(ws, r, 4, "", bg="FFE699")
            sc = dat(ws, r, 5, score_fr, bold=True, bg="FFE699")
            sc.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
            dat(ws, r, 6, "/ 25", bg="FFE699"); ws.merge_cells(f"F{r}:J{r}")
            r += 1
            r += 1

            # ── PRICE IMPACT ──────────────────────────────────────────────
            cat_bg = "FCE4D6"
            imp_gt50  = abs(impact_pct) > 50
            imp_25_50 = 25 < abs(impact_pct) <= 50
            ul_same   = ul_dir == "SAME"
            ul_opp    = ul_dir == "OPPOSITE"

            r = score_factor_row(ws, r, "Price Impact",
                "LP drove > 50% of market price move",
                f"{impact_pct:.1f}%", "> 50%",
                10 if imp_gt50 else 0, imp_gt50, cat_bg)
            r = score_factor_row(ws, r, "",
                "LP drove 25-50% of market price move",
                f"{impact_pct:.1f}%", "25-50%",
                5 if imp_25_50 else 0, imp_25_50, cat_bg)
            r = score_factor_row(ws, r, "",
                "Underlying moved SAME direction as LP",
                ul_dir or "No data", "SAME → discount",
                -8 if ul_same else 0, ul_same, cat_bg)
            r = score_factor_row(ws, r, "",
                "LP traded OPPOSITE to underlying direction",
                ul_dir or "No data", "OPPOSITE → amplify",
                5 if ul_opp else 0, ul_opp, cat_bg)

            hdr(ws, r, 1, "Price Impact Subtotal", bg="F4B183", fg="000000", bold=True, size=10)
            dat(ws, r, 2, "", bg="F4B183"); dat(ws, r, 3, "", bg="F4B183"); dat(ws, r, 4, "", bg="F4B183")
            sc = dat(ws, r, 5, score_imp, bold=True, bg="F4B183")
            sc.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
            dat(ws, r, 6, "/ 20", bg="F4B183"); ws.merge_cells(f"F{r}:J{r}")
            r += 1
            r += 1

            # ── AGGRESSION PATTERN ────────────────────────────────────────
            cat_bg = "E2EFDA"
            agg_gt90    = agg_pct > 90
            agg_75_90   = 75 < agg_pct <= 90
            exp_gt7     = days_exp is not None and int(days_exp) > 7
            exp_1_3     = days_exp is not None and 1 <= int(days_exp) <= 3
            pos_flat    = net_pos == 0
            rloss       = realized_pnl < 0
            bod_winner  = (
                (bod_dir == "NET YES BUYER" and outcome_pr >= 0.90) or
                (bod_dir == "NET NO BUYER"  and outcome_pr <= 0.10)
            ) if sf("bod_net_yes_position") != 0 else False

            r = score_factor_row(ws, r, "Aggression Pattern",
                "LP aggression rate > 90%",
                f"{agg_pct:.1f}%", "> 90%",
                10 if agg_gt90 else 0, agg_gt90, cat_bg)
            r = score_factor_row(ws, r, "",
                "LP aggression rate 75-90%",
                f"{agg_pct:.1f}%", "75-90%",
                5 if agg_75_90 else 0, agg_75_90, cat_bg)
            r = score_factor_row(ws, r, "",
                "Contract expiry > 7 days out",
                f"{days_exp}d" if days_exp is not None else "N/A", "> 7 days",
                5 if exp_gt7 else 0, exp_gt7, cat_bg)
            r = score_factor_row(ws, r, "",
                "Contract expiry 1-3 days (natural convergence)",
                f"{days_exp}d" if days_exp is not None else "N/A", "1-3 days",
                -5 if exp_1_3 else 0, exp_1_3, cat_bg)
            r = score_factor_row(ws, r, "",
                "LP net position flat at end of day",
                f"{net_pos:+.0f} net YES", "= 0",
                -5 if pos_flat else 0, pos_flat, cat_bg)
            r = score_factor_row(ws, r, "",
                "BOD position already aligned with winner",
                f"{bod_dir} | outcome={outcome_pr:.2f}", "Pre-aligned",
                5 if bod_winner else 0, bod_winner, cat_bg)
            r = score_factor_row(ws, r, "",
                "Realized loss confirmed",
                f"${realized_pnl:,.2f}", "< 0",
                -10 if rloss else 0, rloss, cat_bg)

            hdr(ws, r, 1, "Aggression Pattern Subtotal", bg="A9D18E", fg="000000", bold=True, size=10)
            dat(ws, r, 2, "", bg="A9D18E"); dat(ws, r, 3, "", bg="A9D18E"); dat(ws, r, 4, "", bg="A9D18E")
            sc = dat(ws, r, 5, score_agg, bold=True, bg="A9D18E")
            sc.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
            dat(ws, r, 6, "/ 20", bg="A9D18E"); ws.merge_cells(f"F{r}:J{r}")
            r += 1
            r += 1

            # ── COMPOSITE TOTAL ───────────────────────────────────────────
            hdr(ws, r, 1, "COMPOSITE RISK SCORE", bg=tc, fg="FFFFFF", bold=True, size=12)
            ws.merge_cells(f"A{r}:D{r}")
            sc = dat(ws, r, 5, risk_score, bold=True)
            sc.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
            sc.fill = PatternFill("solid", fgColor=tc)
            sc.alignment = Alignment(horizontal="center", vertical="center")
            tier_c = dat(ws, r, 6, f"{tier} ({risk_score}/100)", bold=True)
            tier_c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            tier_c.fill = PatternFill("solid", fgColor=tc)
            ws.merge_cells(f"F{r}:J{r}")
            ws.row_dimensions[r].height = 24
            r += 2

            # ==============================================================
            # SECTION 3 — MARKET CONTEXT
            # ==============================================================
            hdr(ws, r, 1, "MARKET CONTEXT", bg="404040", fg="FFFFFF", size=11)
            ws.merge_cells(f"A{r}:J{r}")
            r += 1

            context_left = [
                ("Open Price",              f"{sf('open_price'):.4f}"),
                ("Close Price",             f"{sf('close_price'):.4f}"),
                ("Price Change",            f"{sf('price_change'):+.4f}"),
                ("Low Price",               f"{sf('low_price'):.4f}"),
                ("High Price",              f"{sf('high_price'):.4f}"),
                ("Market Avg Price",        f"{sf('market_avg_price'):.4f}"),
                ("Total Trades (all)",      str(int(sf('total_trades')))),
                ("Total Volume (all)",      f"{int(sf('total_volume')):,}"),
                ("First LP Aggr Time",      str(row.get("first_agg_time", ""))[:19]),
                ("Last LP Aggr Time",       str(row.get("last_agg_time", ""))[:19]),
            ]
            context_right = [
                ("LP Agg Avg YES Price",    f"{sf('lp_agg_avg_yes_price'):.4f}"),
                ("LP Agg YES Qty",          f"{int(sf('lp_agg_yes_qty')):,}"),
                ("LP Agg Avg NO Price",     f"{sf('lp_agg_avg_no_price'):.4f}"),
                ("LP Agg NO Qty",           f"{int(sf('lp_agg_no_qty')):,}"),
                ("LP Total YES Qty",        f"{int(sf('lp_total_yes_qty')):,}"),
                ("LP Total NO Qty",         f"{int(sf('lp_total_no_qty')):,}"),
                ("LP Net YES Position",     f"{sf('lp_net_yes_position'):+.0f}"),
                ("Price Before LP Aggr",    f"{sf('price_before_lp_aggression'):.4f}"),
                ("Price After LP Aggr",     f"{sf('price_after_lp_aggression'):.4f}"),
                ("LP Price Impact",         f"{sf('lp_price_impact'):+.4f}"),
            ]

            for i in range(max(len(context_left), len(context_right))):
                if i < len(context_left):
                    hdr(ws, r, 1, context_left[i][0], bg="D9D9D9", fg="000000", bold=True, size=10)
                    dat(ws, r, 2, context_left[i][1])
                    ws.merge_cells(f"B{r}:E{r}")
                if i < len(context_right):
                    hdr(ws, r, 6, context_right[i][0], bg="D9D9D9", fg="000000", bold=True, size=10)
                    dat(ws, r, 7, context_right[i][1])
                    ws.merge_cells(f"G{r}:J{r}")
                r += 1

            # Customer comparison row
            r += 1
            hdr(ws, r, 1, "Customer Avg Price During LP Window", bg="D9D9D9", fg="000000", bold=True, size=10)
            dat(ws, r, 2, f"{sf('customer_avg_price_during_lp'):.4f}  ({int(sf('customer_trades_during_lp'))} trades)")
            ws.merge_cells(f"B{r}:E{r}")
            hdr(ws, r, 6, "LP vs Customer Price Diff", bg="D9D9D9", fg="000000", bold=True, size=10)
            c_diff = dat(ws, r, 7, f"{lp_vs_cust:+.4f}")
            if lp_vs_cust > 0.02:
                c_diff.font = Font(name="Arial", bold=True, color="C00000", size=10)
            ws.merge_cells(f"G{r}:J{r}")
            r += 2

            # ==============================================================
            # SECTION 4 — TRADE DETAIL
            # ==============================================================
            hdr(ws, r, 1, "TRADE DETAIL  (LP aggressive rows highlighted)", bg="375623", fg="FFFFFF", size=11)
            ws.merge_cells(f"A{r}:J{r}")
            r += 1

            try:
                detail_query = build_trade_detail_query(inst_id, tdate, LP_ACCOUNTS)
                trades_df = pd.read_sql(detail_query, conn)
                trades_df = strip_tz(trades_df)

                if not trades_df.empty:
                    # Widen tab for trade columns — trades can have 9+ cols
                    for ci, col_name in enumerate(trades_df.columns, 1):
                        hdr(ws, r, ci, col_name, bg="548235", fg="FFFFFF", size=9)
                        # Auto-width hint
                        ws.column_dimensions[
                            ws.cell(r, ci).column_letter
                        ].width = max(len(str(col_name)) + 2, 14)
                    r += 1
                    for _, trow in trades_df.iterrows():
                        is_lp_agg = "LP_AGGRESSIVE" in str(trow.get("flag_reasons", ""))
                        is_fr     = any(x in str(trow.get("flag_reasons", ""))
                                        for x in ["FRONT_RUN_10S", "FRONT_RUN_30S", "FRONT_RUN_60S"])
                        row_bg = "FCE4D6" if is_lp_agg else ("FFF2CC" if is_fr else None)
                        for ci, col_name in enumerate(trades_df.columns, 1):
                            dat(ws, r, ci, trow[col_name], bg=row_bg)
                        r += 1
                print(f"    {tab_name}: {len(trades_df)} trades")
            except Exception as e:
                dat(ws, r, 1, f"Error loading trade detail: {e}")
                r += 1
            r += 1

            # ==============================================================
            # SECTION 5 — UNDERLYING ASSET DATA (if applicable)
            # ==============================================================
            try:
                ext_sym, prod_name, cat, cat2 = get_product_info(conn, inst_id)
                yahoo_ticker = resolve_yahoo_ticker(ext_sym, prod_name, cat2) if ext_sym else None
                if yahoo_ticker:
                    ul_data = fetch_underlying_data(
                        ext_sym, tdate, interval=yf_interval, return_utc=False)
                    if ul_data is not None and not ul_data.empty:
                        dir_label = ul_dir or "No data"
                        dir_color = {"SAME": "70AD47", "OPPOSITE": "C00000"}.get(ul_dir, "808080")
                        hdr(ws, r, 1,
                            f"UNDERLYING ASSET: {yahoo_ticker} ({prod_name})",
                            bg="7030A0", fg="FFFFFF", size=11)
                        ws.merge_cells(f"A{r}:E{r}")
                        hdr(ws, r, 6,
                            f"Direction vs LP Trades: {dir_label}",
                            bg=dir_color, fg="FFFFFF", size=11)
                        ws.merge_cells(f"F{r}:J{r}")
                        r += 1
                        # Context note
                        note = (
                            "Underlying moved in the SAME direction as LP — supports market-following behavior, reduces price impact concern."
                            if ul_dir == "SAME" else
                            "Underlying moved OPPOSITE to LP direction — LP was trading against the market, amplifies concern."
                            if ul_dir == "OPPOSITE" else
                            "No underlying direction data available for this flag."
                        )
                        dat(ws, r, 1, note)
                        ws.merge_cells(f"A{r}:J{r}")
                        r += 1
                        # Column headers
                        for ci, col_name in enumerate(ul_data.columns, 1):
                            hdr(ws, r, ci, col_name, bg="9B59B6", fg="FFFFFF", size=9)
                        r += 1
                        for _, urow in ul_data.iterrows():
                            for ci, col_name in enumerate(ul_data.columns, 1):
                                dat(ws, r, ci, urow[col_name])
                            r += 1
                        print(f"    {tab_name}: +{len(ul_data)} underlying rows ({yahoo_ticker})")
                    else:
                        hdr(ws, r, 1,
                            f"UNDERLYING ASSET: {yahoo_ticker} ({prod_name}) — No intraday data available for {tdate}",
                            bg="808080", fg="FFFFFF", size=10)
                        ws.merge_cells(f"A{r}:J{r}")
                        r += 1
                else:
                    hdr(ws, r, 1,
                        "UNDERLYING ASSET: Not applicable (non-financial product — weather, elections, economic indicators, etc.)",
                        bg="808080", fg="FFFFFF", size=10)
                    ws.merge_cells(f"A{r}:J{r}")
                    r += 1
            except Exception as e:
                dat(ws, r, 1, f"Underlying data error: {e}")
                r += 1

        except Exception as e:
            print(f"    ERROR on tab {tab_name}: {e}")
            import traceback; traceback.print_exc()

    wb.save(output_path)
    wb.close()
    print(f"  Pass 2 complete.")
    print(f"  Excel saved: {output_path}")


def main():
    start_date, end_date = get_prior_month_range()
    month_label = datetime.strptime(start_date, "%Y-%m-%d").strftime("%B %Y")

    print(f"=" * 60)
    print(f"LP Watchlist & Enhanced Monitoring")
    print(f"Report Period: {month_label} ({start_date} to {end_date})")
    print(f"=" * 60)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"\nConnecting to database...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        print(f"  Connected successfully.")
    except Exception as e:
        print(f"  ERROR: Could not connect to database: {e}")
        sys.exit(1)

    print(f"\nRunning enhanced monitoring query for {month_label}...")
    print(f"  (This may take a few minutes depending on trade volume)")
    try:
        query = build_query(start_date, end_date)
        df = pd.read_sql(query, conn)
        print(f"  Query complete. {len(df)} flags returned.")
    except Exception as e:
        print(f"  ERROR: Query failed: {e}")
        conn.close()
        sys.exit(1)

    print(f"\nCalculating risk scores...")
    df = apply_risk_scores(df)
    print(f"  Initial scoring complete. Score range: {df['risk_score'].min()}-{df['risk_score'].max()} "
          f"(avg {df['risk_score'].mean():.1f})")

    underlying_directions = {}
    print(f"\nFetching underlying asset data from Yahoo Finance...")
    try:
        underlying_directions = build_underlying_directions(df, conn)
        if any(v is not None for v in underlying_directions.values()):
            print(f"  Re-scoring with underlying direction adjustments...")
            df = apply_risk_scores(df, underlying_directions=underlying_directions)
            print(f"  Final scoring complete. Score range: {df['risk_score'].min()}-{df['risk_score'].max()} "
                  f"(avg {df['risk_score'].mean():.1f})")
        else:
            print(f"  No underlying data available — scores unchanged.")
    except Exception as e:
        print(f"  WARNING: Underlying direction analysis failed: {type(e).__name__}: {e}")
        print(f"  Proceeding with scores calculated without underlying adjustment.")
        underlying_directions = {}

    if df.empty:
        print(f"\nNo flags generated for {month_label}. Nothing to export.")
        conn.close()
        return

    filename = f"LP_Watchlist_{datetime.strptime(start_date, '%Y-%m-%d').strftime('%Y_%m')}.xlsx"
    output_path = OUTPUT_DIR / filename

    print(f"\nExporting to Excel...")
    print(f"  Dataframe shape: {df.shape}")
    print(f"  Columns: {list(df.columns)}")
    print(f"  Dtypes with issues:")
    for col in df.columns:
        try:
            sample = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
            print(f"    {col}: {df[col].dtype} (sample: {repr(sample)[:60]})")
        except Exception as e:
            print(f"    {col}: ERROR - {e}")

    # Sanitize dataframe before writing — convert problematic types
    import numpy as np
    # Use df.dtypes (Series) not df[col].dtype to avoid issues with duplicate col names
    for col in df.columns.unique():
        try:
            if df[col].dtype == object:
                df[col] = df[col].where(df[col].notna(), None)
        except Exception:
            pass  # duplicate col edge case — apply_risk_scores drop should have fixed this

    try:
        export_to_excel(df, output_path, month_label, conn=conn, underlying_directions=underlying_directions)
    except Exception as e:
        import traceback
        print(f"\n  ERROR during Excel export:")
        traceback.print_exc()
        print(f"\n  Saving CSV fallback instead...")
        csv_path = output_path.with_suffix(".csv")
        df.to_csv(csv_path, index=False)
        print(f"  CSV saved: {csv_path}")

    conn.close()
    print(f"  Database connection closed.")

    print(f"\n{'=' * 60}")
    print(f"SUMMARY - {month_label}")
    print(f"{'=' * 60}")
    print(f"  Total flags:      {len(df)}")
    print(f"  High   (76-100):  {len(df[df['risk_tier'] == 'High'])}")
    print(f"  Elevated (51-75): {len(df[df['risk_tier'] == 'Elevated'])}")
    print(f"  Moderate (26-50): {len(df[df['risk_tier'] == 'Moderate'])}")
    print(f"  Low      (0-25):  {len(df[df['risk_tier'] == 'Low'])}")
    print(f"  Avg risk score:   {df['risk_score'].mean():.1f}")
    print(f"\n  Report saved to: {output_path}")


if __name__ == "__main__":
    main()
