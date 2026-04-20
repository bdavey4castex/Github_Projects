"""
=============================================================================
FORECASTEX INVALID LARGE TRADER REPORT (ILTR) PROCESSOR
=============================================================================
Scans an inbox folder for EP3 ILTR exports (CSV, Excel, TXT), classifies
each alert row, and writes ONE output Excel file per source file with the
original columns plus LTR File Date, Side Date, Instrument Ticker,
EP3 Matched Instrument Name, Match Notes, and Adjudication appended.

USAGE:
    python iltr_processor.py

    Drop an EP3 export into the Inbox folder, then run.
    Processed source files are renamed with "_completed" suffix.

FOLDERS:
    Inbox:   ~/Documents/ForecastEx Compliance/ILTR/Inbox/
    Reports: ~/Documents/ForecastEx Compliance/ILTR/Reports/

ADJUDICATION VALUES:
    Instrument Found in Database       — no_match, instrument exists in DB
    Instrument NOT Found in Database   — no_match, instrument missing from DB
    Known Bug - Multiple Match         — multiple instruments found, known symbol
    New Known Bug - Multiple Match     — multiple instruments found, unknown symbol
    Unknown Alert Type - Flagged       — unrecognized error message
=============================================================================
"""

import os
import re
import sys
import logging
import traceback
from datetime import datetime
from pathlib import Path

import psycopg2
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Windows UTF-8 console safety
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


# =============================================================================
# CONFIGURATION
# =============================================================================

DB_CONFIG = {
    "host":     "web-prod-db.cla246ai8gve.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "database": "postgres",
    "user":     "postgres_read_only",
    "password": os.environ.get("FXDB_PASSWORD", "pgreadonlypw"),
}

# Known bug symbols — multiple instruments found is a known system issue
KNOWN_BUG_SYMBOLS = {"DISSN", "USCE", "GCE", "GSL"}

# Folder layout — always uses local Documents folder.
def _find_documents() -> Path:
    return Path.home() / "Documents"

BASE_DIR      = _find_documents() / "ForecastEx Compliance" / "ILTR"
INBOX_DIR     = BASE_DIR / "Inbox"
REPORTS_DIR   = BASE_DIR / "Reports"
REFERENCE_DIR = BASE_DIR / "Reference"

INBOX_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
REFERENCE_DIR.mkdir(parents=True, exist_ok=True)

# Required columns to validate an ILTR file
REQUIRED_COLUMNS = {"alertName", "alertTime", "description", "id", "status"}

# Recognized error types
ERROR_NO_MATCH = "no matching instruments found"
ERROR_MULTI    = "multiple instruments found"

# Adjudication text — short and clean
ADJ = {
    "in_db":           "Instrument Found in Database",
    "in_db_alias":     "Instrument Found in Database",          # alias match = valid exact match
    "in_db_period":    "Underlying Found - Date Mismatch",      # period-end / date window — not exact
    "in_db_date":      "Underlying Found - Date Mismatch",      # product + strike found, date differs
    "in_db_strike":    "Underlying Found - Strike Mismatch",    # product found, strike not matched
    "not_in_db":       "Instrument NOT Found in Database",
    "known_bug":       "Known Bug - Multiple Match",
    "new_bug":         "New Known Bug - Multiple Match",
    "unknown":         "Unknown Alert Type - Flagged",
}


# =============================================================================
# LOGGING
# =============================================================================

LOG_FILE = BASE_DIR / "iltr_processor.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# =============================================================================
# EP3 INSTRUMENT REFERENCE
# =============================================================================

def load_ep3_reference() -> dict:
    """
    Load the EP3 instruments CSV from the Reference/ folder.
    Returns a dict mapping instrument_id (DB) -> EP3 symbol (id column).
    Since DB instrument_id == EP3 id, this also serves as a validation table.

    If no reference file is found, returns empty dict and logs a warning.
    Place the instruments export CSV in:
        ~/Documents/ForecastEx Compliance/ILTR/Reference/
    """
    candidates = [
        p for p in REFERENCE_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in (".csv", ".xlsx", ".xls")
    ] if REFERENCE_DIR.exists() else []

    if not candidates:
        log.warning(
            f"No EP3 reference file found in {REFERENCE_DIR}. "
            "EP3 Instrument Name column will use DB instrument_id directly. "
            "To enable reference lookup, place the instruments export CSV in the Reference folder."
        )
        return {}

    # Use the most recently modified file if multiple present
    ref_path = max(candidates, key=lambda p: p.stat().st_mtime)
    log.info(f"Loading EP3 reference: {ref_path.name}")

    try:
        if ref_path.suffix.lower() == ".csv":
            df = pd.read_csv(ref_path, low_memory=False, usecols=["id", "productId"])
        else:
            df = pd.read_excel(ref_path, usecols=["id", "productId"])

        # Build lookup: instrument_id -> EP3 symbol
        # EP3 id == DB instrument_id, so this validates and confirms the symbol
        ref = dict(zip(df["id"].astype(str), df["id"].astype(str)))
        log.info(f"  Loaded {len(ref):,} EP3 instrument symbols from reference")
        return ref

    except Exception as e:
        log.warning(f"Failed to load EP3 reference file: {e}")
        return {}


# =============================================================================
# DATABASE
# =============================================================================

def get_engine():
    url = (
        f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
        f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
    )
    return create_engine(url)


def lookup_instrument(engine, product_id: str, time_specifier: str,
                      strike_value: str, side_date: str = "",
                      instrument_raw_date: str = "") -> tuple[str, str | None]:
    """
    Tiered lookup — tries progressively looser matches before giving up.

    Returns (tier, instrument_id | None):
        "in_db"        — exact match (product + date + strike)
        "in_db_period" — matched via period-end date or instrument_raw date window (±7 days)
        "in_db_alias"  — matched via cftc_strike_alias (ordinal) — exact date or any date
        "in_db_date"   — product + strike found, date differs
        "in_db_strike" — product exists but strike/alias not matched anywhere
        "not_in_db"    — product not in DB at all
    """
    try:
        # ── Tier 1: exact match (product + date + strike_value) ──
        df = pd.read_sql("""
            SELECT instrument_id FROM public.instrument_definitions
            WHERE product_id     = %(product_id)s
              AND time_specifier = %(time_specifier)s
              AND strike_value   = %(strike_value)s
            LIMIT 1
        """, engine, params={
            "product_id":     product_id,
            "time_specifier": time_specifier,
            "strike_value":   strike_value,
        })
        if not df.empty:
            return "in_db", df["instrument_id"].iloc[0], "Exact match"

        # ── Tier 2: period-end date candidates (annual / quarterly / monthly) ──
        for candidate_date in period_end_candidates(side_date):
            df = pd.read_sql("""
                SELECT instrument_id FROM public.instrument_definitions
                WHERE product_id     = %(product_id)s
                  AND time_specifier = %(time_specifier)s
                  AND strike_value   = %(strike_value)s
                LIMIT 1
            """, engine, params={
                "product_id":     product_id,
                "time_specifier": candidate_date,
                "strike_value":   strike_value,
            })
            if not df.empty:
                return "in_db_period", df["instrument_id"].iloc[0], f"Matched via period-end date ({candidate_date})"

        # ── Tier 2b: instrument_raw date + nearby window (±7 days) ──
        # The date embedded in instrument_raw is the LTR filing date, which
        # may be close to (but not exactly) the event date in time_specifier.
        # Catches products like USDR where event date isn't derivable from side_date.
        if instrument_raw_date:
            from datetime import date, timedelta
            try:
                base = date(int(instrument_raw_date[:4]),
                            int(instrument_raw_date[4:6]),
                            int(instrument_raw_date[6:]))
                for delta in range(0, 8):  # 0 to +7 days
                    for d in ([base + timedelta(days=delta)] +
                              ([base - timedelta(days=delta)] if delta > 0 else [])):
                        candidate = d.strftime("%Y-%m-%d")
                        df = pd.read_sql("""
                            SELECT instrument_id FROM public.instrument_definitions
                            WHERE product_id     = %(product_id)s
                              AND time_specifier = %(time_specifier)s
                              AND strike_value   = %(strike_value)s
                            LIMIT 1
                        """, engine, params={
                            "product_id":     product_id,
                            "time_specifier": candidate,
                            "strike_value":   strike_value,
                        })
                        if not df.empty:
                            return "in_db_period", df["instrument_id"].iloc[0], f"Matched via date window (instrument date {instrument_raw_date} \u2192 event date {candidate})"
            except (ValueError, TypeError):
                pass

        # ── Tier 3a: alias match with exact date ──
        # cftc_strike_alias is stored as integer — cast strike_value for comparison
        try:
            alias_val = int(strike_value)
            df = pd.read_sql("""
                SELECT instrument_id, strike_value FROM public.instrument_definitions
                WHERE product_id        = %(product_id)s
                  AND time_specifier    = %(time_specifier)s
                  AND cftc_strike_alias = %(alias_val)s
                LIMIT 1
            """, engine, params={
                "product_id":     product_id,
                "time_specifier": time_specifier,
                "alias_val":      alias_val,
            })
            if not df.empty:
                return "in_db_alias", df["instrument_id"].iloc[0], f"Matched via strike alias, exact date (ordinal {alias_val} \u2192 '{df['strike_value'].iloc[0]}')"

            # ── Tier 3b: alias match, any date (FFDEC / GPALR — event date not in LTR) ──
            df = pd.read_sql("""
                SELECT instrument_id, strike_value FROM public.instrument_definitions
                WHERE product_id        = %(product_id)s
                  AND cftc_strike_alias = %(alias_val)s
                ORDER BY time_specifier DESC
                LIMIT 1
            """, engine, params={
                "product_id": product_id,
                "alias_val":  alias_val,
            })
            if not df.empty:
                text_strike = df["strike_value"].iloc[0]
                return "in_db_alias", df["instrument_id"].iloc[0], f"Matched via strike alias, any date (ordinal {alias_val} \u2192 '{text_strike}')"
        except (ValueError, TypeError):
            pass  # strike_value not numeric — skip alias tiers

        # ── Tier 4: product + strike_value, any date (USDR — specific event date) ──
        df = pd.read_sql("""
            SELECT instrument_id FROM public.instrument_definitions
            WHERE product_id   = %(product_id)s
              AND strike_value = %(strike_value)s
            ORDER BY time_specifier DESC
            LIMIT 1
        """, engine, params={
            "product_id":   product_id,
            "strike_value": strike_value,
        })
        if not df.empty:
            return "in_db_date", df["instrument_id"].iloc[0], "Product and strike found but date differs"

        # ── Tier 5: product exists at all ──
        df = pd.read_sql("""
            SELECT instrument_id FROM public.instrument_definitions
            WHERE product_id = %(product_id)s
            LIMIT 1
        """, engine, params={"product_id": product_id})
        if not df.empty:
            return "in_db_strike", None, "Product found but strike/alias not matched"

        return "not_in_db", None, "No matching instrument in database"

    except Exception as e:
        log.warning(f"DB lookup failed for {product_id}/{time_specifier}/{strike_value}: {e}")
        return "not_in_db", None, "DB lookup error"


# =============================================================================
# PARSING
# =============================================================================

DESC_PATTERN = re.compile(
    r"Error finding matching instrument for LTR line in (\S+):\s+"
    r"(\S+)\s+(\S+)\s+(\S+)\s+(\S+)\s+(YES|NO)(\d{6,8})\s*(?:\S+\s+)?\((.+?)\)\s*$"
)


def parse_description(desc: str) -> dict | None:
    """Parse a single ILTR alert description into structured fields."""
    if not isinstance(desc, str):
        return None
    m = DESC_PATTERN.search(desc.strip())
    if not m:
        return None
    instrument_raw = m.group(4)
    # Instrument ticker is the full instrument string used for DB lookup (prefix + date + strike)
    ticker = instrument_raw
    # Extract LTR file date from the filename embedded in the description (group 1)
    # Format: e.g. 534.LTR.20260316.txt — pull the 8-digit date
    file_date_match = re.search(r"(\d{8})", m.group(1))
    ltr_file_date = file_date_match.group(1) if file_date_match else ""
    return {
        "instrument_raw":    instrument_raw,
        "ltr_file_date":     ltr_file_date,   # date from LTR filename, e.g. 20260316
        "side_date":         m.group(7),      # YYYYMMDD — position date for DB lookup
        "instrument_ticker": ticker,          # product prefix, e.g. GCE
        "error_msg":         m.group(8),
    }


def parse_instrument_raw(instrument_raw: str) -> tuple:
    """
    Returns (prefix, strike_str) from instrument_raw.
    Format: {PREFIX}{YYYYMMDD}{STRIKE}
    Strike formatted to match DB storage (no trailing .0 for whole numbers).
    """
    # Prefix is everything before the first 8-digit date block
    # Handles alphanumeric prefixes like G08TX, G27TX etc.
    m = re.match(r"^([A-Z][A-Z0-9]*?)(\d{8})(.+)$", instrument_raw)
    if not m:
        return None, None
    prefix     = m.group(1)
    raw_strike = m.group(3)
    try:
        strike_f   = float(raw_strike)
        strike_str = str(int(strike_f)) if strike_f == int(strike_f) else str(strike_f)
    except ValueError:
        strike_str = raw_strike.lstrip("0") or "0"
    return prefix, strike_str


def format_side_date(side_date: str) -> str:
    """Convert YYYYMMDD or YYYYMM to YYYY-MM-DD for primary DB query."""
    if len(side_date) == 6:
        # Use first of month as primary; period_end_candidates() covers real dates
        return f"{side_date[:4]}-{side_date[4:6]}-01"
    return f"{side_date[:4]}-{side_date[4:6]}-{side_date[6:]}"


def period_end_candidates(side_date: str) -> list[str]:
    """
    For 6-digit YYYYMM side_dates, return candidate period-end dates to try.
    Covers annual (Dec 31), quarterly (Mar/Jun/Sep/Dec 31/30), and monthly (last day).
    Returns empty list for 8-digit dates (already have exact date).
    """
    import calendar
    if len(side_date) != 6:
        return []
    year  = int(side_date[:4])
    month = int(side_date[4:6])
    last_day = calendar.monthrange(year, month)[1]
    seen = set()
    candidates = []
    def add(d):
        if d not in seen:
            seen.add(d)
            candidates.append(d)
    # Annual: Dec 31
    if month == 12:
        add(f"{year}-12-31")
    # Quarterly: last day of quarter-end months
    if month in (3, 6, 9, 12):
        add(f"{year}-{month:02d}-{last_day}")
    # Monthly: last day of the month
    add(f"{year}-{month:02d}-{last_day}")
    return candidates


def get_error_type(error_msg: str) -> str:
    msg = error_msg.lower()
    if ERROR_NO_MATCH in msg:
        return "no_match"
    if ERROR_MULTI in msg:
        return "multiple"
    return "unknown"


# =============================================================================
# CLASSIFICATION  (per instrument group, cached to avoid repeat DB calls)
# =============================================================================

def classify_group(instrument_raw: str, error_type: str, prefix: str,
                   side_date: str, strike_str: str, raw_side_date: str,
                   engine) -> tuple[str, str | None, str]:
    """
    Return (adjudication_string, instrument_id | None, note).
    Note: multiple-match check happens BEFORE any DB lookup so known bug
    symbols (GCE, DISSN, etc.) are never incorrectly classified as date mismatches.
    """
    # ── Check multiple-match FIRST — before any DB lookup ──
    if error_type == "multiple":
        if prefix in KNOWN_BUG_SYMBOLS:
            return ADJ["known_bug"], None, f"Known multiple-match bug ({prefix})"
        else:
            return ADJ["new_bug"], None, f"New multiple-match bug — flag for review ({prefix})"

    if error_type == "no_match":
        # Extract the 8-digit date from instrument_raw for the window lookup
        raw_date_match = re.match(r"^[A-Z][A-Z0-9]*?(\d{8})", instrument_raw)
        instrument_raw_date = raw_date_match.group(1) if raw_date_match else ""

        tier, instrument_id, note = lookup_instrument(
            engine, prefix, side_date, strike_str, raw_side_date, instrument_raw_date
        )
        return ADJ[tier], instrument_id, note

    return ADJ["unknown"], None, "Unrecognized error message format"


# =============================================================================
# FILE READING
# =============================================================================

def read_source_file(path: Path) -> pd.DataFrame | None:
    """Read CSV, Excel, or TXT. Returns DataFrame of ILTR rows or None."""
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            df = pd.read_csv(path)
        elif suffix in (".xlsx", ".xls"):
            df = pd.read_excel(path)
        elif suffix == ".txt":
            for sep in (",", "\t", "|"):
                try:
                    df = pd.read_csv(path, sep=sep)
                    if len(df.columns) > 3:
                        break
                except Exception:
                    continue
            else:
                log.warning(f"Could not parse TXT file: {path.name}")
                return None
        else:
            log.warning(f"Unsupported file type: {path.name}")
            return None

        missing = REQUIRED_COLUMNS - set(df.columns)
        if missing:
            log.warning(f"{path.name} missing columns: {missing} -- skipping")
            return None

        df = df[df["alertName"] == "Invalid Large Trader Report"].copy()
        if df.empty:
            log.warning(f"No ILTR alerts in {path.name} -- skipping")
            return None

        log.info(f"Loaded {len(df)} ILTR alerts from {path.name}")
        return df

    except Exception as e:
        log.warning(f"Failed to read {path.name}: {e} -- skipping")
        return None


# =============================================================================
# EXCEL WRITER  — one flat file, original columns + Adjudication
# =============================================================================

def write_output(df: pd.DataFrame, source_path: Path):
    """
    Write a single Excel file with the original columns plus Adjudication.
    Named after the source file with _analyzed suffix.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ILTR Analysis"

    cols = list(df.columns)

    # ── Header row ──
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 18

    # Adjudication column colors
    adj_colors = {
        ADJ["in_db"]:        "C6EFCE",   # green       — exact match
        ADJ["in_db_alias"]:  "C6EFCE",   # green       — alias match (valid exact)
        ADJ["in_db_period"]: "FFEB9C",   # yellow      — date mismatch
        ADJ["in_db_date"]:   "FFEB9C",   # yellow      — date mismatch
        ADJ["in_db_strike"]: "FCE4D6",   # orange      — strike mismatch
        ADJ["not_in_db"]:    "FFC7CE",   # red         — not found
        ADJ["known_bug"]:    "BDD7EE",   # blue        — known bug
        ADJ["new_bug"]:      "FCE4D6",   # orange      — new bug
        ADJ["unknown"]:      "FFC7CE",   # red         — unknown
    }
    adj_col_idx = cols.index("Adjudication") + 1

    # ── Data rows ──
    for r_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = r_idx + 2
        bg = "FFFFFF" if r_idx % 2 == 0 else "F5F9FF"
        adj_val = str(row.get("Adjudication", ""))
        adj_bg  = adj_colors.get(adj_val, "EDEDED")

        for col_idx, col_name in enumerate(cols, start=1):
            val  = row.get(col_name, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            if col_idx == adj_col_idx:
                cell.fill = PatternFill("solid", fgColor=adj_bg)
                cell.font = Font(bold=True, size=9, name="Calibri")
            else:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[excel_row].height = 14

    # ── Column widths ──
    widths = {
        "id": 22, "alertName": 26, "alertTime": 26,
        "description": 70, "status": 22, "venue": 16,
        "LTR File Date": 16,
        "Side Date": 14,
        "Instrument Ticker": 20,
        "EP3 Matched Instrument Name": 34,
        "Match Notes": 52,
        "Adjudication": 42,
    }
    for col_idx, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 18)

    ws.freeze_panes = "A2"

    # ── Instrument Summary tab ──
    ws2 = wb.create_sheet("Instrument Summary")

    DESC_EXTRACT = re.compile(
        r"Error finding matching instrument for LTR line in \S+:\s+\S+\s+\S+\s+(\S+\s+\S+\s+(?:YES|NO)\S+\s+\S+)\s+\("
    )

    summary_cols = [
        "Alert Description Extract",
        "LTR File Date",
        "Side Date",
        "Instrument Ticker",
        "EP3 Matched Instrument Name",
        "Match Notes",
        "Adjudication",
    ]
    summary_widths = [52, 16, 14, 28, 34, 52, 42]

    for col_idx, col_name in enumerate(summary_cols, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(col_idx)].width = summary_widths[col_idx - 1]
    ws2.row_dimensions[1].height = 18
    ws2.freeze_panes = "A2"

    # Deduplicate on Instrument Ticker — keep first occurrence
    seen_tickers: dict[str, dict] = {}
    for _, row in df.iterrows():
        ticker = row.get("Instrument Ticker", "")
        if not ticker or ticker in seen_tickers:
            continue
        desc = str(row.get("description", ""))
        m = DESC_EXTRACT.search(desc)
        seen_tickers[ticker] = {
            "Alert Description Extract":    m.group(1) if m else "",
            "LTR File Date":                row.get("LTR File Date", ""),
            "Side Date":                    row.get("Side Date", ""),
            "Instrument Ticker":            ticker,
            "EP3 Matched Instrument Name":  row.get("EP3 Matched Instrument Name", ""),
            "Match Notes":                  row.get("Match Notes", ""),
            "Adjudication":                 row.get("Adjudication", ""),
        }

    for r_idx, summary_row in enumerate(sorted(seen_tickers.values(), key=lambda x: x["Instrument Ticker"]), start=2):
        bg = "FFFFFF" if r_idx % 2 == 0 else "F5F9FF"
        adj_val = str(summary_row.get("Adjudication", ""))
        adj_bg  = adj_colors.get(adj_val, "EDEDED")
        for col_idx, col_name in enumerate(summary_cols, start=1):
            val  = summary_row.get(col_name, "")
            cell = ws2.cell(row=r_idx, column=col_idx, value=val)
            is_adj = col_idx == len(summary_cols)
            cell.fill      = PatternFill("solid", fgColor=adj_bg if is_adj else bg)
            cell.font      = Font(bold=is_adj, size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[r_idx].height = 14

    log.info(f"  Instrument Summary: {len(seen_tickers)} unique tickers")

    # ── Output filename: source stem + _analyzed ──
    fname    = re.sub(r'[\\/:*?"<>|]', "_", f"{source_path.stem}_analyzed.xlsx")
    out_path = REPORTS_DIR / fname

    wb.save(out_path)
    log.info(f"Output written: {out_path.name}  ({len(df)} rows)")


# =============================================================================
# MAIN PROCESSING
# =============================================================================

def process_file(path: Path, engine, ep3_ref: dict = {}) -> bool:
    df = read_source_file(path)
    if df is None:
        return False

    # Parse all descriptions
    df["_parsed"] = df["description"].apply(parse_description)
    unparseable   = df["_parsed"].isna().sum()
    if unparseable:
        log.warning(f"  {unparseable} rows could not be parsed -- will be flagged")

    # Build classification + instrument_id cache keyed by instrument_raw
    classification_cache: dict[str, tuple[str, str | None, str]] = {}

    adjudications    = []
    ep3_names        = []
    notes            = []
    ltr_file_dates   = []
    side_dates_out   = []
    tickers          = []

    for _, row in df.iterrows():
        parsed = row["_parsed"]
        if parsed is None:
            adjudications.append(ADJ["unknown"])
            ep3_names.append("")
            notes.append("Unrecognized error message format")
            ltr_file_dates.append("")
            side_dates_out.append("")
            tickers.append("")
            continue

        instrument_raw  = parsed["instrument_raw"]
        raw_side_date   = parsed["side_date"]

        # Collect new granular fields from parsed description
        ltr_file_dates.append(parsed.get("ltr_file_date", ""))
        side_dates_out.append(raw_side_date)
        tickers.append(parsed.get("instrument_ticker", ""))

        if instrument_raw not in classification_cache:
            prefix, strike_str = parse_instrument_raw(instrument_raw)
            error_type         = get_error_type(parsed["error_msg"])
            side_date          = format_side_date(raw_side_date)

            if prefix is None:
                classification_cache[instrument_raw] = (ADJ["unknown"], None, "Could not parse instrument")
            else:
                classification_cache[instrument_raw] = classify_group(
                    instrument_raw, error_type, prefix,
                    side_date, strike_str, raw_side_date, engine
                )
            adj, inst_id, note = classification_cache[instrument_raw]
            log.info(f"  {instrument_raw} -> {adj}" + (f" [{inst_id}]" if inst_id else ""))

        adj, inst_id, note = classification_cache[instrument_raw]
        adjudications.append(adj)
        # Confirm EP3 name against reference table if available
        if inst_id and ep3_ref:
            ep3_name = ep3_ref.get(inst_id, inst_id)  # fallback to DB id if not in ref
        else:
            ep3_name = inst_id or ""
        ep3_names.append(ep3_name)
        notes.append(note)

    # Drop internal column, insert new granular columns then matched instrument/notes/adjudication
    df = df.drop(columns=["_parsed"])
    df["LTR File Date"]               = ltr_file_dates
    df["Side Date"]                   = side_dates_out
    df["Instrument Ticker"]           = tickers
    df["EP3 Matched Instrument Name"] = ep3_names
    df["Match Notes"]                 = notes
    df["Adjudication"]                = adjudications

    write_output(df, path)
    return True


def mark_completed(path: Path):
    new_path = path.parent / f"{path.stem}_completed{path.suffix}"
    try:
        path.rename(new_path)
        log.info(f"  Source marked completed: {new_path.name}")
    except Exception as e:
        log.warning(f"  Could not rename {path.name}: {e}")


def main():
    log.info("=" * 65)
    log.info("ForecastEx ILTR Processor")
    log.info(f"Inbox:     {INBOX_DIR}")
    log.info(f"Reports:   {REPORTS_DIR}")
    log.info(f"Reference: {REFERENCE_DIR}")
    log.info("=" * 65)

    candidates = [
        p for p in INBOX_DIR.iterdir()
        if p.is_file()
        and p.suffix.lower() in (".csv", ".xlsx", ".xls", ".txt")
        and "completed" not in p.name.lower()
    ]

    if not candidates:
        log.info("No new files in Inbox. Nothing to process.")
        return

    log.info(f"Found {len(candidates)} file(s) to process")

    # Load EP3 instrument reference table (optional but recommended)
    ep3_ref = load_ep3_reference()

    try:
        engine = get_engine()
        log.info("Database connection established.")
    except Exception as e:
        log.error(f"Could not connect to database: {e}")
        return

    processed = skipped = 0
    for path in sorted(candidates):
        if process_file(path, engine, ep3_ref):
            mark_completed(path)
            processed += 1
        else:
            skipped += 1

    log.info("=" * 65)
    log.info(f"Done. Processed: {processed}  Skipped: {skipped}")
    log.info(f"Reports: {REPORTS_DIR}")
    log.info("=" * 65)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log.info("Interrupted by user.")
    except Exception:
        crash_path = BASE_DIR / "iltr_processor_crash.txt"
        with open(crash_path, "w", encoding="utf-8") as f:
            f.write(traceback.format_exc())
        log.error(f"Crash details: {crash_path}")
        traceback.print_exc()
    finally:
        input("Press Enter to close...")
