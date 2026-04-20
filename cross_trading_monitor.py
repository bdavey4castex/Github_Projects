"""
=============================================================================
FORECASTEX CROSS TRADING ALERT MONITOR
=============================================================================
Polls public.alerts every 30 minutes for new Cross Trading alerts.
For each new alert, fetches execution and position data, classifies the
activity, and writes a standalone Excel report to the analyst's Documents
folder — but ONLY for alerts that are not confirmed MM dismissals.

USAGE:
    python cross_trading_monitor.py

    Runs indefinitely, polling every 30 minutes. Press Ctrl+C to stop.

OUTPUT FOLDER:
    ~/Documents/ForecastEx Compliance/Cross Trading/

OUTPUT FILES (one per alert pair — only non-MM-confirmed alerts):
    CrossTrading_{AlertID}_{Acct1}_{Acct2}_{Symbol}.xlsx

CLASSIFICATIONS:
    MM_CONFIRMED    Known MM acting as MM (two-sided) — silent dismiss, no file
    CRITERIA_1_BUG  Alert generated solely on Criteria 1 bug — silent dismiss, no file
    NEEDS_REVIEW    Everything else — Excel report generated
=============================================================================
"""

import os
import re
import sys
import json
import time
import logging
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path

import psycopg2
import psycopg2.extras
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURATION
# =============================================================================

DB_CONFIG = {
    "host":     "web-prod-db.cla246ai8gve.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "database": "postgres",
    "user":     "postgres_read_only",
    "password": os.environ.get("FXDB_PASSWORD", "pgreadonlypw"),
}

MARKET_MAKERS = {"FCX001", "RRA001"}
LQ_PROVIDERS  = {"UIBGIECTC", "UIBGIECTMC"}
ALL_MM        = MARKET_MAKERS | LQ_PROVIDERS

POLL_INTERVAL_SECONDS = 30 * 60
LOOKBACK_HOURS        = 96  # 4 days

OUTPUT_DIR  = Path.home() / "Documents" / "ForecastEx Compliance" / "Cross Trading"
REVIEWS_DIR = OUTPUT_DIR / "Reviews"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
REVIEWS_DIR.mkdir(parents=True, exist_ok=True)
STATE_FILE = OUTPUT_DIR / "cross_trading_monitor_state.json"

# Criteria detection phrases
_C1_PHRASE = "on one side and Account 2 traded"
_C2_PHRASE = "Multiple trades found between account 1 and account 2"
_C3_PHRASE = "Quantity traded between account 1 and account 2"


# =============================================================================
# LOGGING
# =============================================================================

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(OUTPUT_DIR / "cross_trading_monitor.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# =============================================================================
# STYLE HELPERS
# =============================================================================

HEX = {
    "navy":       "1F3864",
    "navy2":      "2E4D7B",
    "white":      "FFFFFF",
    "border":     "B8CCE4",
    "label_bg":   "EEF4FB",
    "section_bg": "D6E4F0",
    "alt_row":    "F5F9FF",
    "green":      "C6EFCE",
    "yellow":     "FFEB9C",
    "red":        "FFC7CE",
    "orange":     "FCE4D6",
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def _border(color=None):
    c = color or HEX["border"]
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# =============================================================================
# DATABASE
# =============================================================================

def get_engine():
    url = (
        f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
        f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
    )
    return create_engine(url)


def get_conn():
    return psycopg2.connect(
        host=DB_CONFIG["host"], port=DB_CONFIG["port"],
        dbname=DB_CONFIG["database"], user=DB_CONFIG["user"],
        password=DB_CONFIG["password"],
    )


def fetch_new_cross_trading_alerts(conn, processed_ids: set) -> list[dict]:
    """Pull unprocessed Cross Trading alerts from the last 24 hours (all statuses)."""
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT
            alert_id, accounts, symbol, description,
            alert_time, firm, status
        FROM alerts
        WHERE alert_name = 'Cross Trading'
          AND alert_time >= NOW() - INTERVAL %s
        ORDER BY alert_time DESC
    """, (f"{LOOKBACK_HOURS} hours",))
    rows = cur.fetchall()
    return [dict(r) for r in rows if r["alert_id"] not in processed_ids]


# =============================================================================
# ACCOUNT HELPERS
# =============================================================================

ACCOUNT_RE = re.compile(
    r'\b(FCX\w+|UIBGIECT\w*|RRA\w+|U\d\w*C|RH\w+\d)\b',
    re.IGNORECASE
)

def extract_account(raw: str) -> str:
    raw = str(raw).strip().strip("{}[]")
    if "/" in raw:
        return raw.split("/")[-1].strip().strip("{}[]")
    m = ACCOUNT_RE.search(raw)
    return m.group(1) if m else raw.strip()


def extract_all_accounts(raw) -> list[str]:
    if isinstance(raw, (list, tuple)):
        return [extract_account(str(a)) for a in raw]
    raw = str(raw).strip().strip("{}[]")
    return [extract_account(a.strip()) for a in raw.split(",") if a.strip()]


# =============================================================================
# CRITERIA DETECTION
# =============================================================================

def detect_criteria(description: str) -> tuple[bool, bool, bool]:
    desc = description or ""
    return (_C1_PHRASE in desc, _C2_PHRASE in desc, _C3_PHRASE in desc)


# =============================================================================
# DATA FETCHERS
# =============================================================================

def fetch_executions_for_accounts(conn, accounts: list[str],
                                   symbol: str, alert_time) -> pd.DataFrame:
    """Fetch executions for both accounts on the flagged instrument."""
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT account, symbol_subtype, execution_type,
                   order_quantity, fill_quantity, cum_quantity, leaves_quantity,
                   price, transact_time, trade_id,
                   order_id, order_status, order_type,
                   aggressor, time_in_force
            FROM executions
            WHERE account = ANY(%s)
              AND instrument_id = %s
              AND transact_time >= %s::timestamptz - INTERVAL '1 day'
            ORDER BY transact_time
        """, (accounts, symbol, alert_time))
        rows = cur.fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception as e:
        log.warning(f"  fetch_executions error: {e}")
        return pd.DataFrame()


def fetch_position_summary(conn, accounts: list[str],
                            symbol: str) -> dict[str, dict]:
    """Fetch final position state (P&L, flatness) from position_ledgers.

    after_realized is a running cumulative — take only the LAST row per
    account per symbol_subtype (by transaction_time) to get the true
    end-state P&L, then sum YES + NO sides for the account total.
    """
    result = {}
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        # Pull RESOLUTION rows first (most accurate cost basis and P&L).
        # before_cost at RESOLUTION = total cost paid across all contracts on that side.
        # Fallback to last ORDER_EXECUTION row if no RESOLUTION exists yet.
        cur.execute("""
            SELECT DISTINCT ON (account, symbol_subtype)
                account, symbol_subtype,
                after_realized, after_qty_bought, after_qty_sold,
                before_cost, entry_type,
                CASE WHEN entry_type = 'RESOLUTION' THEN 0 ELSE 1 END AS sort_priority
            FROM position_ledgers
            WHERE account = ANY(%s)
              AND instrument_id = %s
            ORDER BY account, symbol_subtype, sort_priority ASC, transaction_time DESC
        """, (accounts, symbol))
        rows = cur.fetchall()
        # after_realized = gross cash received (always positive).
        # True net P&L = sum(after_realized) - sum(abs(before_cost)) across all sides.
        # Per-side: profit = after_realized - abs(before_cost).
        for row in rows:
            acct       = row["account"]
            side       = row["symbol_subtype"]
            received   = float(row["after_realized"] or 0)
            cost       = abs(float(row["before_cost"] or 0))
            net_pnl    = received - cost
            bought     = int(row["after_qty_bought"] or 0)
            sold       = int(row["after_qty_sold"] or 0)
            qty        = bought if bought > 0 else sold
            avg_cost   = round(cost / qty, 4) if qty > 0 else None
            pnl_per    = round(net_pnl / qty, 4) if qty > 0 else None

            if acct not in result:
                result[acct] = {
                    "pnl": 0.0, "bought": 0, "sold": 0,
                    "sides": {},
                }
            result[acct]["pnl"]    += net_pnl
            result[acct]["bought"]  = max(result[acct]["bought"], bought)
            result[acct]["sold"]    = max(result[acct]["sold"], sold)
            result[acct]["sides"][side] = {
                "net_pnl":  round(net_pnl, 2),
                "received": round(received, 2),
                "cost":     round(cost, 2),
                "qty":      qty,
                "avg_cost": avg_cost,
                "pnl_per":  pnl_per,
                "resolved": row["entry_type"] == "RESOLUTION",
            }
    except Exception as e:
        log.warning(f"  fetch_position_summary error: {e}")
    return result


def fetch_beneficial_ownership(conn, accounts: list[str]) -> dict[str, dict]:
    """Fetch account BO info from accounts table."""
    result = {}
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT account_id, legal_name, state, zip, firm
            FROM accounts
            WHERE account_id = ANY(%s)
        """, (accounts,))
        for row in cur.fetchall():
            result[row["account_id"]] = dict(row)
    except Exception as e:
        log.warning(f"  fetch_beneficial_ownership error: {e}")
    return result


def fetch_prior_alerts(conn, accounts: list[str], alert_time) -> dict[str, int]:
    """Count prior Cross Trading alerts per account."""
    result = {a: 0 for a in accounts}
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        for acct in accounts:
            cur.execute("""
                SELECT COUNT(*) as cnt FROM alerts
                WHERE alert_name = 'Cross Trading'
                  AND accounts::text ILIKE %s
                  AND alert_time < %s::timestamptz
            """, (f"%{acct}%", alert_time))
            row = cur.fetchone()
            result[acct] = int(row["cnt"]) if row else 0
    except Exception as e:
        log.warning(f"  fetch_prior_alerts error: {e}")
    return result


def fetch_instrument_volume(conn, symbol: str, alert_date) -> dict:
    """Total market volume for the instrument on the alert date."""
    result = {"total_qty": 0, "trade_count": 0, "participants": 0}
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT
                SUM(quantity)                          AS total_qty,
                COUNT(*)                               AS trade_count,
                COUNT(DISTINCT aggressor_account)
                  + COUNT(DISTINCT passive_account)    AS approx_participants
            FROM pairs
            WHERE instrument_id = %s
              AND trading_date = %s::date
        """, (symbol, str(alert_date)[:10]))
        row = cur.fetchone()
        if row:
            result["total_qty"]    = int(row["total_qty"] or 0)
            result["trade_count"]  = int(row["trade_count"] or 0)
            result["participants"] = int(row["approx_participants"] or 0)
    except Exception as e:
        log.warning(f"  fetch_instrument_volume error: {e}")
    return result


def fetch_matched_pairs(conn, acct1: str, acct2: str,
                         symbol: str, alert_time) -> pd.DataFrame:
    """All pairs (matched trades) directly between acct1 and acct2."""
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT pair_id, quantity, yes_price, no_price,
                   aggressor_symbol_subtype, aggressor_account,
                   passive_symbol_subtype,  passive_account,
                   pair_time, trading_date
            FROM pairs
            WHERE instrument_id = %s
              AND pair_time >= %s::timestamptz - INTERVAL '1 day'
              AND (
                (aggressor_account = %s AND passive_account = %s)
                OR
                (aggressor_account = %s AND passive_account = %s)
              )
            ORDER BY pair_time
        """, (symbol, alert_time, acct1, acct2, acct2, acct1))
        rows = cur.fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception as e:
        log.warning(f"  fetch_matched_pairs error: {e}")
        return pd.DataFrame()


def fetch_all_market_pairs(conn, symbol: str, alert_date) -> pd.DataFrame:
    """All matched trades in the instrument on the alert date."""
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT pair_id, quantity, yes_price, no_price,
                   aggressor_symbol_subtype, aggressor_account, aggressor_firm,
                   passive_symbol_subtype,  passive_account,  passive_firm,
                   pair_time
            FROM pairs
            WHERE instrument_id = %s
              AND trading_date = %s::date
            ORDER BY pair_time
        """, (symbol, str(alert_date)[:10]))
        rows = cur.fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception as e:
        log.warning(f"  fetch_all_market_pairs error: {e}")
        return pd.DataFrame()


def fetch_prior_activity(conn, accounts: list[str],
                          symbol: str, alert_time) -> pd.DataFrame:
    """
    Prior trading activity for both accounts:
    - Same instrument on previous days
    - Same source agency (instrument root) over last 30 days
    """
    try:
        symbol_root = symbol.split("_")[0] if symbol else ""
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT
                e.account, e.instrument_id, e.symbol_subtype,
                e.execution_type, e.fill_quantity, e.order_quantity,
                e.price, e.transact_time, e.trading_date,
                p.source_agency
            FROM executions e
            LEFT JOIN instrument_definitions id
                ON e.instrument_id = id.instrument_id
            LEFT JOIN products p
                ON id.product_id = p.product_id
            WHERE e.account = ANY(%s)
              AND e.transact_time < %s::timestamptz
              AND e.transact_time >= %s::timestamptz - INTERVAL '30 days'
              AND (
                e.instrument_id = %s
                OR p.source_agency IN (
                    SELECT source_agency FROM products
                    JOIN instrument_definitions id2
                        ON products.product_id = id2.product_id
                    WHERE id2.instrument_id = %s
                    LIMIT 1
                )
              )
              AND e.execution_type IN ('FILL', 'PARTIAL_FILL')
            ORDER BY e.transact_time DESC
        """, (accounts, alert_time, alert_time, symbol, symbol))
        rows = cur.fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception as e:
        log.warning(f"  fetch_prior_activity error: {e}")
        return pd.DataFrame()


def detect_trading_pattern(execs: pd.DataFrame, acct: str) -> dict:
    """
    Classify trading pattern for an account:
    - DAY_TRADER: opens and closes position once by end of day
    - BACK_AND_FORTH: multiple direction reversals (manipulative indicator)
    - DIRECTIONAL: builds position one way, no reversal
    - INSUFFICIENT_DATA: not enough data to classify
    """
    if execs.empty:
        return {"pattern": "INSUFFICIENT_DATA", "reversals": 0, "detail": "No execution data."}

    acct_fills = execs[
        (execs["account"] == acct) &
        (execs["execution_type"].isin(["FILL", "PARTIAL_FILL"]))
    ].sort_values("transact_time").copy()

    if acct_fills.empty:
        return {"pattern": "INSUFFICIENT_DATA", "reversals": 0, "detail": "No fills found."}

    # Build running net position (YES = +qty, NO = -qty)
    acct_fills["signed_qty"] = acct_fills.apply(
        lambda r: r["fill_quantity"] if r["symbol_subtype"] == "YES" else -r["fill_quantity"],
        axis=1
    )
    acct_fills["net_pos"] = acct_fills["signed_qty"].cumsum()

    # Count direction reversals (sign changes in net position)
    signs = acct_fills["net_pos"].apply(lambda x: 1 if x > 0 else (-1 if x < 0 else 0))
    signs = signs[signs != 0]
    reversals = (signs != signs.shift()).sum() - 1 if len(signs) > 1 else 0

    final_pos = acct_fills["net_pos"].iloc[-1]
    max_pos   = acct_fills["net_pos"].abs().max()

    if reversals == 0:
        pattern = "DIRECTIONAL"
        detail  = f"Built position in one direction (net {final_pos:+,}), no reversals."
    elif reversals == 1 and abs(final_pos) <= max_pos * 0.10:
        pattern = "DAY_TRADER"
        detail  = f"Opened then closed position (max {max_pos:,}, ended near flat at {final_pos:+,}). Consistent with day trading."
    else:
        pattern = "BACK_AND_FORTH"
        detail  = f"{reversals} direction reversals detected. Position swung between YES and NO multiple times — consistent with manipulative back-and-forth trading."

    return {"pattern": pattern, "reversals": int(reversals),
            "final_pos": int(final_pos), "max_pos": int(max_pos), "detail": detail}


def count_volume_between(execs: pd.DataFrame,
                          acct1: str, acct2: str) -> int:
    """Count contracts where both accounts traded on the same trade_id."""
    if execs.empty or "trade_id" not in execs.columns:
        return 0
    fills = execs[execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])]
    acct1_trades = set(fills[fills["account"] == acct1]["trade_id"].dropna())
    acct2_trades = set(fills[fills["account"] == acct2]["trade_id"].dropna())
    matched = acct1_trades & acct2_trades
    if not matched:
        return 0
    return int(fills[fills["trade_id"].isin(matched)]["fill_quantity"].sum() // 2)


COMMON_SIZES = {10, 25, 50, 100, 250, 500, 1000, 2500, 5000, 10000}


def analyze_execution_detail(execs: pd.DataFrame,
                              acct1: str, acct2: str) -> dict:
    """
    For each matched trade_id (shared between acct1 and acct2):
    - Compute time gap between the NEW orders placed by each account
    - Compare fill quantities for size similarity
    Returns summary stats used for auto-dismiss determination.
    """
    result = {
        "matched_trade_count":    0,
        "median_timing_gap_secs": None,
        "min_timing_gap_secs":    None,
        "pct_timing_under_30s":   0.0,
        "pct_timing_under_5s":    0.0,
        "pct_similar_size":       0.0,
        "timing_gaps":            [],
        "size_pairs":             [],
        "can_auto_dismiss":       False,
        "auto_dismiss_reasons":   [],
        "flag_reasons":           [],
    }

    if execs.empty or "trade_id" not in execs.columns:
        result["can_auto_dismiss"] = False
        result["flag_reasons"].append("no execution data available for timing analysis")
        return result

    fills = execs[execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])].copy()
    news  = execs[execs["execution_type"] == "NEW"].copy()

    acct1_trade_ids = set(fills[fills["account"] == acct1]["trade_id"].dropna())
    acct2_trade_ids = set(fills[fills["account"] == acct2]["trade_id"].dropna())
    matched_ids = acct1_trade_ids & acct2_trade_ids

    if not matched_ids:
        result["can_auto_dismiss"] = False
        result["flag_reasons"].append("no matched trade IDs found between accounts")
        return result

    result["matched_trade_count"] = len(matched_ids)
    timing_gaps = []
    size_pairs  = []

    for trade_id in matched_ids:
        # Get fill quantities for size comparison
        a1_fills = fills[(fills["account"] == acct1) & (fills["trade_id"] == trade_id)]
        a2_fills = fills[(fills["account"] == acct2) & (fills["trade_id"] == trade_id)]
        a1_qty = int(a1_fills["fill_quantity"].sum()) if not a1_fills.empty else 0
        a2_qty = int(a2_fills["fill_quantity"].sum()) if not a2_fills.empty else 0

        if a1_qty > 0 and a2_qty > 0:
            size_pairs.append((a1_qty, a2_qty))

        # Get NEW order times for timing comparison
        # Look for NEW orders placed before the fill time for each account
        fill_time = fills[fills["trade_id"] == trade_id]["transact_time"].min()
        if fill_time is None:
            continue

        fill_time_dt = pd.to_datetime(fill_time, utc=True)

        a1_news = news[
            (news["account"] == acct1) &
            (pd.to_datetime(news["transact_time"], utc=True) <= fill_time_dt)
        ]
        a2_news = news[
            (news["account"] == acct2) &
            (pd.to_datetime(news["transact_time"], utc=True) <= fill_time_dt)
        ]

        if a1_news.empty or a2_news.empty:
            continue

        a1_order_time = pd.to_datetime(a1_news["transact_time"], utc=True).max()
        a2_order_time = pd.to_datetime(a2_news["transact_time"], utc=True).max()
        gap_secs = abs((a1_order_time - a2_order_time).total_seconds())
        timing_gaps.append(gap_secs)

    result["timing_gaps"] = timing_gaps
    result["size_pairs"]  = size_pairs

    # Timing stats
    if timing_gaps:
        import statistics
        result["median_timing_gap_secs"] = statistics.median(timing_gaps)
        result["min_timing_gap_secs"]    = min(timing_gaps)
        result["pct_timing_under_30s"]   = sum(1 for g in timing_gaps if g < 30) / len(timing_gaps) * 100
        result["pct_timing_under_5s"]    = sum(1 for g in timing_gaps if g < 5)  / len(timing_gaps) * 100

    # Size similarity stats
    similar_count = 0
    for q1, q2 in size_pairs:
        larger = max(q1, q2)
        # Skip common sizes — not suspicious
        if q1 in COMMON_SIZES and q2 in COMMON_SIZES:
            continue
        if larger > 0 and abs(q1 - q2) / larger <= 0.10:
            similar_count += 1
    if size_pairs:
        result["pct_similar_size"] = similar_count / len(size_pairs) * 100

    # --- Auto-dismiss determination ---
    flag_reasons    = []
    dismiss_reasons = []

    # Timing flags
    if timing_gaps:
        median_gap = result["median_timing_gap_secs"]
        pct_under_5 = result["pct_timing_under_5s"]
        pct_under_30 = result["pct_timing_under_30s"]

        if pct_under_5 > 25:
            flag_reasons.append(
                f"{pct_under_5:.0f}% of matched orders placed within 5 seconds of each other"
            )
        elif pct_under_30 > 50:
            flag_reasons.append(
                f"{pct_under_30:.0f}% of matched orders placed within 30 seconds of each other"
            )
        else:
            dismiss_reasons.append(
                f"median order timing gap {median_gap:.0f}s — orders not placed in close proximity"
            )

    # Size flags
    if size_pairs:
        pct_sim = result["pct_similar_size"]
        if pct_sim >= 25:
            flag_reasons.append(
                f"{pct_sim:.0f}% of matched trades had similar non-standard order sizes"
            )
        else:
            dismiss_reasons.append(
                f"only {pct_sim:.0f}% of matched trades had similar sizes — not suggestive of coordination"
            )

    result["flag_reasons"]          = flag_reasons
    result["auto_dismiss_reasons"]  = dismiss_reasons
    result["can_auto_dismiss"]      = len(flag_reasons) == 0 and len(dismiss_reasons) > 0

    return result


def check_two_sided(execs: pd.DataFrame, account: str) -> bool:
    """Returns True if account has fills on both YES and NO sides."""
    if execs.empty:
        return False
    acct_execs = execs[execs["account"] == account]
    sides = acct_execs[
        acct_execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])
    ]["symbol_subtype"].unique()
    return len(set(sides)) >= 2


def compare_bo(bo: dict[str, dict], acct1: str, acct2: str) -> list[str]:
    """Return list of matching BO fields between two accounts."""
    a1 = bo.get(acct1, {})
    a2 = bo.get(acct2, {})
    matches = []
    if not a1 or not a2:
        return matches
    n1 = (a1.get("legal_name") or "").strip().upper()
    n2 = (a2.get("legal_name") or "").strip().upper()
    if n1 and n2:
        if n1 == n2:
            matches.append(f"full name ({a1.get('legal_name')})")
        else:
            ln1 = n1.split()[-1] if n1 else ""
            ln2 = n2.split()[-1] if n2 else ""
            if ln1 and ln2 and ln1 == ln2:
                matches.append(f"last name ({a1.get('legal_name')} / {a2.get('legal_name')})")
    z1 = (a1.get("zip") or "").strip()
    z2 = (a2.get("zip") or "").strip()
    if z1 and z2 and z1 == z2:
        matches.append(f"zip code ({z1})")
    return matches


# =============================================================================
# CLASSIFICATION
# =============================================================================

def classify_alert(alert: dict, conn) -> tuple[str, str, dict]:
    """
    Returns (classification, note, context_dict).
    Classifications:
        MM_CONFIRMED    — known MM, two-sided confirmed — no report
        CRITERIA_1_BUG  — criteria 1 only bug — no report
        NEEDS_REVIEW    — everything else — generate report
    """
    description = alert.get("description", "") or ""
    symbol      = alert.get("symbol", "") or ""
    alert_time  = alert.get("alert_time")
    alert_id    = alert.get("alert_id", "")

    all_accounts = extract_all_accounts(alert.get("accounts", []))
    acct1 = all_accounts[0] if all_accounts else ""
    acct2 = all_accounts[1] if len(all_accounts) > 1 else ""

    c1, c2, c3 = detect_criteria(description)

    # --- Criteria 1 bug dismiss ---
    if c1 and not c2 and not c3:
        note = (
            "Alert generating due to known bug which causes account 1 and account 2's "
            "volume to be compared against each other as opposed to each account's volume "
            "being compared against itself on the other side of the market. None of the "
            "additional criteria were met and the alert can be dismissed."
        )
        return "CRITERIA_1_BUG", note, {}

    # Fetch executions
    execs = fetch_executions_for_accounts(conn, all_accounts, symbol, alert_time)

    # --- MM two-sided check ---
    mm_account = next((a for a in all_accounts if a in ALL_MM), None)
    if mm_account:
        if check_two_sided(execs, mm_account):
            note = (
                f"{mm_account} is a known market maker, has been confirmed to be acting "
                f"as a market maker at this time, and therefore this alert can be dismissed. "
                f"This activity is not considered to be problematic."
            )
            return "MM_CONFIRMED", note, {}
        # MM not two-sided — fall through to NEEDS_REVIEW

    # --- NEEDS_REVIEW — build full context ---
    alert_date  = str(alert_time)[:10] if alert_time else ""
    positions   = fetch_position_summary(conn, all_accounts, symbol)
    bo          = fetch_beneficial_ownership(conn, all_accounts)
    prior       = fetch_prior_alerts(conn, all_accounts, alert_time)
    bo_matches  = compare_bo(bo, acct1, acct2)
    qty_between = count_volume_between(execs, acct1, acct2)
    instr_vol   = fetch_instrument_volume(conn, symbol, alert_date)
    matched_pairs = fetch_matched_pairs(conn, acct1, acct2, symbol, alert_time)
    market_pairs  = fetch_all_market_pairs(conn, symbol, alert_date)
    prior_activity = fetch_prior_activity(conn, all_accounts, symbol, alert_time)

    exec_analysis = analyze_execution_detail(execs, acct1, acct2)
    pattern1 = detect_trading_pattern(execs, acct1)
    pattern2 = detect_trading_pattern(execs, acct2)

    context = {
        "acct1": acct1, "acct2": acct2,
        "symbol": symbol, "description": description,
        "c1": c1, "c2": c2, "c3": c3,
        "execs": execs,
        "positions": positions,
        "bo": bo,
        "bo_matches": bo_matches,
        "prior": prior,
        "qty_between": qty_between,
        "mm_not_acting": mm_account,
        "exec_analysis": exec_analysis,
        "instr_vol": instr_vol,
        "matched_pairs": matched_pairs,
        "market_pairs": market_pairs,
        "prior_activity": prior_activity,
        "pattern1": pattern1,
        "pattern2": pattern2,
        "alert_date": alert_date,
    }

    note = build_review_note(context)
    return "NEEDS_REVIEW", note, context


def build_review_note(ctx: dict) -> str:
    lines      = []
    cross_flags = []   # cross trading specific
    wash_flags  = []   # wash trading specific
    general_flags = []

    acct1  = ctx["acct1"]
    acct2  = ctx["acct2"]
    symbol = ctx["symbol"]
    c1, c2, c3 = ctx["c1"], ctx["c2"], ctx["c3"]

    if ctx.get("mm_not_acting"):
        mm = ctx["mm_not_acting"]
        lines.append(
            f"NOTE: {mm} is a known market maker but has not been confirmed to be acting "
            f"as a market maker during this time interval. "
            f"This account should be treated as any other account in this review."
        )
        general_flags.append(f"{mm} is a known market maker not acting as one")

    # Criteria
    triggered = []
    if c1: triggered.append("Criteria 1 (similar volumes)")
    if c2: triggered.append("Criteria 2 (similar order times/quantities)")
    if c3: triggered.append("Criteria 3 (large % of total volume)")
    criteria_str = ", ".join(triggered) if triggered else "unknown criteria"
    lines.append(f"Alert triggered on {criteria_str}.")
    if len(triggered) >= 2:
        general_flags.append("multiple criteria triggered (higher risk)")

    # ── P&L ──
    pos = ctx["positions"]
    p1_data = pos.get(acct1, {})
    p2_data = pos.get(acct2, {})
    net_pnl1 = p1_data.get("pnl", 0.0)
    net_pnl2 = p2_data.get("pnl", 0.0)
    combined_pnl = net_pnl1 + net_pnl2

    for acct, p in [(acct1, p1_data), (acct2, p2_data)]:
        net_pnl = p.get("pnl", 0.0)
        bought  = p.get("bought", 0)
        sold    = p.get("sold", 0)
        total   = bought + sold
        if total == 0:
            continue

        pnl_str  = f"${net_pnl:,.2f}" if net_pnl >= 0 else f"-${abs(net_pnl):,.2f}"
        net_per  = round(net_pnl / (bought or 1), 4)
        flat_pct = abs(bought - sold) / max(bought, sold) * 100 if max(bought, sold) > 0 else 0

        if bought == sold:
            flatness = "exactly flat"
            wash_flags.append(f"{acct} is exactly flat (stronger wash trade indicator than near-flat)")
        elif flat_pct <= 10:
            flatness = f"near flat ({flat_pct:.1f}% imbalance)"
            wash_flags.append(f"{acct} ended near flat (wash trade indicator)")
        else:
            flatness = f"directional ({flat_pct:.1f}% imbalance)"

        sides   = p.get("sides", {})
        s_parts = []
        for side in ("YES", "NO"):
            s = sides.get(side)
            if not s:
                continue
            qty    = s.get("qty", 0)
            ac     = s.get("avg_cost")
            s_net  = s.get("net_pnl", 0.0)
            s_pp   = s.get("pnl_per")
            res    = s.get("resolved", False)
            if ac is not None and qty > 0:
                pp_str = (f"~${s_pp:.2f}/contract profit" if s_pp is not None and s_pp >= 0
                          else f"~${abs(s_pp):.2f}/contract loss" if s_pp is not None else "")
                s_parts.append(
                    f"{side}: {qty:,} contracts @ avg ${ac:.2f}, "
                    f"net {'+' if s_net >= 0 else ''}{s_net:,.2f} ({pp_str})"
                    + (" [RESOLVED]" if res else "")
                )
        side_str = "; ".join(s_parts) if s_parts else "no side detail"
        lines.append(
            f"{acct}: {bought:,} bought / {sold:,} sold ({flatness}) | "
            f"Net P&L: {pnl_str} (~${net_per:.2f}/contract). "
            f"Side breakdown — {side_str}."
        )

    # Combined P&L signal
    combined_str = f"${combined_pnl:,.2f}" if combined_pnl >= 0 else f"-${abs(combined_pnl):,.2f}"
    lines.append(f"Combined net P&L across both accounts: {combined_str}.")
    if net_pnl1 * net_pnl2 < 0:
        cross_flags.append(
            f"one account profited (${max(net_pnl1,net_pnl2):,.2f}) while the other lost "
            f"(-${abs(min(net_pnl1,net_pnl2)):,.2f}) — consistent with money pass / cross trading"
        )
    elif abs(combined_pnl) < 50:
        wash_flags.append(
            f"combined P&L near zero ({combined_str}) with both accounts flat — "
            f"consistent with wash trading"
        )

    # ── Instrument volume and account concentration ──
    instr_vol = ctx.get("instr_vol", {})
    total_instr_qty = instr_vol.get("total_qty", 0)
    bought1 = p1_data.get("bought", 0)
    bought2 = p2_data.get("bought", 0)
    if total_instr_qty > 0:
        pct1_instr = (bought1 + p1_data.get("sold", 0)) / total_instr_qty * 100
        pct2_instr = (bought2 + p2_data.get("sold", 0)) / total_instr_qty * 100
        lines.append(
            f"Total instrument volume on alert date: {total_instr_qty:,} contracts "
            f"({instr_vol.get('trade_count',0):,} trades). "
            f"{acct1} accounted for {pct1_instr:.1f}% of total volume; "
            f"{acct2} accounted for {pct2_instr:.1f}%."
        )
        if pct1_instr >= 25 or pct2_instr >= 25:
            general_flags.append(
                f"high individual volume concentration "
                f"({acct1}: {pct1_instr:.1f}%, {acct2}: {pct2_instr:.1f}%)"
            )

    # ── Cross-account flatness (were they flat vs each other?) ──
    qty_between = ctx.get("qty_between", 0)
    total1 = p1_data.get("bought", 0) + p1_data.get("sold", 0)
    total2 = p2_data.get("bought", 0) + p2_data.get("sold", 0)
    if qty_between and (total1 or total2):
        pct_of1 = qty_between / total1 * 100 if total1 else 0
        pct_of2 = qty_between / total2 * 100 if total2 else 0
        cross_flat = abs(qty_between * 2 - qty_between * 2) <= 1  # they always trade equal qty
        lines.append(
            f"Volume traded directly between accounts: {qty_between:,} contracts "
            f"({pct_of1:.1f}% of {acct1}'s total, {pct_of2:.1f}% of {acct2}'s total). "
            f"Both accounts are flat relative to each other "
            f"(each traded {qty_between:,} contracts against the other)."
        )
        if pct_of1 >= 25 or pct_of2 >= 25:
            cross_flags.append(
                f"high inter-account volume concentration ({max(pct_of1,pct_of2):.1f}% of "
                f"one account's total activity was directly with the other account)"
            )

    # ── Beneficial ownership ──
    bo    = ctx.get("bo", {})
    a1_bo = bo.get(acct1, {})
    a2_bo = bo.get(acct2, {})
    if a1_bo or a2_bo:
        lines.append(
            f"Beneficial ownership: {acct1}: {a1_bo.get('legal_name','N/A')} "
            f"({a1_bo.get('state','?')}/{a1_bo.get('zip','?')}) | "
            f"{acct2}: {a2_bo.get('legal_name','N/A')} "
            f"({a2_bo.get('state','?')}/{a2_bo.get('zip','?')})."
        )
    bo_matches = ctx.get("bo_matches", [])
    if bo_matches:
        cross_flags.append(f"beneficial ownership match on {', '.join(bo_matches)}")
        lines.append(
            f"POTENTIAL MATCH: accounts share {', '.join(bo_matches)}. "
            f"This substantially increases the risk of cross trading."
        )
    else:
        lines.append("No matching beneficial ownership fields identified.")

    # ── Prior alerts ──
    prior  = ctx.get("prior", {})
    p1_cnt = prior.get(acct1, 0)
    p2_cnt = prior.get(acct2, 0)
    if p1_cnt or p2_cnt:
        lines.append(f"Prior cross trading alerts: {acct1}: {p1_cnt} | {acct2}: {p2_cnt}.")
        general_flags.append("prior cross trading alerts on record")

    # ── Execution timing/sizing ──
    ea = ctx.get("exec_analysis", {})
    if ea:
        matched    = ea.get("matched_trade_count", 0)
        median_gap = ea.get("median_timing_gap_secs")
        pct_5s     = ea.get("pct_timing_under_5s", 0)
        pct_30s    = ea.get("pct_timing_under_30s", 0)
        pct_sim    = ea.get("pct_similar_size", 0)
        if matched:
            timing_str = f"{median_gap:.0f}s median" if median_gap is not None else "N/A"
            lines.append(
                f"Execution analysis: {matched} matched trade(s). "
                f"Order timing gap — {timing_str} "
                f"({pct_5s:.0f}% under 5s, {pct_30s:.0f}% under 30s). "
                f"Similar non-standard sizing: {pct_sim:.0f}% of matched trades."
            )
        for reason in ea.get("flag_reasons", []):
            cross_flags.append(reason)

    # ── Trading pattern ──
    p1_pat = ctx.get("pattern1", {})
    p2_pat = ctx.get("pattern2", {})
    for acct, pat in [(acct1, p1_pat), (acct2, p2_pat)]:
        if pat.get("pattern"):
            lines.append(f"Trading pattern ({acct}): {pat['detail']}")
            if pat["pattern"] == "BACK_AND_FORTH":
                wash_flags.append(
                    f"{acct} shows back-and-forth trading pattern "
                    f"({pat['reversals']} direction reversals)"
                )

    # ── Narrative conclusion ──
    all_flags = cross_flags + wash_flags + general_flags
    has_cross  = len(cross_flags) > 0
    has_wash   = len(wash_flags) > 0

    if has_cross and has_wash:
        narrative = (
            f"This alert exhibits indicators of both cross trading and wash trading. "
            f"Cross trading indicators: {'; '.join(cross_flags)}. "
            f"Wash trading indicators: {'; '.join(wash_flags)}. "
            f"Analyst should review matched trades sheet and escalate to CRO if pre-arranged activity is suspected."
        )
    elif has_cross:
        narrative = (
            f"This alert exhibits indicators consistent with cross trading: {'; '.join(cross_flags)}. "
            f"Wash trading indicators are not present. "
            f"Analyst should review matched trades sheet and escalate to CRO if pre-arranged activity is suspected."
        )
    elif has_wash:
        narrative = (
            f"This alert exhibits indicators consistent with wash trading: {'; '.join(wash_flags)}. "
            f"Cross trading indicators are not present. "
            f"Analyst should review execution pattern before closing."
        )
    elif ea.get("can_auto_dismiss", False):
        dismiss_reasons = ea.get("auto_dismiss_reasons", [])
        reason_str = "; ".join(dismiss_reasons) if dismiss_reasons else "no red flags identified"
        narrative = (
            f"Review of trading activity, beneficial ownership, position data, and execution "
            f"timing and sizing does not suggest pre-arranged trading ({reason_str}). "
            f"This activity is not considered to be problematic."
        )
    else:
        narrative = (
            "No significant red flags identified in the automated review. "
            "Analyst should confirm before closing."
        )
    if general_flags:
        narrative += f" Additional notes: {'; '.join(general_flags)}."
    lines.append(narrative)

    return " ".join(lines)


# =============================================================================
# EXCEL REPORT
# =============================================================================

def write_alert_report(alert: dict, classification: str,
                        note: str, context: dict):
    """Write a standalone Excel report for a NEEDS_REVIEW alert."""
    alert_id = alert.get("alert_id", "UNKNOWN")
    acct1    = context.get("acct1", "")
    acct2    = context.get("acct2", "")
    symbol   = context.get("symbol", alert.get("symbol", ""))

    safe = lambda s: re.sub(r'[\\/:*?"<>|]', '_', str(s))
    filename = f"CrossTrading_{safe(alert_id)}_{safe(acct1)}_{safe(acct2)}_{safe(symbol)}.xlsx"

    wb = Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70

    # Title banner
    ws.merge_cells("A1:B1")
    tc = ws["A1"]
    tc.value     = f"Cross Trading Alert — {acct1} / {acct2} — {symbol}"
    tc.fill      = _fill(HEX["navy"])
    tc.font      = _font(bold=True, color=HEX["white"], size=13)
    tc.alignment = _align("center")
    ws.row_dimensions[1].height = 24

    # Classification banner
    cls_color = HEX["red"] if classification == "NEEDS_REVIEW" else HEX["green"]
    ws.merge_cells("A2:B2")
    cc = ws["A2"]
    cc.value     = f"Classification: {classification}"
    cc.fill      = _fill(cls_color)
    cc.font      = _font(bold=True, size=11)
    cc.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    # Metadata rows
    alert_time = alert.get("alert_time")
    if hasattr(alert_time, "strftime"):
        from zoneinfo import ZoneInfo
        eastern = ZoneInfo("America/New_York")
        try:
            alert_time_str = alert_time.astimezone(eastern).strftime("%Y-%m-%d %H:%M ET")
        except Exception:
            alert_time_str = str(alert_time)
    else:
        alert_time_str = str(alert_time)

    meta = [
        ("Alert ID",    alert_id),
        ("Alert Time",  alert_time_str),
        ("Account 1",   acct1),
        ("Account 2",   acct2),
        ("Symbol",      symbol),
        ("Criteria",    ", ".join(
            [k for k, v in [("C1", context.get("c1")),
                             ("C2", context.get("c2")),
                             ("C3", context.get("c3"))] if v]
        ) or "Unknown"),
        ("Description", alert.get("description", "")),
    ]

    for ri, (label, value) in enumerate(meta, start=3):
        lc = ws.cell(ri, 1, label)
        lc.fill      = _fill(HEX["section_bg"])
        lc.font      = _font(bold=True, size=10)
        lc.border    = _border()
        lc.alignment = _align()

        vc = ws.cell(ri, 2, value)
        vc.fill      = _fill(HEX["label_bg"] if ri % 2 == 0 else HEX["white"])
        vc.font      = _font(size=10)
        vc.border    = _border()
        vc.alignment = _align(wrap=True)
        ws.row_dimensions[ri].height = 30 if label == "Description" else 16

    # Adjudication note
    note_row = len(meta) + 4
    ws.merge_cells(f"A{note_row}:B{note_row}")
    nh = ws.cell(note_row, 1, "Adjudication Note")
    nh.fill      = _fill(HEX["navy2"])
    nh.font      = _font(bold=True, color=HEX["white"], size=11)
    nh.alignment = _align("center")
    ws.row_dimensions[note_row].height = 18

    note_row += 1
    ws.merge_cells(f"A{note_row}:B{note_row}")
    nc = ws.cell(note_row, 1, note)
    nc.fill      = _fill("FFFBE6")
    nc.font      = _font(size=10)
    nc.alignment = _align(wrap=True)
    ws.row_dimensions[note_row].height = max(15 * (len(note) // 80 + 1), 60)

    # Conclusion banner — extract and highlight the final sentence
    sentences = [s.strip() for s in note.split(". ") if s.strip()]
    conclusion = sentences[-1] if sentences else note
    if not conclusion.endswith("."):
        conclusion += "."
    red_flags_present = "warrant further review" in note
    conclusion_color = HEX["red"] if red_flags_present else HEX["green"]

    conclusion_row = note_row + 1
    ws.merge_cells(f"A{conclusion_row}:B{conclusion_row}")
    cc2 = ws.cell(conclusion_row, 1, f"Conclusion: {conclusion}")
    cc2.fill      = _fill(conclusion_color)
    cc2.font      = _font(bold=True, size=11)
    cc2.alignment = _align(wrap=True)
    ws.row_dimensions[conclusion_row].height = max(15 * (len(conclusion) // 80 + 1), 22)

    # ── Sheet 2: Execution Detail ──
    execs = context.get("execs", pd.DataFrame())
    ws2 = wb.create_sheet("Execution Detail")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:O1")
    t2 = ws2["A1"]
    t2.value     = f"Execution Detail — {acct1} / {acct2} — {symbol}"
    t2.fill      = _fill(HEX["navy"])
    t2.font      = _font(bold=True, color=HEX["white"], size=12)
    t2.alignment = _align("center")
    ws2.row_dimensions[1].height = 22

    exec_headers = [
        ("Account",      "account",         18),
        ("Time (UTC)",   "transact_time",   22),
        ("Side",         "symbol_subtype",   7),
        ("Exec Type",    "execution_type",  15),
        ("Order Qty",    "order_quantity",  10),
        ("Fill Qty",     "fill_quantity",   10),
        ("Cum Qty",      "cum_quantity",    10),
        ("Leaves Qty",   "leaves_quantity", 10),
        ("Price ($)",    "price",           10),
        ("Aggressor",    "aggressor",       10),
        ("Order Type",   "order_type",      12),
        ("Order Status", "order_status",    16),
        ("TIF",          "time_in_force",   16),
        ("Order ID",     "order_id",        28),
        ("Trade ID",     "trade_id",        28),
    ]

    if execs.empty:
        ws2["A3"] = "No execution data available."
        ws2["A3"].font = _font(italic=True, color="888888")
    else:
        for ci, (hdr, _, width) in enumerate(exec_headers, 1):
            c = ws2.cell(2, ci, hdr)
            c.fill      = _fill(HEX["navy2"])
            c.font      = _font(bold=True, color=HEX["white"], size=10)
            c.alignment = _align("center")
            c.border    = _border()
            ws2.column_dimensions[get_column_letter(ci)].width = width
        ws2.row_dimensions[2].height = 18

        for ri, (_, row) in enumerate(execs.iterrows(), 3):
            acct     = str(row.get("account", ""))
            is_fill  = str(row.get("execution_type", "")) in ("FILL", "PARTIAL_FILL")
            if acct == acct1:
                bg = HEX["yellow"] if is_fill else "FFFDE7"
            elif acct == acct2:
                bg = "C5E8FF" if is_fill else "EBF5FF"
            else:
                bg = HEX["white"]

            for ci, (_, col_key, _) in enumerate(exec_headers, 1):
                val = row.get(col_key, "")
                if col_key == "transact_time" and hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d %H:%M:%S")
                elif col_key == "price" and val is not None:
                    try: val = round(float(val), 4)
                    except Exception: pass
                elif col_key == "aggressor" and val is not None:
                    val = "Yes" if val else "No"
                c = ws2.cell(ri, ci, val if val is not None else "")
                c.fill      = _fill(bg)
                c.font      = _font(size=10)
                c.border    = _border()
                c.alignment = _align("left" if ci == 1 else "center")
            ws2.row_dimensions[ri].height = 14

        ws2.freeze_panes = "A3"

        # Color key
        leg = len(execs) + 5
        ws2.cell(leg, 1, "Color key:").font = _font(bold=True, size=9)
        for ci, (lbl, bg) in enumerate([
            (f"{acct1} order", "FFFDE7"),
            (f"{acct1} fill",  HEX["yellow"]),
            (f"{acct2} order", "EBF5FF"),
            (f"{acct2} fill",  "C5E8FF"),
        ], 2):
            c = ws2.cell(leg, ci, lbl)
            c.fill      = _fill(bg)
            c.font      = _font(size=9)
            c.border    = _border()
            c.alignment = _align("center")
            ws2.column_dimensions[get_column_letter(ci)].width = max(
                ws2.column_dimensions[get_column_letter(ci)].width, len(lbl) + 4
            )

    # ── Sheet 3: Position Summary ──
    positions = context.get("positions", {})
    bo        = context.get("bo", {})
    ws3 = wb.create_sheet("Position & BO")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 20
    ws3.column_dimensions["E"].width = 20

    ws3.merge_cells("A1:E1")
    t3 = ws3["A1"]
    t3.value     = "Position Summary & Beneficial Ownership"
    t3.fill      = _fill(HEX["navy"])
    t3.font      = _font(bold=True, color=HEX["white"], size=12)
    t3.alignment = _align("center")
    ws3.row_dimensions[1].height = 22

    # Position headers
    pos_hdrs = ["Account", "Bought", "Sold", "Net", "Realized P&L"]
    for ci, h in enumerate(pos_hdrs, 1):
        c = ws3.cell(2, ci, h)
        c.fill      = _fill(HEX["navy2"])
        c.font      = _font(bold=True, color=HEX["white"], size=10)
        c.alignment = _align("center")
        c.border    = _border()
    ws3.row_dimensions[2].height = 16

    for ri, acct in enumerate([acct1, acct2], 3):
        p      = positions.get(acct, {})
        bought = p.get("bought", 0)
        sold   = p.get("sold", 0)
        pnl    = p.get("pnl", 0.0)
        vals   = [acct, bought, sold, bought - sold,
                  f"${pnl:,.2f}" if pnl >= 0 else f"-${abs(pnl):,.2f}"]
        bg = HEX["alt_row"] if ri % 2 == 0 else HEX["white"]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(ri, ci, val)
            c.fill      = _fill(bg)
            c.font      = _font(size=10)
            c.border    = _border()
            c.alignment = _align("center" if ci > 1 else "left")
        ws3.row_dimensions[ri].height = 15

    # BO section
    bo_start = 6
    ws3.merge_cells(f"A{bo_start}:E{bo_start}")
    bh = ws3.cell(bo_start, 1, "Beneficial Ownership")
    bh.fill      = _fill(HEX["navy2"])
    bh.font      = _font(bold=True, color=HEX["white"], size=11)
    bh.alignment = _align("center")
    ws3.row_dimensions[bo_start].height = 18

    bo_hdrs = ["Account", "Legal Name", "State", "ZIP", "Firm"]
    for ci, h in enumerate(bo_hdrs, 1):
        c = ws3.cell(bo_start + 1, ci, h)
        c.fill      = _fill(HEX["section_bg"])
        c.font      = _font(bold=True, size=10)
        c.border    = _border()
        c.alignment = _align("center")

    bo_matches = context.get("bo_matches", [])
    for ri, acct in enumerate([acct1, acct2], bo_start + 2):
        a = bo.get(acct, {})
        vals = [
            acct,
            a.get("legal_name", "N/A"),
            a.get("state", "N/A"),
            a.get("zip", "N/A"),
            a.get("firm", "N/A"),
        ]
        # Highlight row red if there's a BO match
        bg = HEX["red"] if bo_matches else (HEX["alt_row"] if ri % 2 == 0 else HEX["white"])
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(ri, ci, val)
            c.fill      = _fill(bg)
            c.font      = _font(size=10)
            c.border    = _border()
            c.alignment = _align("center" if ci > 1 else "left")
        ws3.row_dimensions[ri].height = 15

    if bo_matches:
        match_row = bo_start + 4
        ws3.merge_cells(f"A{match_row}:E{match_row}")
        mc = ws3.cell(match_row, 1,
                      f"⚠ MATCH IDENTIFIED: {', '.join(bo_matches)}")
        mc.fill      = _fill(HEX["red"])
        mc.font      = _font(bold=True, size=11)
        mc.alignment = _align("center")
        ws3.row_dimensions[match_row].height = 18

    # ── Sheet 3: Matched Trades (between the two accounts only) ──
    matched_pairs = context.get("matched_pairs", pd.DataFrame())
    execs_df      = context.get("execs", pd.DataFrame())
    ws_mt = wb.create_sheet("Matched Trades")
    ws_mt.sheet_view.showGridLines = False

    ws_mt.merge_cells("A1:J1")
    th = ws_mt["A1"]
    th.value     = f"Matched Trades — {acct1} / {acct2} — {symbol}"
    th.fill      = _fill(HEX["navy"])
    th.font      = _font(bold=True, color=HEX["white"], size=12)
    th.alignment = _align("center")
    ws_mt.row_dimensions[1].height = 22

    if matched_pairs.empty:
        ws_mt["A3"] = "No direct matched trades found between these accounts."
        ws_mt["A3"].font = _font(italic=True, color="888888")
    else:
        mt_hdrs = [
            ("Time (UTC)",        "pair_time",               22),
            ("Qty",               "quantity",                  8),
            ("Aggressor Acct",    "aggressor_account",        18),
            ("Aggressor Side",    "aggressor_symbol_subtype",  14),
            ("Passive Acct",      "passive_account",          18),
            ("Passive Side",      "passive_symbol_subtype",   14),
            ("YES Price",         "yes_price",                10),
            ("NO Price",          "no_price",                 10),
            ("Pair ID",           "pair_id",                  28),
        ]
        ws_mt.merge_cells(f"A1:{get_column_letter(len(mt_hdrs))}1")
        for ci, (hdr, _, w) in enumerate(mt_hdrs, 1):
            c = ws_mt.cell(2, ci, hdr)
            c.fill = _fill(HEX["navy2"]); c.font = _font(bold=True, color=HEX["white"], size=10)
            c.alignment = _align("center"); c.border = _border()
            ws_mt.column_dimensions[get_column_letter(ci)].width = w
        ws_mt.row_dimensions[2].height = 16

        total_qty_matched = 0
        for ri, (_, row) in enumerate(matched_pairs.iterrows(), 3):
            qty = int(row.get("quantity", 0))
            total_qty_matched += qty
            bg = HEX["yellow"] if str(row.get("aggressor_account","")) == acct1 else "C5E8FF"
            for ci, (_, col, _) in enumerate(mt_hdrs, 1):
                val = row.get(col, "")
                if col == "pair_time" and hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d %H:%M:%S")
                elif col in ("yes_price", "no_price") and val is not None:
                    try: val = round(float(val), 4)
                    except: pass
                c = ws_mt.cell(ri, ci, val if val is not None else "")
                c.fill = _fill(bg); c.font = _font(size=10)
                c.border = _border(); c.alignment = _align("center" if ci > 1 else "left")
            ws_mt.row_dimensions[ri].height = 14

        # Summary row
        sum_row = len(matched_pairs) + 4
        ws_mt.cell(sum_row, 1, "Total matched qty:").font = _font(bold=True, size=10)
        ws_mt.cell(sum_row, 2, total_qty_matched).font    = _font(bold=True, size=10)
        ws_mt.freeze_panes = "A3"

    # ── Sheet 4: Market Activity (all trades in instrument on alert date) ──
    market_pairs = context.get("market_pairs", pd.DataFrame())
    ws_mkt = wb.create_sheet("Market Activity")
    ws_mkt.sheet_view.showGridLines = False

    ws_mkt.merge_cells("A1:J1")
    mh = ws_mkt["A1"]
    mh.value     = f"All Market Activity — {symbol} — {context.get('alert_date','')}"
    mh.fill      = _fill(HEX["navy"])
    mh.font      = _font(bold=True, color=HEX["white"], size=12)
    mh.alignment = _align("center")
    ws_mkt.row_dimensions[1].height = 22

    if market_pairs.empty:
        ws_mkt["A3"] = "No market activity data available."
        ws_mkt["A3"].font = _font(italic=True, color="888888")
    else:
        mkt_hdrs = [
            ("Time (UTC)",        "pair_time",               22),
            ("Qty",               "quantity",                  8),
            ("Aggressor Acct",    "aggressor_account",        18),
            ("Aggressor Side",    "aggressor_symbol_subtype",  14),
            ("Aggressor Firm",    "aggressor_firm",           18),
            ("Passive Acct",      "passive_account",          18),
            ("Passive Side",      "passive_symbol_subtype",   14),
            ("Passive Firm",      "passive_firm",             18),
            ("YES Price",         "yes_price",                10),
            ("NO Price",          "no_price",                 10),
        ]
        ws_mkt.merge_cells(f"A1:{get_column_letter(len(mkt_hdrs))}1")
        for ci, (hdr, _, w) in enumerate(mkt_hdrs, 1):
            c = ws_mkt.cell(2, ci, hdr)
            c.fill = _fill(HEX["navy2"]); c.font = _font(bold=True, color=HEX["white"], size=10)
            c.alignment = _align("center"); c.border = _border()
            ws_mkt.column_dimensions[get_column_letter(ci)].width = w
        ws_mkt.row_dimensions[2].height = 16

        involved = {acct1, acct2}
        for ri, (_, row) in enumerate(market_pairs.iterrows(), 3):
            agg = str(row.get("aggressor_account", ""))
            pas = str(row.get("passive_account", ""))
            if agg in involved and pas in involved:
                bg = HEX["red"]       # both accounts involved — cross trade
            elif agg in involved or pas in involved:
                bg = HEX["yellow"]    # one of the accounts involved
            else:
                bg = HEX["white"]
            for ci, (_, col, _) in enumerate(mkt_hdrs, 1):
                val = row.get(col, "")
                if col == "pair_time" and hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d %H:%M:%S")
                elif col in ("yes_price", "no_price") and val is not None:
                    try: val = round(float(val), 4)
                    except: pass
                c = ws_mkt.cell(ri, ci, val if val is not None else "")
                c.fill = _fill(bg); c.font = _font(size=10)
                c.border = _border(); c.alignment = _align("center" if ci > 1 else "left")
            ws_mkt.row_dimensions[ri].height = 14

        # Color key
        leg = len(market_pairs) + 4
        ws_mkt.cell(leg, 1, "Color key:").font = _font(bold=True, size=9)
        for ci, (lbl, bg) in enumerate([
            ("Both accounts traded (cross)", HEX["red"]),
            ("One account involved",         HEX["yellow"]),
            ("Third-party trades",           HEX["white"]),
        ], 2):
            c = ws_mkt.cell(leg, ci, lbl)
            c.fill = _fill(bg); c.font = _font(size=9)
            c.border = _border(); c.alignment = _align("center")
        ws_mkt.freeze_panes = "A3"

    # ── Sheet 5: Prior Activity (last 30 days) ──
    prior_activity = context.get("prior_activity", pd.DataFrame())
    ws_prior = wb.create_sheet("Prior Activity")
    ws_prior.sheet_view.showGridLines = False

    ws_prior.merge_cells("A1:J1")
    ph = ws_prior["A1"]
    ph.value     = f"Prior Activity (30 Days) — {acct1} / {acct2}"
    ph.fill      = _fill(HEX["navy"])
    ph.font      = _font(bold=True, color=HEX["white"], size=12)
    ph.alignment = _align("center")
    ws_prior.row_dimensions[1].height = 22

    if prior_activity.empty:
        ws_prior["A3"] = "No prior activity found in the last 30 days for this instrument/source agency."
        ws_prior["A3"].font = _font(italic=True, color="888888")
    else:
        pr_hdrs = [
            ("Account",       "account",        18),
            ("Date",          "trading_date",   12),
            ("Instrument",    "instrument_id",  22),
            ("Source Agency", "source_agency",  28),
            ("Side",          "symbol_subtype",  8),
            ("Exec Type",     "execution_type",  16),
            ("Fill Qty",      "fill_quantity",   10),
            ("Price",         "price",           10),
            ("Time (UTC)",    "transact_time",   22),
        ]
        ws_prior.merge_cells(f"A1:{get_column_letter(len(pr_hdrs))}1")
        for ci, (hdr, _, w) in enumerate(pr_hdrs, 1):
            c = ws_prior.cell(2, ci, hdr)
            c.fill = _fill(HEX["navy2"]); c.font = _font(bold=True, color=HEX["white"], size=10)
            c.alignment = _align("center"); c.border = _border()
            ws_prior.column_dimensions[get_column_letter(ci)].width = w
        ws_prior.row_dimensions[2].height = 16

        for ri, (_, row) in enumerate(prior_activity.iterrows(), 3):
            acct = str(row.get("account", ""))
            bg = HEX["yellow"] if acct == acct1 else "C5E8FF"
            for ci, (_, col, _) in enumerate(pr_hdrs, 1):
                val = row.get(col, "")
                if col == "transact_time" and hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d %H:%M:%S")
                elif col == "trading_date" and hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                elif col == "price" and val is not None:
                    try: val = round(float(val), 4)
                    except: pass
                c = ws_prior.cell(ri, ci, val if val is not None else "")
                c.fill = _fill(bg); c.font = _font(size=10)
                c.border = _border(); c.alignment = _align("center" if ci > 1 else "left")
            ws_prior.row_dimensions[ri].height = 14

        # Color key
        leg = len(prior_activity) + 4
        ws_prior.cell(leg, 1, "Color key:").font = _font(bold=True, size=9)
        for ci, (lbl, bg) in enumerate([
            (acct1, HEX["yellow"]),
            (acct2, "C5E8FF"),
        ], 2):
            c = ws_prior.cell(leg, ci, lbl)
            c.fill = _fill(bg); c.font = _font(size=9)
            c.border = _border(); c.alignment = _align("center")
        ws_prior.freeze_panes = "A3"

    # ── Sheet 6: Position & BO (enhanced with per-side cost basis) ──
    positions = context.get("positions", {})
    bo        = context.get("bo", {})
    ws_pos = wb.create_sheet("Position & BO")
    ws_pos.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [18, 12, 12, 12, 14, 14, 20]):
        ws_pos.column_dimensions[col].width = w

    ws_pos.merge_cells("A1:G1")
    t3 = ws_pos["A1"]
    t3.value = "Position Summary & Beneficial Ownership"
    t3.fill = _fill(HEX["navy"]); t3.font = _font(bold=True, color=HEX["white"], size=12)
    t3.alignment = _align("center"); ws_pos.row_dimensions[1].height = 22

    pos_hdrs = ["Account", "Bought", "Sold", "Net", "Net P&L", "Combined P&L", "Notes"]
    for ci, h in enumerate(pos_hdrs, 1):
        c = ws_pos.cell(2, ci, h)
        c.fill = _fill(HEX["navy2"]); c.font = _font(bold=True, color=HEX["white"], size=10)
        c.alignment = _align("center"); c.border = _border()
    ws_pos.row_dimensions[2].height = 16

    combined_pnl = sum(positions.get(a, {}).get("pnl", 0) for a in [acct1, acct2])
    for ri, acct in enumerate([acct1, acct2], 3):
        p      = positions.get(acct, {})
        bought = p.get("bought", 0)
        sold   = p.get("sold", 0)
        net    = bought - sold
        pnl    = p.get("pnl", 0.0)
        pnl_str = f"${pnl:,.2f}" if pnl >= 0 else f"-${abs(pnl):,.2f}"
        combined_str = f"${combined_pnl:,.2f}" if ri == 3 else ""
        note_str = "Exactly flat" if bought == sold else (
            "Near flat" if abs(bought-sold) <= max(bought,sold)*0.10 else "Directional"
        )
        bg = HEX["alt_row"] if ri % 2 == 0 else HEX["white"]
        for ci, val in enumerate([acct, bought, sold, net, pnl_str, combined_str, note_str], 1):
            c = ws_pos.cell(ri, ci, val)
            c.fill = _fill(bg); c.font = _font(size=10)
            c.border = _border(); c.alignment = _align("center" if ci > 1 else "left")
        ws_pos.row_dimensions[ri].height = 15

    # Per-side cost basis detail
    ws_pos.merge_cells("A6:G6")
    sd_hdr = ws_pos["A6"]
    sd_hdr.value = "Per-Side Cost Basis"
    sd_hdr.fill = _fill(HEX["navy2"]); sd_hdr.font = _font(bold=True, color=HEX["white"], size=11)
    sd_hdr.alignment = _align("center"); ws_pos.row_dimensions[6].height = 18

    side_hdrs = ["Account", "Side", "Qty", "Cost Paid", "Cash Received", "Net P&L", "Avg Cost/Contract"]
    for ci, h in enumerate(side_hdrs, 1):
        c = ws_pos.cell(7, ci, h)
        c.fill = _fill(HEX["section_bg"]); c.font = _font(bold=True, size=10)
        c.border = _border(); c.alignment = _align("center")
    ws_pos.row_dimensions[7].height = 16

    ri = 8
    for acct in [acct1, acct2]:
        sides = positions.get(acct, {}).get("sides", {})
        for side in ("YES", "NO"):
            s = sides.get(side)
            if not s:
                continue
            qty     = s.get("qty", 0)
            cost    = s.get("cost", 0.0)
            recv    = s.get("received", 0.0)
            net     = s.get("net_pnl", 0.0)
            ac      = s.get("avg_cost")
            bg = HEX["alt_row"] if ri % 2 == 0 else HEX["white"]
            vals = [
                acct, side, qty,
                f"${cost:,.2f}",
                f"${recv:,.2f}",
                f"${net:,.2f}" if net >= 0 else f"-${abs(net):,.2f}",
                f"${ac:.4f}" if ac is not None else "N/A",
            ]
            for ci, val in enumerate(vals, 1):
                c = ws_pos.cell(ri, ci, val)
                c.fill = _fill(bg); c.font = _font(size=10)
                c.border = _border(); c.alignment = _align("center" if ci > 1 else "left")
            ws_pos.row_dimensions[ri].height = 15
            ri += 1

    # BO section
    bo_start = ri + 1
    ws_pos.merge_cells(f"A{bo_start}:G{bo_start}")
    bh = ws_pos[f"A{bo_start}"]
    bh.value = "Beneficial Ownership"
    bh.fill = _fill(HEX["navy2"]); bh.font = _font(bold=True, color=HEX["white"], size=11)
    bh.alignment = _align("center"); ws_pos.row_dimensions[bo_start].height = 18

    bo_hdrs = ["Account", "Legal Name", "State", "ZIP", "Firm", "", ""]
    for ci, h in enumerate(bo_hdrs, 1):
        c = ws_pos.cell(bo_start + 1, ci, h)
        c.fill = _fill(HEX["section_bg"]); c.font = _font(bold=True, size=10)
        c.border = _border(); c.alignment = _align("center")

    bo_matches = context.get("bo_matches", [])
    for ri2, acct in enumerate([acct1, acct2], bo_start + 2):
        a = bo.get(acct, {})
        vals = [acct, a.get("legal_name","N/A"), a.get("state","N/A"),
                a.get("zip","N/A"), a.get("firm","N/A"), "", ""]
        bg = HEX["red"] if bo_matches else (HEX["alt_row"] if ri2 % 2 == 0 else HEX["white"])
        for ci, val in enumerate(vals, 1):
            c = ws_pos.cell(ri2, ci, val)
            c.fill = _fill(bg); c.font = _font(size=10)
            c.border = _border(); c.alignment = _align("center" if ci > 1 else "left")
        ws_pos.row_dimensions[ri2].height = 15

    if bo_matches:
        match_row = bo_start + 4
        ws_pos.merge_cells(f"A{match_row}:G{match_row}")
        mc = ws_pos[f"A{match_row}"]
        mc.value = f"⚠ MATCH IDENTIFIED: {', '.join(bo_matches)}"
        mc.fill = _fill(HEX["red"]); mc.font = _font(bold=True, size=11)
        mc.alignment = _align("center"); ws_pos.row_dimensions[match_row].height = 18

    filepath = REVIEWS_DIR / filename
    wb.save(filepath)
    log.info(f"  Report saved: {filename}")
    return filepath


# =============================================================================
# STATE MANAGEMENT
# =============================================================================

def load_state() -> dict:
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return {"processed_alert_ids": [], "last_run": None}


def prune_state(conn, processed_ids: set) -> set:
    """Remove IDs older than the lookback window from state file."""
    if not processed_ids:
        return processed_ids
    try:
        cur = conn.cursor()
        id_list = list(processed_ids)
        cur.execute("""
            SELECT alert_id FROM alerts
            WHERE alert_id = ANY(%s)
              AND alert_time >= NOW() - INTERVAL %s
        """, (id_list, f"{LOOKBACK_HOURS} hours"))
        still_valid = {row[0] for row in cur.fetchall()}
        pruned = processed_ids - still_valid
        if pruned:
            log.info(f"  Pruned {len(pruned)} expired alert ID(s) from state file.")
        return still_valid
    except Exception as e:
        log.warning(f"  State pruning skipped: {e}")
        return processed_ids


def save_state(state: dict):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2, default=str)


# =============================================================================
# POLL CYCLE
# =============================================================================

def run_once(conn, processed_ids: set) -> set:
    alerts = fetch_new_cross_trading_alerts(conn, processed_ids)
    if not alerts:
        log.info("  No new alerts.")
        return processed_ids

    log.info(f"  Found {len(alerts)} new alert(s).")

    for alert in alerts:
        alert_id = alert["alert_id"]
        symbol   = alert.get("symbol", "")
        accounts = extract_all_accounts(alert.get("accounts", []))
        log.info(f"  Processing {alert_id} — {'/'.join(accounts)} — {symbol}")

        try:
            classification, note, context = classify_alert(alert, conn)
            log.info(f"    -> {classification}")

            if classification in ("MM_CONFIRMED", "CRITERIA_1_BUG"):
                log.info(f"    Silent dismiss — no report generated.")
            else:
                write_alert_report(alert, classification, note, context)

            processed_ids.add(alert_id)

        except Exception as e:
            log.error(f"  Error on {alert_id}: {e}\n{traceback.format_exc()}")

    return processed_ids


# =============================================================================
# MAIN
# =============================================================================

def main():
    log.info("=" * 65)
    log.info("ForecastEx Cross Trading Alert Monitor")
    log.info(f"Output  : {OUTPUT_DIR}")
    log.info(f"Interval: every {POLL_INTERVAL_SECONDS // 60} minutes")
    log.info("Press Ctrl+C to stop.")
    log.info("=" * 65)

    state         = load_state()
    processed_ids = set(state.get("processed_alert_ids", []))

    try:
        conn = get_conn()
        conn.set_session(readonly=True)
        log.info("Database connection established.")
    except Exception as e:
        log.error(f"Failed to connect to database: {e}")
        return

    while True:
        log.info(f"--- Poll: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---")
        try:
            processed_ids = run_once(conn, processed_ids)
        except Exception as e:
            log.error(f"Poll cycle error: {e}\n{traceback.format_exc()}")
            # Reconnect on DB errors
            try:
                conn.close()
            except Exception:
                pass
            try:
                conn = get_conn()
                conn.set_session(readonly=True)
                log.info("  Reconnected to database.")
            except Exception as e2:
                log.error(f"  Reconnect failed: {e2}")

        processed_ids = prune_state(conn, processed_ids)
        state["processed_alert_ids"] = list(processed_ids)
        state["last_run"] = datetime.now().isoformat()
        save_state(state)

        log.info(f"Sleeping {POLL_INTERVAL_SECONDS // 60} min...")
        try:
            time.sleep(POLL_INTERVAL_SECONDS)
        except KeyboardInterrupt:
            log.info("Stopped by user.")
            break

    try:
        conn.close()
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except Exception as _e:
        _crash_path = OUTPUT_DIR / "cross_trading_monitor_crash.txt"
        try:
            _crash_path.write_text(
                datetime.now().isoformat() + "\n" + traceback.format_exc(),
                encoding="utf-8"
            )
            print(f"CRASH — details written to: {_crash_path}")
        except Exception:
            print(f"CRASH: {_e}")
    finally:
        input("\nPress Enter to close...")
