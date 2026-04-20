"""
=============================================================================
FORECASTEX LARGE TRADER REPORT DISCREPANCY PROCESSOR
=============================================================================
Scans an inbox folder for EP3 LTR Discrepancy exports (Excel, CSV),
classifies each alert row, and writes one output Excel file per source file.

USAGE:
    python ltr_discrepancy_processor.py

    Drop an EP3 export into the Inbox folder, then run.
    Processed source files are renamed with "_completed" suffix.

FOLDERS:
    Inbox:   ~/Documents/ForecastEx Compliance/LTR Discrepancy/Inbox/
    Reports: ~/Documents/ForecastEx Compliance/LTR Discrepancy/Reports/

CLASSIFICATION (evaluated in order):
    OMNIBUS_DISMISS       — account is a known omnibus account
    LATE_NETTING_DISMISS  — first appearance (no match in prior 14 days of
                            Reports files) AND EP3 > LTR
    INQUIRY_REQUIRED      — anything else
    PARSE_ERROR           — description could not be parsed

AUDIT TRAIL COLUMNS:
    Omnibus Account       — YES if account is in OMNIBUS_ACCOUNTS list
    First Appearance      — YES if no matching (account, symbol, subtype)
                            found in Reports files from the last 14 days
    Nettable              — YES if EOD position holds both YES and NO sides
    EP3 > LTR             — YES if EP3 long position exceeds LTR long position
    LTR Matches Netted    — YES if LTR qty equals calculated netted qty
                            (informational only — not used in classification)
=============================================================================
"""

import os
import re
import sys
import logging
import traceback
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import psycopg2
import psycopg2.extras
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Windows UTF-8 console safety
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# =============================================================================
# CONFIGURATION
# =============================================================================

DB_CONFIG = {
    "host":     "web-prod-db.cla246ai8gve.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "dbname":   "postgres",
    "user":     "postgres_read_only",
    "password": os.environ.get("FXDB_PASSWORD", "pgreadonlypw"),
}

# Known omnibus accounts — add new accounts here as identified
OMNIBUS_ACCOUNTS = {
    "RH0000001749",
}

# How many calendar days of Reports history to scan for prior appearances
PRIOR_APPEARANCE_WINDOW_DAYS = 14


def _find_documents() -> Path:
    return Path.home() / "Documents"

BASE_DIR    = _find_documents() / "ForecastEx Compliance" / "LTR Discrepancy"
INBOX_DIR   = BASE_DIR / "Inbox"
REPORTS_DIR = BASE_DIR / "Reports"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# =============================================================================
# REGEX PATTERNS
# =============================================================================

DESC_PATTERN = re.compile(
    r'account\s+"([^"]+)",\s+symbol\s+"([^"]+)".*?'
    r'subtype\s+"([^"]+)"\s+was\s+Long\s+(\d+)\s*/\s*Short\s+(\d+),\s+'
    r'while the corresponding LTR position was\s+Long\s+(\d+)\s*/\s*Short\s+(\d+)',
    re.IGNORECASE | re.DOTALL,
)
DATE_PATTERN = re.compile(r'\.LTR\.(\d{8})\.')

# Pattern to extract the run date from a Report filename:
# e.g. LTR_Discrepancy_Report_Large_Trader_Report_Discrepancy_2026-03-24_20260324_0401.xlsx
REPORT_DATE_PATTERN = re.compile(r'_(\d{8})_\d{4}\.xlsx$')


# =============================================================================
# PARSERS
# =============================================================================

def parse_description(description: str):
    """Returns (account, symbol, subtype, ep3_long, ep3_short, ltr_long, ltr_short) or None."""
    m = DESC_PATTERN.search(description or "")
    if not m:
        return None
    return (
        m.group(1),
        m.group(2),
        m.group(3).upper(),
        int(m.group(4)),
        int(m.group(5)),
        int(m.group(6)),
        int(m.group(7)),
    )


def parse_report_date(metadata: str):
    """Extract report date from metadata string like {file_name:510.LTR.20260323.txt,...}"""
    if not metadata:
        return None
    m = DATE_PATTERN.search(str(metadata))
    if m:
        s = m.group(1)
        return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    return None


# =============================================================================
# FILE LOADING
# =============================================================================

def load_alert_file(path: Path) -> pd.DataFrame:
    """Load EP3 export — supports .xlsx, .xls, .csv, .txt (tab-sep)."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        df = pd.read_excel(path, dtype=str)
    elif suffix == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        df = pd.read_csv(path, sep="\t", dtype=str)
    df.columns = [c.strip() for c in df.columns]
    return df


def find_inbox_files() -> list:
    INBOX_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    files = []
    for ext in ("*.xlsx", "*.xls", "*.csv", "*.txt"):
        for f in INBOX_DIR.glob(ext):
            if "_completed" not in f.stem:
                files.append(f)
    return sorted(files)


# =============================================================================
# PRIOR APPEARANCE — FILE-BASED (14-DAY ROLLING WINDOW)
# =============================================================================

def load_prior_appearances() -> set:
    """
    Scans Reports folder for output files generated in the last
    PRIOR_APPEARANCE_WINDOW_DAYS days. Returns a set of
    (account, symbol, subtype) tuples found in those files.

    Uses the run-date embedded in the output filename
    (e.g. ..._20260324_0401.xlsx → 2026-03-24) to determine age.
    Falls back to file modification time if the pattern doesn't match.
    """
    cutoff = date.today() - timedelta(days=PRIOR_APPEARANCE_WINDOW_DAYS)
    prior  = set()

    for f in REPORTS_DIR.glob("LTR_Discrepancy_Report_*.xlsx"):
        # Determine the file's run date
        m = REPORT_DATE_PATTERN.search(f.name)
        if m:
            s        = m.group(1)
            file_date = date(int(s[:4]), int(s[4:6]), int(s[6:8]))
        else:
            # Fall back to file modification time
            file_date = date.fromtimestamp(f.stat().st_mtime)

        if file_date < cutoff:
            continue  # Too old — skip

        try:
            # Row 0 is the summary banner, row 1 is the header
            df = pd.read_excel(f, header=1, dtype=str)
            if "Account" not in df.columns or "Symbol" not in df.columns or "Subtype" not in df.columns:
                continue
            for _, row in df.iterrows():
                # Only dismissed-as-inquiry rows count as prior appearances.
                # Dismissed (late netting / omnibus) rows do not block
                # first-appearance on subsequent days.
                if str(row.get("Classification", "")).strip() != "INQUIRY_REQUIRED":
                    continue
                acct    = str(row.get("Account",  "") or "").strip()
                sym     = str(row.get("Symbol",   "") or "").strip()
                subtype = str(row.get("Subtype",  "") or "").strip().upper()
                if acct and sym and subtype:
                    prior.add((acct, sym, subtype))
        except Exception as e:
            log.warning(f"  Could not read prior report {f.name}: {e}")

    log.info(f"  Prior appearances loaded: {len(prior)} unique (account, symbol, subtype) combos "
             f"from Reports files within last {PRIOR_APPEARANCE_WINDOW_DAYS} days.")
    return prior


# =============================================================================
# DB QUERY — EOD POSITIONS ONLY
# =============================================================================

def batch_eod_positions(conn, lookups: list) -> dict:
    """
    lookups: list of (account, symbol, report_date)
    Returns: dict keyed (account, symbol, report_date) -> {"YES": qty, "NO": qty}
    """
    if not lookups:
        return {}

    placeholders = ",".join(["(%s, %s, %s::date)"] * len(lookups))
    flat_params  = []
    for account, symbol, rd in lookups:
        flat_params.extend([account, symbol, rd])

    sql = f"""
        WITH lookup_keys AS (
            SELECT * FROM (VALUES {placeholders}) AS t(account, instrument_id, report_date)
        ),
        ranked AS (
            SELECT
                pl.account,
                pl.instrument_id,
                pl.symbol_subtype,
                pl.after_net_position,
                lk.report_date,
                ROW_NUMBER() OVER (
                    PARTITION BY pl.account, pl.instrument_id, pl.symbol_subtype, lk.report_date
                    ORDER BY pl.transaction_time DESC
                ) AS rn
            FROM position_ledgers pl
            JOIN lookup_keys lk
                ON  pl.account       = lk.account
                AND pl.instrument_id = lk.instrument_id
                AND pl.transaction_time::date = lk.report_date
        )
        SELECT account, instrument_id, symbol_subtype, after_net_position, report_date
        FROM ranked WHERE rn = 1
    """
    results = defaultdict(lambda: {"YES": 0, "NO": 0})
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, flat_params)
        for row in cur.fetchall():
            key = (row["account"], row["instrument_id"], row["report_date"])
            results[key][row["symbol_subtype"]] = row["after_net_position"]
    return results


# =============================================================================
# CLASSIFICATION
# =============================================================================

def classify_row(row: dict, eod_positions: dict, prior_appearances: set) -> dict:
    result = {
        "account":             "",
        "symbol":              "",
        "subtype":             "",
        "ep3_long":            None,
        "ep3_short":           None,
        "ltr_long":            None,
        "ltr_short":           None,
        "report_date":         None,
        "eod_yes":             None,
        "eod_no":              None,
        "netted_side":         "",
        "netted_qty":          None,
        "is_omnibus":          False,
        "ltr_matches_netted":  False,
        "is_first_appearance": None,
        "is_nettable":         False,
        "ep3_gt_ltr":          False,
        "classification":      "PARSE_ERROR",
        "dismissal_note":      "",
        "inquiry_email":       "",
        "jira_summary":        "",
        "jira_description":    "",
        "parse_error":         "",
    }

    parsed = parse_description(row.get("description", ""))
    if not parsed:
        result["parse_error"] = "Could not parse description"
        return result

    account, symbol, subtype, ep3_long, ep3_short, ltr_long, ltr_short = parsed
    result.update({
        "account":   account,
        "symbol":    symbol,
        "subtype":   subtype,
        "ep3_long":  ep3_long,
        "ep3_short": ep3_short,
        "ltr_long":  ltr_long,
        "ltr_short": ltr_short,
    })

    report_date         = parse_report_date(row.get("metadata", ""))
    result["report_date"] = report_date
    rd_str              = report_date.strftime("%Y-%m-%d") if report_date else "UNKNOWN DATE"
    firm_name           = row.get("firm") or "the Clearing Member"

    # ── Audit trail checks ──────────────────────────────────────────────────

    # 1. Omnibus account
    is_omnibus          = account in OMNIBUS_ACCOUNTS
    result["is_omnibus"] = is_omnibus

    # 2. Prior appearance (file-based, 14-day window)
    is_first            = (account, symbol, subtype) not in prior_appearances
    result["is_first_appearance"] = is_first

    # 3. EP3 > LTR (both YES and NO positions stored as Long in DB)
    ep3_qty             = ep3_long
    ltr_qty             = ltr_long
    ep3_gt_ltr          = ep3_qty > ltr_qty
    result["ep3_gt_ltr"] = ep3_gt_ltr

    # 4. EOD position / nettable
    if report_date:
        pos     = eod_positions.get((account, symbol, report_date), {"YES": 0, "NO": 0})
        eod_yes = pos["YES"]
        eod_no  = pos["NO"]
    else:
        eod_yes = eod_no = 0

    result["eod_yes"]     = eod_yes
    result["eod_no"]      = eod_no
    is_nettable           = eod_yes > 0 and eod_no > 0
    result["is_nettable"] = is_nettable

    # 5. LTR matches netted qty (informational — not used in classification)
    if is_nettable:
        if eod_yes >= eod_no:
            netted_side, netted_qty = "YES", eod_yes - eod_no
        else:
            netted_side, netted_qty = "NO",  eod_no  - eod_yes
        result["netted_side"] = netted_side
        result["netted_qty"]  = netted_qty
        if netted_side == subtype and ltr_qty == netted_qty:
            result["ltr_matches_netted"] = True

    # ── Classification (evaluated in priority order) ─────────────────────────

    if is_omnibus:
        result["classification"] = "OMNIBUS_DISMISS"
        result["dismissal_note"] = (
            "Discrepancy due to RH reporting proprietary positions "
            "being held by omnibus account."
        )

    elif is_first and ep3_gt_ltr:
        result["classification"] = "LATE_NETTING_DISMISS"
        result["dismissal_note"] = (
            "Discrepancy due to late netting report. "
            "No action necessary unless discrepancy continues on future dates."
        )

    else:
        result["classification"] = "INQUIRY_REQUIRED"
        result["inquiry_email"]  = (
            f"We noticed the following discrepancy in your {rd_str} LTR submission. "
            f"On your LTR, {account} is listed as having a {ltr_qty} contract {subtype} position "
            f"in {symbol}, but our records show this account with a {ep3_qty} contract {subtype} "
            f"position in that contract. Can you please explain this discrepancy?"
        )
        result["jira_summary"]     = f"LTR Discrepancy Inquiry — {account} / {symbol} / {rd_str}"
        result["jira_description"] = (
            f"Alert ID: {row.get('id', '')}\n"
            f"Account: {account}\nSymbol: {symbol}\nSubtype: {subtype}\n"
            f"Report Date: {rd_str}\n"
            f"EP3 Position: Long {ep3_long} / Short {ep3_short}\n"
            f"LTR Position: Long {ltr_long} / Short {ltr_short}\n\n"
            f"Inquiry sent to {firm_name} asking for explanation for the discrepancy."
        )

    return result


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def build_excel(source_df: pd.DataFrame, results: list, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "LTR Discrepancy"

    GREEN_FILL   = PatternFill("solid", start_color="C6EFCE")
    BLUE_FILL    = PatternFill("solid", start_color="BDD7EE")
    RED_FILL     = PatternFill("solid", start_color="FFC7CE")
    YELLOW_FILL  = PatternFill("solid", start_color="FFEB9C")
    HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    BODY_FONT    = Font(name="Arial", size=9)
    BOLD_FONT    = Font(name="Arial", size=9, bold=True)

    headers = [
        "Alert ID", "Alert Time", "Report Date", "Firm", "Account",
        "Symbol", "Subtype",
        "EP3 Long", "EP3 Short", "LTR Long", "LTR Short",
        "EOD YES", "EOD NO", "Netted Side", "Netted Qty",
        # Audit trail columns
        "Omnibus Account", "First Appearance", "Nettable",
        "EP3 > LTR", "LTR Matches Netted",
        # Result
        "Classification",
        "Dismissal Note",
        "Inquiry Email",
        "Jira Summary", "Jira Description",
        "Parse Error",
    ]
    col_widths = [
        20, 22, 14, 20, 16,
        28, 10,
        10, 10, 10, 10,
        10, 10, 12, 12,
        16, 16, 10,
        10, 18,
        22,
        55,
        80,
        45, 60,
        30,
    ]

    # Classification column index (1-based) — update if headers change
    CLS_COL = headers.index("Classification") + 1

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, (df_row, r) in enumerate(zip(source_df.itertuples(index=False), results), 2):
        cls = r["classification"]
        row_fill = (
            BLUE_FILL   if cls == "OMNIBUS_DISMISS"      else
            GREEN_FILL  if cls == "LATE_NETTING_DISMISS" else
            RED_FILL    if cls == "INQUIRY_REQUIRED"     else
            YELLOW_FILL
        )

        alert_id   = getattr(df_row, "id",        "") or ""
        alert_time = getattr(df_row, "alertTime",  "") or ""
        firm       = getattr(df_row, "firm",        "") or ""

        values = [
            alert_id, alert_time, r["report_date"], firm, r["account"],
            r["symbol"], r["subtype"],
            r["ep3_long"], r["ep3_short"], r["ltr_long"], r["ltr_short"],
            r["eod_yes"], r["eod_no"], r["netted_side"], r["netted_qty"],
            # Audit trail
            "YES" if r["is_omnibus"]          else "NO",
            "YES" if r["is_first_appearance"] else "NO",
            "YES" if r["is_nettable"]         else "NO",
            "YES" if r["ep3_gt_ltr"]          else "NO",
            "YES" if r["ltr_matches_netted"]  else "NO",
            # Result
            cls,
            r["dismissal_note"],
            r["inquiry_email"],
            r["jira_summary"], r["jira_description"],
            r["parse_error"],
        ]

        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.fill      = row_fill
            cell.font      = BOLD_FONT if col == CLS_COL else BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(col > CLS_COL))

        ws.row_dimensions[row_idx].height = 60 if cls == "INQUIRY_REQUIRED" else 18

    # Summary banner
    ws.insert_rows(1)
    omnibus = sum(1 for r in results if r["classification"] == "OMNIBUS_DISMISS")
    dismiss = sum(1 for r in results if r["classification"] == "LATE_NETTING_DISMISS")
    inquiry = sum(1 for r in results if r["classification"] == "INQUIRY_REQUIRED")
    errors  = sum(1 for r in results if r["classification"] == "PARSE_ERROR")

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    c       = ws["A1"]
    c.value = (
        f"Total: {len(results)}  |  Late Netting Dismiss: {dismiss}  |  "
        f"Omnibus Dismiss: {omnibus}  |  Inquiry Required: {inquiry}  |  "
        f"Parse Errors: {errors}  |  Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    c.font      = Font(bold=True, name="Arial", size=10)
    c.fill      = PatternFill("solid", start_color="D9E1F2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    wb.save(output_path)
    return dismiss, omnibus, inquiry, errors


# =============================================================================
# MAIN
# =============================================================================

def process_file(path: Path, conn, prior_appearances: set):
    log.info(f"Loading: {path.name}")
    df = load_alert_file(path)
    log.info(f"  {len(df)} rows found.")

    if df.empty:
        log.warning("  No rows — skipping.")
        return

    # Build EOD position lookup keys
    position_keys = set()
    for _, row in df.iterrows():
        parsed = parse_description(row.get("description", ""))
        rd     = parse_report_date(row.get("metadata", ""))
        if parsed and rd:
            account, symbol = parsed[0], parsed[1]
            position_keys.add((account, symbol, rd))

    log.info(f"  Fetching EOD positions ({len(position_keys)} unique combos)...")
    eod_positions = batch_eod_positions(conn, list(position_keys))

    log.info("  Classifying...")
    results = [classify_row(row.to_dict(), eod_positions, prior_appearances)
               for _, row in df.iterrows()]

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_name = f"LTR_Discrepancy_Report_{path.stem}_{timestamp}.xlsx"
    output_path = REPORTS_DIR / output_name

    dismiss, omnibus, inquiry, errors = build_excel(df, results, output_path)

    log.info(f"  Saved: {output_path.name}")
    log.info(
        f"  Summary — Total: {len(results)} | Late Netting: {dismiss} | "
        f"Omnibus: {omnibus} | Inquiry: {inquiry} | Errors: {errors}"
    )

    # Mark source file as completed
    completed = path.with_stem(path.stem + "_completed")
    path.rename(completed)
    log.info(f"  Source renamed: {completed.name}")


def main():
    print("=" * 56)
    print("  ForecastEx LTR Discrepancy Processor")
    print("=" * 56)
    print()

    files = find_inbox_files()
    if not files:
        print(f"No files found in inbox:")
        print(f"  {INBOX_DIR}")
        print()
        print("Drop an EP3 LTR Discrepancy export into that folder and run again.")
        input("\nPress Enter to close...")
        return

    print(f"Found {len(files)} file(s) to process.")
    print()

    # Load prior appearances once — shared across all files in this run
    print(f"Scanning Reports folder for prior appearances (last {PRIOR_APPEARANCE_WINDOW_DAYS} days)...")
    prior_appearances = load_prior_appearances()
    print()

    print("Connecting to database...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        print("Connected.\n")
    except Exception as e:
        print(f"ERROR: Could not connect to database — {e}")
        input("\nPress Enter to close...")
        sys.exit(1)

    try:
        for path in files:
            try:
                process_file(path, conn, prior_appearances)
            except Exception:
                log.error(f"Failed on {path.name}:\n{traceback.format_exc()}")
    finally:
        conn.close()

    print()
    print("All files processed.")
    print(f"Reports saved to: {REPORTS_DIR}")
    input("\nPress Enter to close...")


if __name__ == "__main__":
    main()
