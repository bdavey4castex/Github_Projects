#!/usr/bin/env python3
"""
NWS Daily Temperature Monitor v3.0
ForecastEx Compliance Tool - Multi-Timezone Edition

Features:
- Per-city local timezone handling
- 48-hour forecast collection (24 hrs before + 24 hrs during target date)
- Daily HIGH and LOW temperature forecasts from 3 sources
- Current temperature (NWS only) for target date
- One Excel tab per city with charts
- Automatic multi-file updates based on city timezones
"""

import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import time
import os
import sys
import hashlib
import logging
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from typing import Optional, Tuple, Dict, List, Set
from dataclasses import dataclass, field
from functools import wraps

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════

# Feature toggles
ENABLE_DAILY_LOW = True  # Daily low temperature tracking
ENABLE_EMAIL_ALERTS = True  # Send email alerts for issues

# File paths
CITIES_FILE = "cities.txt"
OUTPUT_DIR = "daily_highs"
LOG_DIR = "logs"

# API Keys
WEATHERAPI_KEY = os.environ.get("WEATHERAPI_KEY", "d375fd19a77b4cf4940121940262801")

# Email configuration (Gmail)
ALERT_EMAIL_TO = "bdavey@4castex.com"
ALERT_EMAIL_FROM = os.environ.get("ALERT_EMAIL_FROM", "forecastex.alerts@gmail.com")
SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "forecastex.alerts@gmail.com")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "rqfiuyoxgigqxjpd")

# Retry settings
MAX_RETRIES = 3
RETRY_BASE_DELAY = 2
API_DELAY_BETWEEN_CITIES = 1

# Create directories
Path(OUTPUT_DIR).mkdir(exist_ok=True)
Path(LOG_DIR).mkdir(exist_ok=True)

# Logging setup
CT = ZoneInfo("America/Chicago")  # For log timestamps
log_file = Path(LOG_DIR) / f"nws_monitor_{datetime.now(CT).strftime('%Y%m%d')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Excel styling
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
TIMESTAMP_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

# ═══════════════════════════════════════════════════════════════════════════
# DATA STRUCTURES
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class City:
    """City configuration"""
    name: str
    station_id: str
    lat: float
    lon: float
    timezone: str

    def get_tz(self) -> ZoneInfo:
        return ZoneInfo(self.timezone)

    def local_now(self) -> datetime:
        return datetime.now(self.get_tz())

    def local_today(self) -> datetime:
        return self.local_now().date()

    def local_tomorrow(self) -> datetime:
        return self.local_today() + timedelta(days=1)

@dataclass
class TemperatureReading:
    """Temperature reading from a single source"""
    source: str
    high: Optional[int] = None
    low: Optional[int] = None
    current: Optional[int] = None
    error: Optional[str] = None

@dataclass
class CityPullResult:
    """Results from a single pull for a city"""
    city: City
    timestamp_local: datetime
    today_date: datetime  # date object
    tomorrow_date: datetime  # date object
    readings: Dict[str, TemperatureReading] = field(default_factory=dict)

# ═══════════════════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def retry_with_backoff(max_retries: int = MAX_RETRIES, base_delay: float = RETRY_BASE_DELAY):
    """Decorator for retry logic with exponential backoff"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    last_exception = e
                    if attempt < max_retries - 1:
                        delay = base_delay * (2 ** attempt)
                        time.sleep(delay)
            raise last_exception
        return wrapper
    return decorator

def load_cities() -> List[City]:
    """Load cities from cities.txt"""
    if not Path(CITIES_FILE).exists():
        logger.error(f"Cities file not found: {CITIES_FILE}")
        return []

    cities = []
    try:
        with open(CITIES_FILE, "r", encoding="utf-8") as f:
            for line_num, raw in enumerate(f, 1):
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue

                parts = line.split("|")
                if len(parts) == 5:
                    name, station_id, lat, lon, timezone = parts
                    try:
                        cities.append(City(
                            name=name.strip(),
                            station_id=station_id.strip(),
                            lat=float(lat.strip()),
                            lon=float(lon.strip()),
                            timezone=timezone.strip()
                        ))
                    except (ValueError, KeyError) as e:
                        logger.warning(f"Line {line_num}: Invalid data: {line} - {e}")
                else:
                    logger.warning(f"Line {line_num}: Expected 5 fields: {line}")

        logger.info(f"Loaded {len(cities)} cities from {CITIES_FILE}")
        return cities

    except Exception as e:
        logger.error(f"Error loading cities: {e}")
        return []

def get_cities_hash() -> Optional[str]:
    """Get MD5 hash of cities file for change detection"""
    try:
        with open(CITIES_FILE, "rb") as f:
            return hashlib.md5(f.read()).hexdigest()
    except Exception:
        return None

def validate_temperature(temp, min_valid: int = -60, max_valid: int = 150) -> Optional[int]:
    """Validate temperature is within reasonable range"""
    try:
        temp_int = int(round(float(temp)))
        if min_valid <= temp_int <= max_valid:
            return temp_int
        return None
    except (TypeError, ValueError):
        return None

def get_filename_for_date(target_date) -> Path:
    """Get Excel filename for a target date"""
    date_str = target_date.strftime("%Y-%m-%d")
    return Path(OUTPUT_DIR) / f"Daily_Temps_{date_str}.xlsx"

# ═══════════════════════════════════════════════════════════════════════════
# EMAIL ALERTING
# ═══════════════════════════════════════════════════════════════════════════

def send_alert(subject: str, message: str, is_critical: bool = False):
    """Send email alert"""
    if not ENABLE_EMAIL_ALERTS:
        logger.info(f"Alert (email disabled): {subject}")
        return

    if not SMTP_USERNAME or not SMTP_PASSWORD:
        logger.warning(f"Alert (SMTP not configured): {subject}")
        return

    try:
        msg = MIMEMultipart()
        msg['Subject'] = f"{'[CRITICAL] ' if is_critical else ''}[NWS Monitor] {subject}"
        msg['From'] = ALERT_EMAIL_FROM
        msg['To'] = ALERT_EMAIL_TO

        body = f"""
NWS Temperature Monitor Alert
=============================
Time: {datetime.now(CT).strftime('%Y-%m-%d %H:%M:%S CT')}
Priority: {'CRITICAL' if is_critical else 'Normal'}

{message}

---
Automated message from ForecastEx NWS Temperature Monitor.
        """
        msg.attach(MIMEText(body, 'plain'))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)

        logger.info(f"Alert sent: {subject}")

    except Exception as e:
        logger.error(f"Failed to send alert: {e}")

# ═══════════════════════════════════════════════════════════════════════════
# WEATHER API FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

NWS_HEADERS = {
    "User-Agent": "ForecastEx Daily Temp Monitor (compliance@forecastex.com)",
    "Accept": "application/geo+json"
}

@retry_with_backoff()
def fetch_nws(city: City, target_date) -> TemperatureReading:
    """Fetch from NWS API"""
    try:
        # Get grid point
        r = requests.get(
            f"https://api.weather.gov/points/{city.lat},{city.lon}",
            headers=NWS_HEADERS,
            timeout=15
        )
        r.raise_for_status()
        forecast_url = r.json()["properties"]["forecast"]

        # Get forecast
        r2 = requests.get(forecast_url, headers=NWS_HEADERS, timeout=15)
        r2.raise_for_status()
        periods = r2.json()["properties"]["periods"]

        target_weekday = target_date.strftime("%A")
        high_temp = None
        low_temp = None
        current_temp = None

        # First period is current conditions (approximately)
        if periods:
            current_temp = validate_temperature(periods[0]["temperature"])

        for period in periods:
            period_name = period["name"]
            # Match target day for high (daytime) and low (night)
            if target_weekday in period_name:
                if period["isDaytime"]:
                    high_temp = validate_temperature(period["temperature"])
                else:
                    low_temp = validate_temperature(period["temperature"])

        return TemperatureReading(
            source="NWS",
            high=high_temp,
            low=low_temp,
            current=current_temp
        )

    except requests.exceptions.Timeout:
        return TemperatureReading(source="NWS", error="Timeout")
    except requests.exceptions.HTTPError as e:
        return TemperatureReading(source="NWS", error=f"HTTP {e.response.status_code}")
    except Exception as e:
        return TemperatureReading(source="NWS", error=str(e)[:50])

@retry_with_backoff()
def fetch_open_meteo(city: City, target_date) -> TemperatureReading:
    """Fetch from Open-Meteo API"""
    try:
        url = (
            f"https://api.open-meteo.com/v1/forecast"
            f"?latitude={city.lat}&longitude={city.lon}"
            f"&daily=temperature_2m_max,temperature_2m_min"
            f"&current=temperature_2m"
            f"&temperature_unit=fahrenheit"
            f"&timezone={city.timezone}"
            f"&forecast_days=3"
        )

        r = requests.get(url, timeout=15)
        r.raise_for_status()
        data = r.json()

        current_temp = validate_temperature(data.get("current", {}).get("temperature_2m"))

        daily = data.get("daily", {})
        dates = daily.get("time", [])
        highs = daily.get("temperature_2m_max", [])
        lows = daily.get("temperature_2m_min", [])

        target_str = target_date.strftime("%Y-%m-%d")
        high_temp = None
        low_temp = None

        for i, d in enumerate(dates):
            if d == target_str:
                if i < len(highs):
                    high_temp = validate_temperature(highs[i])
                if i < len(lows):
                    low_temp = validate_temperature(lows[i])
                break

        return TemperatureReading(
            source="Open-Meteo",
            high=high_temp,
            low=low_temp,
            current=current_temp
        )

    except requests.exceptions.Timeout:
        return TemperatureReading(source="Open-Meteo", error="Timeout")
    except Exception as e:
        return TemperatureReading(source="Open-Meteo", error=str(e)[:50])

@retry_with_backoff()
def fetch_weatherapi(city: City, target_date) -> TemperatureReading:
    """Fetch from WeatherAPI.com"""
    if WEATHERAPI_KEY == "YOUR_API_KEY_HERE":
        return TemperatureReading(source="WeatherAPI", error="API key not configured")

    try:
        url = (
            f"https://api.weatherapi.com/v1/forecast.json"
            f"?key={WEATHERAPI_KEY}"
            f"&q={city.lat},{city.lon}"
            f"&days=3"
            f"&aqi=no"
        )

        r = requests.get(url, timeout=15)
        r.raise_for_status()
        data = r.json()

        current_temp = validate_temperature(data.get("current", {}).get("temp_f"))

        target_str = target_date.strftime("%Y-%m-%d")
        high_temp = None
        low_temp = None

        for day in data.get("forecast", {}).get("forecastday", []):
            if day.get("date") == target_str:
                high_temp = validate_temperature(day.get("day", {}).get("maxtemp_f"))
                low_temp = validate_temperature(day.get("day", {}).get("mintemp_f"))
                break

        return TemperatureReading(
            source="WeatherAPI",
            high=high_temp,
            low=low_temp,
            current=current_temp
        )

    except requests.exceptions.Timeout:
        return TemperatureReading(source="WeatherAPI", error="Timeout")
    except requests.exceptions.HTTPError as e:
        return TemperatureReading(source="WeatherAPI", error=f"HTTP {e.response.status_code}")
    except Exception as e:
        return TemperatureReading(source="WeatherAPI", error=str(e)[:50])

def fetch_all_sources(city: City, target_date) -> Dict[str, TemperatureReading]:
    """Fetch temperature data from all sources for a specific target date"""
    readings = {}

    readings["NWS"] = fetch_nws(city, target_date)
    time.sleep(API_DELAY_BETWEEN_CITIES / 3)

    readings["Open-Meteo"] = fetch_open_meteo(city, target_date)
    time.sleep(API_DELAY_BETWEEN_CITIES / 3)

    readings["WeatherAPI"] = fetch_weatherapi(city, target_date)

    return readings

# ═══════════════════════════════════════════════════════════════════════════
# EXCEL FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def format_cell(cell, fill=None, font=None, alignment=None, border=True):
    """Apply formatting to a cell"""
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    elif alignment is None:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    if border:
        cell.border = THIN_BORDER

def safe_sheet_name(name: str) -> str:
    """Make a valid Excel sheet name (max 31 chars, no special chars)"""
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '')
    return name[:31]

def get_or_create_workbook(filepath: Path, cities: List[City], target_date) -> openpyxl.Workbook:
    """Load existing workbook or create new one with city tabs"""
    if filepath.exists():
        return load_workbook(filepath)

    # Create new workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active

    # Create a tab for each city
    for i, city in enumerate(cities):
        sheet_name = safe_sheet_name(city.name)
        if i == 0:
            default_sheet.title = sheet_name
            ws = default_sheet
        else:
            ws = wb.create_sheet(sheet_name)

        setup_city_sheet(ws, city, target_date)

    return wb

def setup_city_sheet(ws, city: City, target_date):
    """Set up a city sheet with headers and structure"""
    date_str = target_date.strftime("%Y-%m-%d")

    # Title row
    ws['A1'] = f"{city.name} - Daily Temperature Data"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    ws['A2'] = f"Target Date: {date_str} | Station: {city.station_id} | Timezone: {city.timezone}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:E2')

    # HIGH FORECAST TABLE - starts at row 4
    ws['A4'] = "HIGH TEMPERATURE FORECASTS"
    ws['A4'].font = Font(bold=True, size=12)
    format_cell(ws['A4'], fill=SECTION_FILL, font=Font(bold=True, size=12, color="FFFFFF"))
    ws.merge_cells('A4:E4')

    # High forecast headers - row 5
    high_headers = ["Timestamp (Local)", "NWS High", "Open-Meteo High", "WeatherAPI High"]
    for col, header in enumerate(high_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        format_cell(cell, fill=HEADER_FILL, font=HEADER_FONT)

    # LOW FORECAST TABLE - starts at row 58 (after 48 data rows + headers + spacing)
    ws['A57'] = "LOW TEMPERATURE FORECASTS"
    ws['A57'].font = Font(bold=True, size=12)
    format_cell(ws['A57'], fill=SECTION_FILL, font=Font(bold=True, size=12, color="FFFFFF"))
    ws.merge_cells('A57:E57')

    # Low forecast headers - row 58
    low_headers = ["Timestamp (Local)", "NWS Low", "Open-Meteo Low", "WeatherAPI Low"]
    for col, header in enumerate(low_headers, 1):
        cell = ws.cell(row=58, column=col, value=header)
        format_cell(cell, fill=HEADER_FILL, font=HEADER_FONT)

    # CURRENT TEMPERATURE TABLE - starts at row 110
    ws['A110'] = "CURRENT TEMPERATURE (NWS)"
    ws['A110'].font = Font(bold=True, size=12)
    format_cell(ws['A110'], fill=SECTION_FILL, font=Font(bold=True, size=12, color="FFFFFF"))
    ws.merge_cells('A110:B110')

    # Current temp headers - row 111
    current_headers = ["Timestamp (Local)", "NWS Current"]
    for col, header in enumerate(current_headers, 1):
        cell = ws.cell(row=111, column=col, value=header)
        format_cell(cell, fill=HEADER_FILL, font=HEADER_FONT)

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15

def find_next_empty_row(ws, start_row: int, max_rows: int = 50) -> int:
    """Find the next empty row in a section"""
    for row in range(start_row, start_row + max_rows):
        if ws.cell(row=row, column=1).value is None:
            return row
    return start_row + max_rows

def add_forecast_data(ws, timestamp_str: str, readings: Dict[str, TemperatureReading], data_type: str):
    """Add forecast data row to the appropriate table"""
    if data_type == "high":
        start_row = 6  # Data starts after header row 5
        max_rows = 50
    else:  # low
        start_row = 59  # Data starts after header row 58
        max_rows = 50

    row = find_next_empty_row(ws, start_row, max_rows)

    # Timestamp
    cell = ws.cell(row=row, column=1, value=timestamp_str)
    format_cell(cell, fill=TIMESTAMP_FILL)

    # NWS
    nws = readings.get("NWS")
    value = getattr(nws, data_type) if nws else None
    if value is None and nws and nws.error:
        value = nws.error
    cell = ws.cell(row=row, column=2, value=value if value is not None else "N/A")
    format_cell(cell, fill=ERROR_FILL if isinstance(value, str) and value != "N/A" else None)

    # Open-Meteo
    om = readings.get("Open-Meteo")
    value = getattr(om, data_type) if om else None
    if value is None and om and om.error:
        value = om.error
    cell = ws.cell(row=row, column=3, value=value if value is not None else "N/A")
    format_cell(cell, fill=ERROR_FILL if isinstance(value, str) and value != "N/A" else None)

    # WeatherAPI
    wa = readings.get("WeatherAPI")
    value = getattr(wa, data_type) if wa else None
    if value is None and wa and wa.error:
        value = wa.error
    cell = ws.cell(row=row, column=4, value=value if value is not None else "N/A")
    format_cell(cell, fill=ERROR_FILL if isinstance(value, str) and value != "N/A" else None)

def add_current_temp(ws, timestamp_str: str, nws_reading: TemperatureReading):
    """Add current temperature row"""
    start_row = 112  # Data starts after header row 111
    row = find_next_empty_row(ws, start_row, 30)

    # Timestamp
    cell = ws.cell(row=row, column=1, value=timestamp_str)
    format_cell(cell, fill=TIMESTAMP_FILL)

    # NWS Current
    value = nws_reading.current if nws_reading else None
    if value is None and nws_reading and nws_reading.error:
        value = nws_reading.error
    cell = ws.cell(row=row, column=2, value=value if value is not None else "N/A")
    format_cell(cell, fill=ERROR_FILL if isinstance(value, str) and value != "N/A" else None)

def create_charts(ws, city_name: str):
    """Create or update charts for a city sheet"""
    # Remove existing charts
    ws._charts = []

    # HIGH TEMPERATURE CHART
    high_chart = LineChart()
    high_chart.title = f"{city_name} - High Temperature Forecast"
    high_chart.style = 2  # Style 2 has visible gridlines
    high_chart.y_axis.title = "Temperature (°F)"
    high_chart.x_axis.title = "Timestamp"
    high_chart.height = 12
    high_chart.width = 22

    # Ensure axes are visible with labels
    high_chart.y_axis.delete = False
    high_chart.x_axis.delete = False
    high_chart.y_axis.majorUnit = 5  # Gridlines every 5 degrees
    high_chart.x_axis.tickLblPos = "low"
    high_chart.y_axis.tickLblPos = "low"
    high_chart.x_axis.tickLblSkip = 4  # Show x-axis label every 4 hours

    # Legend on right
    high_chart.legend.position = "r"

    # Find how many data rows we have for high temps
    high_data_end = find_next_empty_row(ws, 6, 50) - 1
    if high_data_end >= 6:
        # Data references (columns B, C, D for the 3 sources)
        data = Reference(ws, min_col=2, max_col=4, min_row=5, max_row=high_data_end)
        cats = Reference(ws, min_col=1, min_row=6, max_row=high_data_end)
        high_chart.add_data(data, titles_from_data=True)
        high_chart.set_categories(cats)

        # Add current temp line if we have data
        current_data_end = find_next_empty_row(ws, 112, 30) - 1
        if current_data_end >= 112:
            current_data = Reference(ws, min_col=2, max_col=2, min_row=111, max_row=current_data_end)
            high_chart.add_data(current_data, titles_from_data=True)

    ws.add_chart(high_chart, "G4")

    # LOW TEMPERATURE CHART
    low_chart = LineChart()
    low_chart.title = f"{city_name} - Low Temperature Forecast"
    low_chart.style = 2  # Style 2 has visible gridlines
    low_chart.y_axis.title = "Temperature (°F)"
    low_chart.x_axis.title = "Timestamp"
    low_chart.height = 12
    low_chart.width = 22

    # Ensure axes are visible with labels
    low_chart.y_axis.delete = False
    low_chart.x_axis.delete = False
    low_chart.y_axis.majorUnit = 5  # Gridlines every 5 degrees
    low_chart.x_axis.tickLblPos = "low"
    low_chart.y_axis.tickLblPos = "low"
    low_chart.x_axis.tickLblSkip = 4  # Show x-axis label every 4 hours

    # Legend on right
    low_chart.legend.position = "r"

    # Find how many data rows we have for low temps
    low_data_end = find_next_empty_row(ws, 59, 50) - 1
    if low_data_end >= 59:
        data = Reference(ws, min_col=2, max_col=4, min_row=58, max_row=low_data_end)
        cats = Reference(ws, min_col=1, min_row=59, max_row=low_data_end)
        low_chart.add_data(data, titles_from_data=True)
        low_chart.set_categories(cats)

        # Add current temp line if we have data
        current_data_end = find_next_empty_row(ws, 112, 30) - 1
        if current_data_end >= 112:
            current_data = Reference(ws, min_col=2, max_col=2, min_row=111, max_row=current_data_end)
            low_chart.add_data(current_data, titles_from_data=True)

    ws.add_chart(low_chart, "G27")

def save_workbook_with_retry(wb, filepath: Path, max_attempts: int = 3) -> bool:
    """Save workbook with retry for file locking"""
    for attempt in range(max_attempts):
        try:
            wb.save(filepath)
            return True
        except PermissionError:
            if attempt < max_attempts - 1:
                logger.warning(f"File locked, retry {attempt + 1}/{max_attempts}...")
                time.sleep(2)
            else:
                logger.error(f"Could not save {filepath} - file is locked")
                send_alert(f"File Save Failed", f"Could not save {filepath.name}", is_critical=True)
                return False
    return False

def update_excel_for_city(city: City, target_date, readings: Dict[str, TemperatureReading],
                          is_today: bool, cities: List[City]):
    """Update Excel file for a specific city and target date"""
    filepath = get_filename_for_date(target_date)
    timestamp_local = city.local_now().strftime("%Y-%m-%d %H:%M")

    # Get or create workbook
    wb = get_or_create_workbook(filepath, cities, target_date)

    # Find city sheet
    sheet_name = safe_sheet_name(city.name)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        setup_city_sheet(ws, city, target_date)
    else:
        ws = wb[sheet_name]

    # Add forecast data (HIGH and LOW) - always
    add_forecast_data(ws, timestamp_local, readings, "high")
    if ENABLE_DAILY_LOW:
        add_forecast_data(ws, timestamp_local, readings, "low")

    # Add current temp - only if this is "today" for this city
    if is_today:
        nws_reading = readings.get("NWS")
        if nws_reading:
            add_current_temp(ws, timestamp_local, nws_reading)

    # Update charts
    create_charts(ws, city.name)

    # Save
    save_workbook_with_retry(wb, filepath)

# ═══════════════════════════════════════════════════════════════════════════
# MAIN LOOP
# ═══════════════════════════════════════════════════════════════════════════

def run_update_cycle(cities: List[City]):
    """Run a single update cycle for all cities"""
    logger.info("-" * 60)
    logger.info(f"Starting update cycle at {datetime.now(CT).strftime('%Y-%m-%d %H:%M:%S CT')}")

    # Track which files we're updating
    files_updated: Set[str] = set()

    for city in cities:
        try:
            local_now = city.local_now()
            today_local = city.local_today()
            tomorrow_local = city.local_tomorrow()

            logger.info(f"Processing {city.name} (local time: {local_now.strftime('%Y-%m-%d %H:%M %Z')})")

            # Fetch data for TODAY (local)
            logger.info(f"  Fetching forecasts for {today_local}...")
            today_readings = fetch_all_sources(city, today_local)

            # Log results
            nws = today_readings.get("NWS")
            if nws and nws.high:
                logger.info(f"    Today HIGH: NWS={nws.high}, OM={today_readings.get('Open-Meteo').high}, WA={today_readings.get('WeatherAPI').high}")

            # Update today's file (forecast + current)
            update_excel_for_city(city, today_local, today_readings, is_today=True, cities=cities)
            files_updated.add(str(today_local))

            # Fetch data for TOMORROW (local)
            logger.info(f"  Fetching forecasts for {tomorrow_local}...")
            tomorrow_readings = fetch_all_sources(city, tomorrow_local)

            # Update tomorrow's file (forecast only)
            update_excel_for_city(city, tomorrow_local, tomorrow_readings, is_today=False, cities=cities)
            files_updated.add(str(tomorrow_local))

            time.sleep(API_DELAY_BETWEEN_CITIES)

        except Exception as e:
            logger.error(f"Error processing {city.name}: {e}")
            send_alert(f"Error: {city.name}", str(e))

    logger.info(f"Update cycle complete. Files updated: {', '.join(sorted(files_updated))}")

def main():
    """Main monitoring loop"""
    logger.info("=" * 70)
    logger.info("ForecastEx NWS Daily Temperature Monitor v3.0")
    logger.info("Multi-Timezone Edition")
    logger.info("=" * 70)
    logger.info(f"Features:")
    logger.info(f"  - Per-city local timezone handling")
    logger.info(f"  - Daily HIGH forecasts: YES")
    logger.info(f"  - Daily LOW forecasts: {'YES' if ENABLE_DAILY_LOW else 'NO'}")
    logger.info(f"  - Current temperature (NWS): YES")
    logger.info(f"  - Data sources: NWS, Open-Meteo, WeatherAPI")
    logger.info(f"  - Email alerts: {'YES' if ENABLE_EMAIL_ALERTS else 'NO'}")
    logger.info("=" * 70)

    # Load cities
    cities = load_cities()
    if not cities:
        logger.error("No cities loaded. Exiting.")
        return

    last_hash = get_cities_hash()
    logger.info(f"Monitoring {len(cities)} cities:")
    for city in cities:
        logger.info(f"  - {city.name} ({city.station_id}) - {city.timezone}")

    logger.info(f"Output directory: {OUTPUT_DIR}")
    logger.info(f"Update schedule: :02 past each hour")
    logger.info("Press Ctrl+C to stop")
    logger.info("=" * 70)

    # Send startup notification
    send_alert(
        "Monitor Started",
        f"NWS Temperature Monitor v3.0 has started.\n"
        f"Monitoring {len(cities)} cities with local timezone support."
    )

    previous_cities = cities.copy()

    try:
        while True:
            # Check for cities.txt changes
            current_hash = get_cities_hash()
            if current_hash != last_hash:
                logger.info("Cities file changed - reloading...")
                new_cities = load_cities()
                if new_cities:
                    cities = new_cities
                    previous_cities = cities.copy()
                else:
                    logger.error("Reload failed - keeping previous list")
                    cities = previous_cities
                last_hash = current_hash

            # Run update cycle
            run_update_cycle(cities)

            # Calculate next run time (:02 past each hour)
            now = datetime.now(CT)
            next_run = (now + timedelta(hours=1)).replace(minute=2, second=0, microsecond=0)
            sleep_seconds = (next_run - now).total_seconds()

            logger.info(f"Next update: {next_run.strftime('%Y-%m-%d %H:%M CT')} ({int(sleep_seconds/60)} min)")
            time.sleep(max(sleep_seconds, 0))

    except KeyboardInterrupt:
        logger.info("\n" + "=" * 70)
        logger.info("Monitoring stopped by user")
        logger.info("=" * 70)

    except Exception as e:
        logger.error(f"Fatal error: {e}")
        send_alert("Monitor Crashed", f"Fatal error: {e}", is_critical=True)
        raise

if __name__ == "__main__":
    main()
