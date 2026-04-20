"""
=============================================================================
FORECASTEX UNIFIED ALERT ADJUDICATION PROCESSOR
=============================================================================
Reads an EP3 alert export from the Inbox folder, applies adjudication logic
for all alert types, queries the database for spoofing and UOA-Size context,
and writes a formatted Excel report to the Reports folder.

USAGE:
    python alert_adjudication_processor.py

FOLDER STRUCTURE (auto-detected from any user's Documents):
    ~/Documents/ForecastEx Compliance/Alert Adjudication/
        Inbox/      <- Drop EP3 export here (newest .xlsx auto-detected)
        Reports/    <- Output written here automatically

OUTPUT FILENAME:
    Alert Adjudication YYYY-MM-DD HH-MM.xlsx
    (timestamp = most recent alert_time in the file)

ALERT TYPES HANDLED:
    Unusual Order Activity - Frequency      Canned note by account type
    Unusual Order Activity - Size           Instrument parsed from DB description
    Manipulative Activity - Day to Day - Price Fluctuation
                                            Weather -> weather note
                                            Same-day expiry -> price fluctuation note
                                            Multi-day -> blank
    Cross Trading                           MM two-sided check, criteria analysis,
                                            execution timing/sizing, BO comparison
    Insider Trading - Price Move            Full description + disposition
    Spoofing                                DB query -> spoofing classifier -> note
=============================================================================
"""

import os
import re
import sys
import json
import logging
import traceback
from datetime import datetime, timezone, date
from pathlib import Path

import psycopg2
import psycopg2.extras
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURATION
# =============================================================================

DB_CONFIG = {
    "host":     "web-prod-db.cla246ai8gve.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "dbname":   "postgres",
    "user":     "postgres_read_only",
    "password": os.environ.get("FXDB_PASSWORD", "pgreadonlypw"),
}

# Folder structure — works for any Windows user
BASE_DIR   = Path.home() / "Documents" / "ForecastEx Compliance" / "Alert Adjudication"
INBOX_DIR  = BASE_DIR / "Inbox"
REPORT_DIR = BASE_DIR / "Reports"

INBOX_DIR.mkdir(parents=True, exist_ok=True)
REPORT_DIR.mkdir(parents=True, exist_ok=True)

# Known market maker / LP accounts
MARKET_MAKERS  = {"FCX001", "RRA001"}
LQ_PROVIDERS   = {"UIBGIECTC", "UIBGIECTMC"}
ALL_MM_ACCOUNTS = MARKET_MAKERS | LQ_PROVIDERS

# Weather symbol prefixes
WEATHER_PREFIXES = ("DH", "DL", "UH", "UL")

# Insider Trading - Price Move thresholds by product category
# (category, category_l2) -> (profit_threshold, timing_days, timing_minutes, label)
# timing_days = None means use timing_minutes instead; both None = analyst reviews timing
INSIDER_THRESHOLDS = {
    # Weather
    ("Environmental", "Temperatures"):         (5000,  None, 5,    "weather"),
    ("Environmental", "Weather"):              (5000,  None, 5,    "weather"),
    # Economic
    ("Economic Indicators", None):             (500,   14,   None, "economic"),
    # Government
    ("Government", None):                      (500,   14,   None, "government"),
    # Environmental / non-weather (climate, agriculture) — lowest applicable = economic
    ("Environmental", None):                   (500,   14,   None, "economic"),
    # Elections — national
    ("Elections", "United States Presidential"): (10000, 5,  None, "elections - national"),
    ("Elections", "United States Congress"):     (10000, 5,  None, "elections - national"),
    # Elections — state (Senate uses lower threshold per policy)
    ("Elections", "United States Senate"):       (5000,  5,  None, "elections - state"),
    ("Elections", "United States Gubernatorial"):(5000,  5,  None, "elections - state"),
    ("Elections", "United States House"):        (5000,  5,  None, "elections - state"),
    # Elections — local
    ("Elections", "Mayoral"):                    (1000,  5,  None, "elections - local"),
    # Other elections (foreign, honors, broad, etc.) — lowest applicable = local
    ("Elections", None):                         (1000,  5,  None, "elections - local"),
    ("Elected Officials Control", None):         (1000,  5,  None, "elections - local"),
    ("Elected Officials Resign", None):          (1000,  5,  None, "elections - local"),
    # Crypto
    ("Financial Markets", "Cryptocurrency"):     (10000, 1,  None, "crypto"),
    # Broad-based indexes
    ("Financial Markets", "Stock Market Indices"):(20000, 7,  None, "broad-based index"),
    ("Financial Markets", "Index Performance"):  (20000, 7,  None, "broad-based index"),
    # Other financials
    ("Financial Markets", None):                 (10000, 1,  None, "other financials"),
    # Sports — timing is complex (before game start), leave for analyst
    ("Sports", None):                            (10000, None, None, "sports"),
    # Technology / other — lowest applicable
    ("Technology", None):                        (500,   14,  None, "economic"),
}

# Fallback threshold when category can't be determined
INSIDER_THRESHOLD_DEFAULT = (500, 14, None, "unknown")

# Keep for backward compat in weather check
WEATHER_INSIDER_PROFIT_THRESHOLD = 5000.0


def get_insider_threshold(conn, symbol: str) -> tuple:
    """
    Look up product category from DB and return the applicable
    (profit_threshold, timing_days, timing_minutes, label) tuple.
    Falls back to lowest threshold if category not found.
    """
    if not conn or not symbol:
        return INSIDER_THRESHOLD_DEFAULT
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT p.category, p.category_l2
            FROM instrument_definitions id
            JOIN products p ON id.product_id = p.product_id
            WHERE id.instrument_id = %s
            LIMIT 1
        """, (symbol,))
        row = cur.fetchone()
        if not row:
            return INSIDER_THRESHOLD_DEFAULT
        cat  = row["category"]
        cat2 = row["category_l2"]

        # Try exact match first, then category-only fallback
        key_exact   = (cat, cat2)
        key_general = (cat, None)
        return (
            INSIDER_THRESHOLDS.get(key_exact)
            or INSIDER_THRESHOLDS.get(key_general)
            or INSIDER_THRESHOLD_DEFAULT
        )
    except Exception as e:
        log.warning(f"  get_insider_threshold error: {e}")
        return INSIDER_THRESHOLD_DEFAULT

# Spoofing classification thresholds (must match spoofing_alert_monitor.py)
GTC_PRICE_DISTANCE_THRESHOLD    = 0.05
NON_COMPETITIVE_PRICE_THRESHOLD = 0.05
RESTING_DURATION_HOURS          = 4


# =============================================================================
# LOGGING
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# =============================================================================
# FILE HELPERS
# =============================================================================

def find_inbox_file() -> Path:
    """Find the most recently modified .xlsx in the Inbox, ignoring _completed files."""
    files = sorted(
        [p for p in INBOX_DIR.glob("*.xlsx") if "_completed" not in p.stem],
        key=lambda p: p.stat().st_mtime,
        reverse=True
    )
    if not files:
        log.error(f"No unprocessed .xlsx files found in {INBOX_DIR}")
        log.error("(Files ending in _completed are skipped.)")
        sys.exit(1)
    log.info(f"  Found inbox file: {files[0].name}")
    return files[0]


def mark_inbox_complete(path: Path) -> None:
    """Rename inbox file to filename_completed.xlsx to prevent re-processing."""
    completed_path = path.with_stem(path.stem + "_completed")
    try:
        path.rename(completed_path)
        log.info(f"  Inbox file renamed to: {completed_path.name}")
    except Exception as e:
        log.warning(f"  Could not rename inbox file: {e}")


def load_inbox(path: Path) -> pd.DataFrame:
    """Load EP3 export. Accepts either xlsx or csv.
    Handles both Carlos-style adjudicated format and raw EP3 export format.
    """
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path, engine="openpyxl")

    # Normalize column names (strip whitespace)
    df.columns = [c.strip() for c in df.columns]

    # --- Detect and remap raw EP3 export columns ---
    # Raw EP3 uses: alertName, id, accounts, symbol, notes, alertTime
    # Carlos format uses: Alert Type, Alert ID, Accounts, Symbol, Notes From Admin Tool
    col_map = {
        "alertName":  "Alert Type",
        "id":         "Alert ID",
        "accounts":   "Accounts",
        "symbol":     "Symbol",
        "notes":      "Notes From Admin Tool",
        "alertTime":  "alert_time",
        "description": "description",
        "firm":       "firm",
    }
    # Only remap if we detect raw EP3 format (has alertName but not Alert Type)
    if "alertName" in df.columns and "Alert Type" not in df.columns:
        log.info("  Detected raw EP3 export format — remapping columns.")
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

    # Required columns
    required = {"Alert Type", "Alert ID"}
    missing = required - set(df.columns)
    if missing:
        log.error(f"Missing required columns: {missing}")
        log.error(f"Found columns: {list(df.columns)}")
        sys.exit(1)

    # Optional columns — fill blank if missing
    for col in ["Accounts", "Symbol", "Notes From Admin Tool",
                "First Reviewer", "Second Reviewer", "alert_time", "description"]:
        if col not in df.columns:
            df[col] = ""

    # Log resolved alert count for info but keep all alerts
    if "status" in df.columns:
        resolved = (df["status"] == "ALERT_STATUS_RESOLVED").sum()
        if resolved:
            log.info(f"  Note: {resolved} already-resolved alerts included (kept for report).")

    df = df.fillna("")
    log.info(f"  Loaded {len(df)} rows from {path.name}")
    return df


# =============================================================================
# ACCOUNT EXTRACTION
# =============================================================================

ACCOUNT_RE = re.compile(
    r'\b(FCX\w+|UIBGIECT\w*|RRA\w+|U\d\w*[C]|RH\w+\d)\b',
    re.IGNORECASE
)

def extract_account(raw: str) -> str:
    """Extract clean account ID from 'firms/Robinhood/accounts/RH0001234' or plain ID."""
    if not raw:
        return ""
    raw = str(raw).strip().strip("{}[]")
    # firms/X/accounts/ACCOUNTID
    if "/" in raw:
        return raw.split("/")[-1].strip().strip("{}[]")
    m = ACCOUNT_RE.search(raw)
    return m.group(1) if m else raw.strip().strip("{}[]")


def extract_account_from_description(desc: str) -> str:
    """Pull account ID out of a DB description string."""
    if not desc:
        return ""
    # 'Account firms/Robinhood/accounts/RH0001234 has placed...'
    m = re.search(r'accounts/([^\s/]+)', desc)
    if m:
        return m.group(1)
    m = ACCOUNT_RE.search(desc)
    return m.group(1) if m else ""


def extract_symbol_from_description(desc: str) -> str:
    """Pull symbol out of UOA-Size description: 'in symbol UHLAX_040526_75 with'"""
    if not desc:
        return ""
    m = re.search(r'in symbol\s+(\S+)', desc, re.IGNORECASE)
    return m.group(1).strip() if m else ""


# =============================================================================
# INSTRUMENT DATE HELPERS
# =============================================================================

def symbol_expiry_date(symbol: str) -> date | None:
    """
    Extract expiry date from symbol like JPUSD_040226_159 or UHLAX_033126_69.
    Date portion is MMDDYY (6 digits after first underscore).
    Returns a date object or None if unparseable.
    """
    if not symbol:
        return None
    parts = symbol.split("_")
    if len(parts) < 2:
        return None
    date_part = parts[1]
    if len(date_part) != 6 or not date_part.isdigit():
        return None
    try:
        mm, dd, yy = date_part[:2], date_part[2:4], date_part[4:6]
        return date(2000 + int(yy), int(mm), int(dd))
    except ValueError:
        return None


def is_same_day_expiry(symbol: str, alert_time_str: str, run_date: date) -> bool:
    """
    Returns True if the instrument expires on the same calendar day as run_date.
    alert_time_str is used as fallback reference if run_date is unavailable.
    """
    expiry = symbol_expiry_date(symbol)
    if expiry is None:
        return False
    return expiry == run_date


# =============================================================================
# DATABASE HELPERS
# =============================================================================

def get_db_connection():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        conn.set_session(readonly=True)
        return conn
    except Exception as e:
        log.error(f"DB connection failed: {e}")
        return None


def fetch_db_alert(conn, alert_id: str) -> dict | None:
    """Fetch raw alert row from DB by alert_id."""
    if not conn:
        return None
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT alert_id, alert_name, accounts, symbol, description,
                   alert_time, metadata
            FROM alerts
            WHERE alert_id = %s
            LIMIT 1
        """, (alert_id,))
        return dict(cur.fetchone()) if cur.rowcount or True else None
    except Exception:
        return None


def fetch_spoofing_executions(conn, alert_id: str,
                              fallback_desc: str = "",
                              fallback_account: str = "",
                              fallback_symbol: str = "",
                              fallback_alert_time: str = "now()") -> tuple[dict, pd.DataFrame]:
    """
    For a spoofing alert, parse the description and fetch execution data.
    If the alert is not in the DB (e.g. EP3-only alerts), falls back to
    the description/account/symbol from the EP3 export file.
    Returns (parsed_dict, executions_df).
    """
    if not conn:
        return {}, pd.DataFrame()

    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Try DB first
    cur.execute("""
        SELECT alert_id, accounts, symbol, description, alert_time
        FROM alerts WHERE alert_id = %s LIMIT 1
    """, (alert_id,))
    row = cur.fetchone()

    if row:
        desc       = row["description"] or ""
        symbol     = row["symbol"] or ""
        account    = extract_account(str(row["accounts"]))
        alert_time = row["alert_time"]
    elif fallback_desc or fallback_account or fallback_symbol:
        # Alert not in DB — use EP3 file data
        log.info(f"  Alert {alert_id} not in DB — using EP3 fallback data.")
        desc       = fallback_desc
        symbol     = fallback_symbol
        account    = fallback_account
        alert_time = fallback_alert_time
    else:
        return {}, pd.DataFrame()

    # Parse description for dominant/fill side quantities
    parsed = parse_spoofing_description(desc, account, symbol)

    # Fetch executions
    try:
        cur.execute("""
            SELECT
                symbol_subtype, execution_type, price, time_in_force,
                transact_time, order_quantity, fill_quantity,
                cum_quantity, leaves_quantity, order_id, order_status,
                order_type, aggressor
            FROM executions
            WHERE account = %s
              AND instrument_id = %s
              AND transact_time >= %s::timestamptz - INTERVAL '14 days'
            ORDER BY transact_time
        """, (account, symbol, alert_time))
        rows = cur.fetchall()
        execs = pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception:
        execs = pd.DataFrame()

    return parsed, execs


def split_symbol_root(symbol: str) -> str:
    return symbol.split("_")[0] if symbol else ""


def fetch_uoa_size_description(conn, alert_id: str) -> str:
    """Fetch description from DB for a UOA-Size alert to extract instrument."""
    if not conn:
        return ""
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT description FROM alerts WHERE alert_id = %s LIMIT 1",
            (alert_id,)
        )
        row = cur.fetchone()
        return row[0] or "" if row else ""
    except Exception:
        return ""


# =============================================================================
# SPOOFING DESCRIPTION PARSER
# =============================================================================

_DOM_QTY_RE  = re.compile(r'(\d+[\d,]*)\s+(YES|NO)\s+side.*?dominant', re.IGNORECASE)
_MIN_QTY_RE  = re.compile(r'(\d+[\d,]*)\s+(YES|NO)\s+side.*?(?:minor|fill)', re.IGNORECASE)
_MULT_RE     = re.compile(r'([\d.]+)x\s+imbalance', re.IGNORECASE)
_FILL_SIDE_RE = re.compile(r'fill.*?\b(YES|NO)\b', re.IGNORECASE)
_DOM_SIDE_RE  = re.compile(r'dominant.*?\b(YES|NO)\b', re.IGNORECASE)

# Simpler approach — parse the known EP3 description format:
# "Account X had Y [side] side orders vs Z [side] orders (Nx imbalance). Fill on [side] side."
_DESC_PATTERN = re.compile(
    r'(\d[\d,]*)\s+(YES|NO)\s*\(.*?\)\s*vs\s*(\d[\d,]*)\s+(YES|NO)',
    re.IGNORECASE
)
_FILL_PATTERN = re.compile(r'executions?\s+on\s+(YES|NO)\s+side\s+\((\d[\d,]*)\)', re.IGNORECASE)
_MULT_PATTERN = re.compile(r'\((\d+\.?\d*)x\s+imbalance\)', re.IGNORECASE)


def parse_spoofing_description(desc: str, account: str, symbol: str) -> dict:
    parsed = {
        "account_id":    account,
        "instrument":    symbol,
        "dominant_side": "",
        "fill_side":     "",
        "dominant_qty":  0,
        "minor_qty":     0,
        "minor_side":    "",
        "fill_qty":      0,
        "multiplier":    0.0,
    }
    if not desc:
        return parsed

    # Format 1 (old): "NO side (85000) vs YES side (4884)"
    m = re.search(
        r'(YES|NO)\s+side\s+\((\d[\d,]*)\)\s+vs\s+(YES|NO)\s+side\s+\((\d[\d,]*)\)',
        desc, re.IGNORECASE
    )
    if m:
        parsed["dominant_side"] = m.group(1).upper()
        parsed["dominant_qty"]  = int(m.group(2).replace(",", ""))
        parsed["minor_side"]    = m.group(3).upper()
        parsed["minor_qty"]     = int(m.group(4).replace(",", ""))

    # Format 2 (new): "more active order quantity on the NO side (5555) than the YES side (65)"
    if not parsed["dominant_side"]:
        m2 = re.search(
            r'more active order quantity on the (YES|NO) side \((\d[\d,]*)\) than the (YES|NO) side \((\d[\d,]*)\)',
            desc, re.IGNORECASE
        )
        if m2:
            parsed["dominant_side"] = m2.group(1).upper()
            parsed["dominant_qty"]  = int(m2.group(2).replace(",", ""))
            parsed["minor_side"]    = m2.group(3).upper()
            parsed["minor_qty"]     = int(m2.group(4).replace(",", ""))

    # Format 1 fill: "executions on YES side (65)"
    mf = re.search(
        r'executions?\s+on\s+(YES|NO)\s+side\s+\((\d[\d,]*)\)',
        desc, re.IGNORECASE
    )
    if mf:
        parsed["fill_side"] = mf.group(1).upper()
        parsed["fill_qty"]  = int(mf.group(2).replace(",", ""))

    # Format 2 fill: "filled exclusively on the YES side (65)"
    if not parsed["fill_side"]:
        mf2 = re.search(
            r'filled exclusively on the (YES|NO) side \((\d[\d,]*)\)',
            desc, re.IGNORECASE
        )
        if mf2:
            parsed["fill_side"] = mf2.group(1).upper()
            parsed["fill_qty"]  = int(mf2.group(2).replace(",", ""))

    # multiplier — format 1: "(85x imbalance)"
    mm = re.search(r'\((\d+\.?\d*)x\s+imbalance\)', desc, re.IGNORECASE)
    if mm:
        parsed["multiplier"] = float(mm.group(1))

    # multiplier — format 2: "85.46 times more"
    if not parsed["multiplier"]:
        mm2 = re.search(r'([\d.]+)\s+times more', desc, re.IGNORECASE)
        if mm2:
            parsed["multiplier"] = float(mm2.group(1))

    return parsed


# =============================================================================
# SPOOFING CLASSIFIER  (mirrors spoofing_alert_monitor.py logic)
# =============================================================================

def _dom_new_prices(dom_execs: pd.DataFrame) -> pd.Series:
    return dom_execs[dom_execs["execution_type"] == "NEW"]["price"].astype(float)

def _fill_prices(execs: pd.DataFrame, fill_side: str) -> pd.Series:
    return execs[
        (execs["symbol_subtype"] == fill_side) &
        (execs["execution_type"].isin(["FILL", "PARTIAL_FILL"]))
    ]["price"].astype(float)

def _resting_hours(execs: pd.DataFrame, fill_time) -> float:
    new_ts = execs[execs["execution_type"] == "NEW"]["transact_time"]
    if new_ts.empty or fill_time is None:
        return 0.0
    try:
        delta = pd.to_datetime(fill_time, utc=True) - pd.to_datetime(new_ts.min(), utc=True)
        return max(delta.total_seconds() / 3600, 0.0)
    except Exception:
        return 0.0


def classify_spoofing(parsed: dict, execs: pd.DataFrame) -> tuple[str, str]:
    """
    Mirrors spoofing_alert_monitor.py classify_alert().
    Returns (classification, note).
    """
    acct     = parsed.get("account_id", "")
    symbol   = parsed.get("instrument", "")
    dom_side = parsed.get("dominant_side", "")
    fil_side = parsed.get("fill_side", "")
    dom_qty  = parsed.get("dominant_qty", 0) or 0
    min_qty  = parsed.get("minor_qty", 0) or 0
    min_side = parsed.get("minor_side", "")
    fil_qty  = parsed.get("fill_qty", 0) or 0
    mult     = parsed.get("multiplier", 0)

    # 1 — MARKET MAKER (verify two-sided)
    if acct in ALL_MM_ACCOUNTS:
        if not execs.empty:
            sides_with_fills = execs[
                execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])
            ]["symbol_subtype"].unique()
            is_two_sided = len(set(sides_with_fills)) >= 2
        else:
            is_two_sided = False

        if is_two_sided:
            return "MARKET_MAKER", (
                f"Account {acct} is a known market maker. "
                f"Activity reviewed and found to be compliant with applicable rules. "
                f"Two-sided executions confirmed (fills on both YES and NO sides). "
                f"No indicators of disorderly trading were identified. No further action required."
            )

    if not execs.empty and dom_side and fil_side:
        dom_execs = execs[execs["symbol_subtype"] == dom_side].copy()
        fil_execs = execs[execs["symbol_subtype"] == fil_side].copy()
        dom_px    = _dom_new_prices(dom_execs)
        fil_px    = _fill_prices(execs, fil_side)
        avg_dom   = dom_px.mean() if not dom_px.empty else None
        avg_fil   = fil_px.mean() if not fil_px.empty else None

        # 2 — RESTING BIDS LONG DURATION
        fil_events = fil_execs[fil_execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])]
        if not fil_events.empty:
            first_fill = fil_events["transact_time"].min()
            dom_h = _resting_hours(dom_execs, first_fill)
            fil_h = _resting_hours(fil_execs, first_fill)
            if dom_h >= RESTING_DURATION_HOURS and fil_h >= RESTING_DURATION_HOURS:
                def _hstr(h):
                    return f"{int(h//24)} days" if h >= 48 else f"{h:.1f} hours"
                return "RESTING_BIDS_LONG_DUR", (
                    f"At time of flagged activity, account {acct} had {dom_qty:,} {dom_side} "
                    f"orders vs {min_qty:,} {min_side} orders ({mult}x imbalance) in {symbol}. "
                    f"Dominant-side ({dom_side}) bids had rested in the market unmodified for "
                    f"{_hstr(dom_h)} before the flagged fill-side execution. "
                    f"Fill-side ({fil_side}) bids had rested unmodified for {_hstr(fil_h)}. "
                    f"Leaving bids open on both sides for an extended period without modification "
                    f"is not indicative of spoofing behavior. "
                    f"Exchange does not find this activity problematic."
                )

        # 3 — NON-COMPETITIVE BIDS
        if avg_dom is not None and avg_fil is not None:
            gap        = abs(avg_dom - avg_fil)
            fill_ratio = fil_qty / dom_qty if dom_qty > 0 else 1.0
            if gap >= NON_COMPETITIVE_PRICE_THRESHOLD and fill_ratio <= 0.20:
                return "NON_COMPETITIVE_BIDS", (
                    f"At time of flagged activity, account {acct} had {dom_qty:,} {dom_side} "
                    f"orders vs {min_qty:,} {min_side} orders ({mult}x imbalance) in {symbol}. "
                    f"Dominant-side ({dom_side}) bids were priced an average of ${gap:.2f} "
                    f"away from prevailing {fil_side}-side execution prices "
                    f"(dominant avg: ${avg_dom:.2f}, fill-side avg: ${avg_fil:.2f}). "
                    f"Bids priced this far from the market are not competitive and not consistent "
                    f"with spoofing, which requires a credible threat of execution. "
                    f"No other signs of potential manipulation. "
                    f"Exchange does not find this activity problematic."
                )

        # 4 — GTC RESTING
        if not dom_execs.empty:
            tif_vals   = dom_execs["time_in_force"].unique()
            is_gtc     = all(t in ("GOOD_TILL_TIME", "GOOD_TILL_CANCEL") for t in tif_vals)
            never_fill = not any(
                t in ("FILL", "PARTIAL_FILL") for t in dom_execs["execution_type"].unique()
            )
            gap_gtc   = abs(avg_dom - avg_fil) if avg_dom is not None and avg_fil is not None else None
            price_far = gap_gtc is not None and gap_gtc >= GTC_PRICE_DISTANCE_THRESHOLD

            if is_gtc and never_fill and price_far:
                new_ts = dom_execs[dom_execs["execution_type"] == "NEW"]["transact_time"]
                exp_ts = dom_execs[
                    dom_execs["execution_type"].isin(["EXPIRED", "CANCELED", "DONE_FOR_DAY"])
                ]["transact_time"]
                t_sit = None
                if not new_ts.empty and not exp_ts.empty:
                    t_sit = pd.to_datetime(exp_ts.max(), utc=True) - pd.to_datetime(new_ts.min(), utc=True)
                t_str   = str(t_sit).split(".")[0] if t_sit else "an extended period"
                gap_str = f"${gap_gtc:.2f}" if gap_gtc else "significantly"
                return "GTC_RESTING", (
                    f"Account {acct} had unexecuted GTT/GTC orders totaling {dom_qty:,} "
                    f"{dom_side} contracts in {symbol} that sat on the market for {t_str} "
                    f"and were priced {gap_str} away from prevailing market prices at the time "
                    f"of execution. "
                    f"Inactive orders resting away from the market are not indicative of spoofing. "
                    f"No concerns."
                )

    # 5 — NEEDS REVIEW
    return "NEEDS_REVIEW", (
        f"Alert requires analyst review. "
        f"Account {acct} had {dom_qty:,} {dom_side} orders vs {min_qty:,} {min_side} "
        f"orders ({mult}x imbalance) in {symbol}. "
        f"Fills occurred on {fil_side} side. "
        f"Execution pattern does not match known auto-dismiss scenarios."
    )


# =============================================================================
# PER-ALERT-TYPE ADJUDICATION LOGIC
# =============================================================================

def adjudicate_uoa_frequency(account: str, symbol: str) -> str:
    if account in MARKET_MAKERS:
        return (
            f"{account} is a Susquehanna market making account. Market makers must be able "
            f"to place many orders to provide liquidity. No other signs of manipulative "
            f"activity. Exchange does not find this activity problematic."
        )
    if account in LQ_PROVIDERS:
        return (
            f"{account} is a ForecastEx institutional customer. No other signs of "
            f"manipulative activity. Exchange does not find this activity problematic."
        )
    return (
        f"{account} placed a large number of orders in a 15 minute interval. This high "
        f"frequency of orders is most likely due to the account utilizing "
        f"computer-generated orders. This activity is not considered to be problematic."
    )


def adjudicate_uoa_size(account: str, symbol: str) -> str:
    return (
        f"{account} has placed a large order order in symbol {symbol} with an unusual "
        f"size. This activity is consistent with the account attempting to accumulate a "
        f"position. This activity is not considered to be problematic."
    )


def adjudicate_day_to_day(account: str, symbol: str, run_date: date) -> str:
    # Weather contract — always canned note regardless of expiry
    if symbol.upper().startswith(WEATHER_PREFIXES):
        return (
            f"{symbol} is a weather-based contract and price volatility in weather-based "
            f"contracts expected in hours before expiration. No other signs of manipulative "
            f"activity. Exchange does not find this activity problematic."
        )
    # Same-day expiry — price fluctuation note
    if is_same_day_expiry(symbol, "", run_date):
        return (
            f"{account} represented [X]% of the daily trade volume in instrument {symbol} "
            f"when an unusual change in price was detected. Price fluctuations are "
            f"anticipated in instruments prior to expiration and near the strike price. "
            f"This activity is not considered to be problematic."
        )
    # Multi-day — blank, analyst to review
    return ""


def extract_all_accounts(raw: str) -> list[str]:
    """Extract all account IDs from a postgres array like {RH0037146667,FCX001}."""
    raw = str(raw).strip().strip("{}")
    return [extract_account(a.strip()) for a in raw.split(",") if a.strip()]


# Criteria detection phrases from the alert engine template
_CRITERIA_2_PHRASE = "Multiple trades found between account 1 and account 2"
_CRITERIA_3_PHRASE = "Quantity traded between account 1 and account 2"
_CRITERIA_1_PHRASE = "on one side and Account 2 traded"


def detect_criteria(description: str) -> tuple[bool, bool, bool]:
    """Returns (criteria1, criteria2, criteria3) booleans based on description text."""
    desc = description or ""
    c1 = _CRITERIA_1_PHRASE in desc
    c2 = _CRITERIA_2_PHRASE in desc
    c3 = _CRITERIA_3_PHRASE in desc
    return c1, c2, c3


def fetch_cross_trading_executions(conn, account: str, symbol: str, alert_time) -> pd.DataFrame:
    """Fetch executions for MM two-sided verification on a cross trading alert.
    instrument_id in executions IS the symbol directly (e.g. UHOKC_040626_72).
    """
    if not conn:
        return pd.DataFrame()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT symbol_subtype, execution_type, price, time_in_force,
                   transact_time, order_quantity, fill_quantity,
                   cum_quantity, leaves_quantity, order_id, order_status,
                   order_type, aggressor
            FROM executions
            WHERE account = %s
              AND instrument_id = %s
              AND transact_time >= %s::timestamptz - INTERVAL '1 day'
            ORDER BY transact_time
        """, (account, symbol, alert_time))
        rows = cur.fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception as e:
        log.warning(f"  fetch_cross_trading_executions error: {e}")
        return pd.DataFrame()


def is_acting_as_market_maker(execs: pd.DataFrame) -> bool:
    """Check if account has fills on both sides — Carlos's two-sided verification."""
    if execs.empty:
        return False
    sides_with_fills = execs[
        execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])
    ]["symbol_subtype"].unique()
    return len(set(sides_with_fills)) >= 2


COMMON_SIZES = {10, 25, 50, 100, 250, 500, 1000, 2500, 5000, 10000}


def fetch_executions_both_accounts(conn, acct1: str, acct2: str,
                                    symbol: str, alert_time) -> pd.DataFrame:
    """Fetch executions for both accounts on the flagged instrument."""
    if not conn:
        return pd.DataFrame()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT account, symbol_subtype, execution_type,
                   order_quantity, fill_quantity, cum_quantity, leaves_quantity,
                   price, transact_time, trade_id,
                   order_id, order_status, order_type,
                   aggressor, time_in_force
            FROM executions
            WHERE account = ANY(%s)
              AND instrument_id = %s
              AND transact_time >= %s::timestamptz - INTERVAL '1 day'
            ORDER BY transact_time
        """, ([acct1, acct2], symbol, alert_time))
        rows = cur.fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception as e:
        log.warning(f"  fetch_executions_both_accounts error: {e}")
        return pd.DataFrame()


def analyze_execution_detail(execs: pd.DataFrame,
                              acct1: str, acct2: str) -> dict:
    """
    For each matched trade_id shared between acct1 and acct2:
    - Compute time gap between the NEW orders placed by each account
    - Compare fill quantities for size similarity
    Returns summary stats and auto-dismiss determination.
    """
    result = {
        "matched_trade_count":    0,
        "median_timing_gap_secs": None,
        "pct_timing_under_30s":   0.0,
        "pct_timing_under_5s":    0.0,
        "pct_similar_size":       0.0,
        "can_auto_dismiss":       False,
        "auto_dismiss_reasons":   [],
        "flag_reasons":           [],
    }
    if execs.empty or "trade_id" not in execs.columns:
        result["flag_reasons"].append("no execution data available for timing analysis")
        return result

    import statistics
    fills = execs[execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])].copy()
    news  = execs[execs["execution_type"] == "NEW"].copy()

    acct1_trade_ids = set(fills[fills["account"] == acct1]["trade_id"].dropna())
    acct2_trade_ids = set(fills[fills["account"] == acct2]["trade_id"].dropna())
    matched_ids = acct1_trade_ids & acct2_trade_ids

    if not matched_ids:
        result["flag_reasons"].append("no matched trade IDs found between accounts")
        return result

    result["matched_trade_count"] = len(matched_ids)
    timing_gaps = []
    size_pairs  = []

    for trade_id in matched_ids:
        a1_fills = fills[(fills["account"] == acct1) & (fills["trade_id"] == trade_id)]
        a2_fills = fills[(fills["account"] == acct2) & (fills["trade_id"] == trade_id)]
        a1_qty = int(a1_fills["fill_quantity"].sum()) if not a1_fills.empty else 0
        a2_qty = int(a2_fills["fill_quantity"].sum()) if not a2_fills.empty else 0
        if a1_qty > 0 and a2_qty > 0:
            size_pairs.append((a1_qty, a2_qty))

        fill_time = fills[fills["trade_id"] == trade_id]["transact_time"].min()
        if fill_time is None:
            continue
        fill_time_dt = pd.to_datetime(fill_time, utc=True)

        a1_news = news[(news["account"] == acct1) &
                       (pd.to_datetime(news["transact_time"], utc=True) <= fill_time_dt)]
        a2_news = news[(news["account"] == acct2) &
                       (pd.to_datetime(news["transact_time"], utc=True) <= fill_time_dt)]
        if a1_news.empty or a2_news.empty:
            continue

        a1_t = pd.to_datetime(a1_news["transact_time"], utc=True).max()
        a2_t = pd.to_datetime(a2_news["transact_time"], utc=True).max()
        timing_gaps.append(abs((a1_t - a2_t).total_seconds()))

    if timing_gaps:
        result["median_timing_gap_secs"] = statistics.median(timing_gaps)
        result["pct_timing_under_30s"]   = sum(1 for g in timing_gaps if g < 30) / len(timing_gaps) * 100
        result["pct_timing_under_5s"]    = sum(1 for g in timing_gaps if g < 5)  / len(timing_gaps) * 100

    similar_count = 0
    for q1, q2 in size_pairs:
        if q1 in COMMON_SIZES and q2 in COMMON_SIZES:
            continue
        larger = max(q1, q2)
        if larger > 0 and abs(q1 - q2) / larger <= 0.10:
            similar_count += 1
    if size_pairs:
        result["pct_similar_size"] = similar_count / len(size_pairs) * 100

    flag_reasons    = []
    dismiss_reasons = []

    if timing_gaps:
        median_gap   = result["median_timing_gap_secs"]
        pct_under_5  = result["pct_timing_under_5s"]
        pct_under_30 = result["pct_timing_under_30s"]
        if pct_under_5 > 25:
            flag_reasons.append(f"{pct_under_5:.0f}% of matched orders placed within 5 seconds of each other")
        elif pct_under_30 > 50:
            flag_reasons.append(f"{pct_under_30:.0f}% of matched orders placed within 30 seconds of each other")
        else:
            dismiss_reasons.append(f"median order timing gap {median_gap:.0f}s — orders not placed in close proximity")

    if size_pairs:
        pct_sim = result["pct_similar_size"]
        if pct_sim >= 25:
            flag_reasons.append(f"{pct_sim:.0f}% of matched trades had similar non-standard order sizes")
        else:
            dismiss_reasons.append(f"only {pct_sim:.0f}% of matched trades had similar sizes — not suggestive of coordination")

    result["flag_reasons"]         = flag_reasons
    result["auto_dismiss_reasons"] = dismiss_reasons
    result["can_auto_dismiss"]     = len(flag_reasons) == 0 and len(dismiss_reasons) > 0
    return result


def fetch_cross_trading_analysis(conn, acct1: str, acct2: str,
                                  symbol: str, alert_time) -> dict:
    """
    Pull position_ledgers final state for both accounts to compute:
    - Realized P&L per account
    - Flatness (qty bought vs sold)
    - Volume between accounts vs total volume
    - Prior cross trading alerts for either account
    """
    result = {
        "acct1_pnl": None, "acct1_bought": 0, "acct1_sold": 0,
        "acct2_pnl": None, "acct2_bought": 0, "acct2_sold": 0,
        "qty_between": 0, "acct1_total": 0, "acct2_total": 0,
        "prior_alerts_acct1": 0, "prior_alerts_acct2": 0,
    }
    if not conn:
        return result
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        # Final position state — prefer RESOLUTION rows for accurate cost basis.
        # sort_priority=0 for RESOLUTION, 1 for everything else, so DISTINCT ON
        # picks the RESOLUTION row first without violating ORDER BY requirements.
        cur.execute("""
            SELECT DISTINCT ON (account, symbol_subtype)
                account, symbol_subtype,
                after_realized, after_qty_bought, after_qty_sold,
                before_cost, entry_type,
                CASE WHEN entry_type = 'RESOLUTION' THEN 0 ELSE 1 END AS sort_priority
            FROM position_ledgers
            WHERE account = ANY(%s)
              AND instrument_id = %s
            ORDER BY account, symbol_subtype, sort_priority ASC, transaction_time DESC
        """, ([acct1, acct2], symbol))
        rows = cur.fetchall()

        # Build per-side detail and sum P&L across YES + NO per account
        sides_data = {acct1: {}, acct2: {}}
        for row in rows:
            acct   = row["account"]
            side   = row["symbol_subtype"]
            pnl    = float(row["after_realized"] or 0)
            bought = int(row["after_qty_bought"] or 0)
            sold   = int(row["after_qty_sold"] or 0)
            cost   = abs(float(row["before_cost"] or 0))
            qty    = bought if bought > 0 else sold
            avg_cost = round(cost / qty, 4) if qty > 0 else None
            implied_pnl_per = round(pnl / qty, 4) if qty > 0 else None

            if acct in sides_data:
                sides_data[acct][side] = {
                    "pnl": round(pnl, 2), "qty": qty,
                    "avg_cost": avg_cost,
                    "implied_pnl_per": implied_pnl_per,
                    "resolved": row["entry_type"] == "RESOLUTION",
                }

            if acct == acct1:
                result["acct1_pnl"] = (result["acct1_pnl"] or 0) + pnl
                result["acct1_bought"] = max(result["acct1_bought"], bought)
                result["acct1_sold"]   = max(result["acct1_sold"], sold)
            elif acct == acct2:
                result["acct2_pnl"] = (result["acct2_pnl"] or 0) + pnl
                result["acct2_bought"] = max(result["acct2_bought"], bought)
                result["acct2_sold"]   = max(result["acct2_sold"], sold)

        result["acct1_sides"] = sides_data.get(acct1, {})
        result["acct2_sides"] = sides_data.get(acct2, {})

        result["acct1_total"] = result["acct1_bought"] + result["acct1_sold"]
        result["acct2_total"] = result["acct2_bought"] + result["acct2_sold"]

        # Volume between accounts from executions (fills matching both accounts)
        cur.execute("""
            SELECT SUM(fill_quantity) as qty
            FROM executions
            WHERE account = %s
              AND instrument_id = %s
              AND execution_type IN ('FILL', 'PARTIAL_FILL')
              AND trade_id IN (
                  SELECT trade_id FROM executions
                  WHERE account = %s AND instrument_id = %s
                  AND execution_type IN ('FILL', 'PARTIAL_FILL')
              )
        """, (acct1, symbol, acct2, symbol))
        row = cur.fetchone()
        result["qty_between"] = int(row["qty"] or 0) if row and row["qty"] else 0

        # Prior cross trading alerts for either account
        for acct, key in [(acct1, "prior_alerts_acct1"), (acct2, "prior_alerts_acct2")]:
            cur.execute("""
                SELECT COUNT(*) as cnt FROM alerts
                WHERE alert_name = 'Cross Trading'
                  AND accounts::text ILIKE %s
                  AND alert_time < %s::timestamptz
            """, (f"%{acct}%", alert_time))
            row = cur.fetchone()
            result[key] = int(row["cnt"]) if row else 0

        # Beneficial ownership comparison
        cur.execute("""
            SELECT account_id, legal_name, state, zip, firm
            FROM accounts
            WHERE account_id = ANY(%s)
        """, ([acct1, acct2],))
        acct_rows = {r["account_id"]: dict(r) for r in cur.fetchall()}
        a1 = acct_rows.get(acct1, {})
        a2 = acct_rows.get(acct2, {})

        bo_matches = []
        if a1 and a2:
            # Last name match
            n1 = (a1.get("legal_name") or "").strip().upper()
            n2 = (a2.get("legal_name") or "").strip().upper()
            ln1 = n1.split()[-1] if n1 else ""
            ln2 = n2.split()[-1] if n2 else ""
            if ln1 and ln2 and ln1 == ln2:
                bo_matches.append(f"last name ({a1.get('legal_name')} / {a2.get('legal_name')})")
            # Full name match
            if n1 and n2 and n1 == n2:
                bo_matches.append(f"full name ({a1.get('legal_name')})")
            # ZIP match
            z1 = (a1.get("zip") or "").strip()
            z2 = (a2.get("zip") or "").strip()
            if z1 and z2 and z1 == z2:
                bo_matches.append(f"zip code ({z1})")
            # State match (only flag if also zip matches to avoid noise)
            s1 = (a1.get("state") or "").strip().upper()
            s2 = (a2.get("state") or "").strip().upper()

        result["bo_matches"]  = bo_matches
        result["bo_acct1"]    = a1
        result["bo_acct2"]    = a2

    except Exception as e:
        log.warning(f"  fetch_cross_trading_analysis error: {e}")

    return result


def build_cross_trading_inquiry_note(acct1: str, acct2: str, symbol: str,
                                      criteria_str: str, analysis: dict,
                                      mm_not_acting: str = None) -> str:
    """
    Build a structured review note from DB analysis data.
    Presents findings neutrally — only flags escalation if something looks suspicious.
    """
    lines = []
    red_flags = []

    # Flag if known MM is not acting as one
    if mm_not_acting:
        lines.append(
            f"NOTE: {mm_not_acting} is a known market maker but has not been confirmed "
            f"to be acting as a market maker during this time interval. "
            f"This account should be treated as any other account in this review."
        )

    # Criteria triggered
    lines.append(f"Alert triggered on {criteria_str}.")

    # P&L and flatness — per-side breakdown with avg cost and implied profit/contract
    def _side_detail(sides: dict, red_flags_list: list, acct_label: str) -> str:
        parts = []
        for side in ("YES", "NO"):
            s = sides.get(side)
            if not s:
                continue
            qty  = s.get("qty", 0)
            ac   = s.get("avg_cost")
            sp   = s.get("pnl", 0.0)
            ipp  = s.get("implied_pnl_per")
            res  = s.get("resolved", False)
            if ac is not None and qty > 0:
                pnl_per_str = (
                    f"~${ipp:.2f}/contract profit" if ipp is not None and ipp >= 0
                    else f"~${abs(ipp):.2f}/contract loss" if ipp is not None
                    else ""
                )
                parts.append(
                    f"{side}: {qty:,} contracts, avg cost ${ac:.2f}, "
                    f"realized ${sp:,.2f} ({pnl_per_str})"
                    + (" [RESOLVED]" if res else "")
                )
                if ipp is not None and ipp > 0.60:
                    red_flags_list.append(
                        f"{acct_label} {side}-side profit of ${ipp:.2f}/contract is unusually high "
                        f"(avg cost ${ac:.2f}, suggesting contracts bought at a steep discount)"
                    )
        return "; ".join(parts) if parts else "no side detail available"

    pnl1 = analysis["acct1_pnl"]
    pnl2 = analysis["acct2_pnl"]
    b1, s1 = analysis["acct1_bought"], analysis["acct1_sold"]
    b2, s2 = analysis["acct2_bought"], analysis["acct2_sold"]
    flat1 = abs(b1 - s1) <= max(b1, s1) * 0.10 if max(b1, s1) > 0 else True
    flat2 = abs(b2 - s2) <= max(b2, s2) * 0.10 if max(b2, s2) > 0 else True

    if pnl1 is not None and (b1 or s1):
        pnl1_str = f"${pnl1:,.2f}" if pnl1 >= 0 else f"-${abs(pnl1):,.2f}"
        detail1  = _side_detail(analysis.get("acct1_sides", {}), red_flags, acct1)
        lines.append(
            f"Realized P&L: {acct1}: {b1:,} bought / {s1:,} sold "
            f"({'near flat' if flat1 else 'directional'}) | Total: {pnl1_str}. "
            f"Side breakdown — {detail1}."
        )
        if flat1 and (b1 + s1) > 0:
            red_flags.append(f"{acct1} ended near flat (wash trade indicator)")

    if pnl2 is not None and (b2 or s2):
        pnl2_str = f"${pnl2:,.2f}" if pnl2 >= 0 else f"-${abs(pnl2):,.2f}"
        detail2  = _side_detail(analysis.get("acct2_sides", {}), red_flags, acct2)
        lines.append(
            f"Realized P&L: {acct2}: {b2:,} bought / {s2:,} sold "
            f"({'near flat' if flat2 else 'directional'}) | Total: {pnl2_str}. "
            f"Side breakdown — {detail2}."
        )
        if flat2 and (b2 + s2) > 0:
            red_flags.append(f"{acct2} ended near flat (wash trade indicator)")

    if pnl1 is not None and pnl2 is not None:
        if pnl1 * pnl2 < 0:
            red_flags.append("one account profited while the other had a loss (consistent with money pass)")
        elif abs(pnl1) < 10 and abs(pnl2) < 10:
            red_flags.append("both accounts near flat on P&L (consistent with wash trading)")

    # Volume concentration
    qty_between = analysis["qty_between"]
    total1 = analysis["acct1_total"]
    total2 = analysis["acct2_total"]
    if qty_between and (total1 or total2):
        pct1 = (qty_between / total1 * 100) if total1 else 0
        pct2 = (qty_between / total2 * 100) if total2 else 0
        lines.append(
            f"Volume traded between accounts: {qty_between:,} contracts "
            f"({pct1:.1f}% of {acct1}'s total, {pct2:.1f}% of {acct2}'s total)."
        )
        if pct1 >= 25 or pct2 >= 25:
            red_flags.append(f"high volume concentration between accounts ({max(pct1,pct2):.1f}%)")

    # Beneficial ownership
    bo_matches = analysis.get("bo_matches", [])
    a1_info    = analysis.get("bo_acct1", {})
    a2_info    = analysis.get("bo_acct2", {})
    if a1_info or a2_info:
        a1_name = a1_info.get("legal_name", "N/A")
        a2_name = a2_info.get("legal_name", "N/A")
        a1_loc  = f"{a1_info.get('state','?')}/{a1_info.get('zip','?')}"
        a2_loc  = f"{a2_info.get('state','?')}/{a2_info.get('zip','?')}"
        lines.append(
            f"Beneficial ownership: {acct1}: {a1_name} ({a1_loc}) | "
            f"{acct2}: {a2_name} ({a2_loc})."
        )
    if bo_matches:
        red_flags.append(f"beneficial ownership match on {', '.join(bo_matches)}")
    else:
        lines.append("No matching beneficial ownership fields identified.")

    # Prior alerts
    p1 = analysis.get("prior_alerts_acct1", 0)
    p2 = analysis.get("prior_alerts_acct2", 0)
    if p1 or p2:
        lines.append(f"Prior cross trading alerts: {acct1}: {p1} | {acct2}: {p2}.")
        if p1 > 0 or p2 > 0:
            red_flags.append("prior cross trading alerts on record")

    # Execution timing and sizing analysis
    ea = analysis.get("exec_analysis", {})
    if ea:
        matched  = ea.get("matched_trade_count", 0)
        med_gap  = ea.get("median_timing_gap_secs")
        pct_5s   = ea.get("pct_timing_under_5s", 0)
        pct_30s  = ea.get("pct_timing_under_30s", 0)
        pct_sim  = ea.get("pct_similar_size", 0)
        if matched:
            gap_str = f"{med_gap:.0f}s median" if med_gap is not None else "N/A"
            lines.append(
                f"Execution analysis: {matched} matched trade(s). "
                f"Order timing gap — {gap_str} "
                f"({pct_5s:.0f}% under 5s, {pct_30s:.0f}% under 30s). "
                f"Similar non-standard sizing: {pct_sim:.0f}% of matched trades."
            )
        for reason in ea.get("flag_reasons", []):
            red_flags.append(reason)

    # Conclusion
    all_clear = len(red_flags) == 0 and ea.get("can_auto_dismiss", False) if ea else False

    if red_flags:
        flag_str = "; ".join(red_flags)
        lines.append(
            f"The following factors warrant further review: {flag_str}. "
            f"Analyst should escalate to CRO if pre-arranged trading is suspected."
        )
    elif all_clear:
        dismiss_str = "; ".join(ea.get("auto_dismiss_reasons", []))
        lines.append(
            f"Review of trading activity, beneficial ownership, position data, and "
            f"execution timing and sizing does not suggest pre-arranged trading "
            f"({dismiss_str}). This activity is not considered to be problematic."
        )
    else:
        lines.append(
            "No significant red flags identified in the automated review. "
            "Analyst should confirm before closing."
        )

    return " ".join(lines)


def adjudicate_cross_trading(accounts_raw: str, symbol: str, description: str,
                              conn=None, alert_id: str = "") -> str:
    all_accounts = extract_all_accounts(accounts_raw)
    c1, c2, c3 = detect_criteria(description)

    # --- Criteria 1 bug dismiss (Criteria 1 only, no C2 or C3) ---
    if c1 and not c2 and not c3:
        return (
            "Alert generating to known bug which causes account 1 and account 2's volume "
            "to be compared against each other as opposed to each accounts volume being "
            "compared against itself on the other side of the market. None of the additional "
            "criteria were met and the alert can be dismissed."
        )

    # --- Market maker checks ---
    mm_account = next((a for a in all_accounts if a in MARKET_MAKERS), None)
    lq_account = next((a for a in all_accounts if a in LQ_PROVIDERS), None)
    mm_or_lq   = mm_account or lq_account
    mm_label   = mm_account or lq_account

    if mm_or_lq:
        # Fetch executions to verify two-sided activity
        alert_row = fetch_db_alert(conn, alert_id) if conn and alert_id else {}
        alert_time = (alert_row or {}).get("alert_time", "now()")
        execs = fetch_cross_trading_executions(conn, mm_or_lq, symbol, alert_time)
        two_sided = is_acting_as_market_maker(execs)

        # MM two-sided verified — single confirmed note regardless of which criteria triggered
        if two_sided:
            return (
                f"{mm_label} is a known market maker, has been confirmed to be acting as "
                f"a market maker at this time, and therefore this alert can be dismissed. "
                f"This activity is not considered to be problematic."
            )

        # MM not acting as expected — fall through to INQUIRY_REQUIRED
        # The inquiry note will flag that a known MM is not acting as one

    # --- LQ provider (UIB) standard dismiss if two-sided ---
    if lq_account and not mm_account:
        alert_row  = fetch_db_alert(conn, alert_id) if conn and alert_id else {}
        alert_time = (alert_row or {}).get("alert_time", "now()")
        execs      = fetch_cross_trading_executions(conn, lq_account, symbol, alert_time)
        if is_acting_as_market_maker(execs):
            counterparty = next((a for a in all_accounts if a != lq_account), "")
            with_str = f" with {counterparty}" if counterparty else ""
            sym_str  = f" on instrument {symbol}" if symbol else ""
            return (
                f"{lq_account} is suspected of cross trading{with_str}{sym_str}. "
                f"{lq_account} is a ForecastEx liquidity provider that uses a known algorithm. "
                f"This activity is not considered to be problematic."
            )

    # --- INQUIRY_REQUIRED — pull analysis from DB ---
    acct1 = all_accounts[0] if all_accounts else "Account 1"
    acct2 = all_accounts[1] if len(all_accounts) > 1 else "Account 2"
    criteria_triggered = []
    if c1: criteria_triggered.append("Criteria 1 (similar volumes)")
    if c2: criteria_triggered.append("Criteria 2 (similar order times/quantities)")
    if c3: criteria_triggered.append("Criteria 3 (large % of total volume)")
    criteria_str = ", ".join(criteria_triggered) if criteria_triggered else "unknown criteria"

    alert_row  = fetch_db_alert(conn, alert_id) if conn and alert_id else {}
    alert_time = (alert_row or {}).get("alert_time", "now()")
    analysis   = fetch_cross_trading_analysis(conn, acct1, acct2, symbol, alert_time)

    # Execution timing and sizing analysis
    both_execs   = fetch_executions_both_accounts(conn, acct1, acct2, symbol, alert_time)
    exec_analysis = analyze_execution_detail(both_execs, acct1, acct2)
    analysis["exec_analysis"] = exec_analysis

    # Flag if a known MM is involved but not acting as one
    mm_not_acting = mm_or_lq if mm_or_lq else None

    return build_cross_trading_inquiry_note(
        acct1, acct2, symbol, criteria_str, analysis, mm_not_acting=mm_not_acting
    )


def fetch_insider_executions(conn, account: str, symbol: str) -> pd.DataFrame:
    """Fetch all fills for account on instrument for insider timing analysis."""
    if not conn:
        return pd.DataFrame()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("""
            SELECT execution_type, symbol_subtype, fill_quantity, price, transact_time
            FROM executions
            WHERE account = %s
              AND instrument_id = %s
              AND execution_type IN ('FILL', 'PARTIAL_FILL')
            ORDER BY transact_time
        """, (account, symbol))
        rows = cur.fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception as e:
        log.warning(f"  fetch_insider_executions error: {e}")
        return pd.DataFrame()


def analyze_insider_timing(execs: pd.DataFrame,
                            timing_window_minutes: int = 5) -> dict:
    """
    Detect the price move time from execution data and determine what
    fraction of the position was built within the timing window before it.

    Price move time = timestamp of first fill where price jumped > $0.10
    from the median price of prior fills.

    Returns dict with:
        price_move_time      — detected timestamp of price move
        qty_in_window        — contracts accumulated in timing window
        qty_total            — total contracts accumulated
        pct_in_window        — percentage in window
        within_threshold     — True if majority was within window
        price_before         — median price before move
        price_after          — price at move
        price_jump           — size of jump
        detail               — human-readable summary
    """
    result = {
        "price_move_time":  None,
        "qty_in_window":    0,
        "qty_total":        0,
        "pct_in_window":    0.0,
        "within_threshold": None,
        "price_before":     None,
        "price_after":      None,
        "price_jump":       None,
        "detail":           "Insufficient execution data for timing analysis.",
    }

    if execs.empty:
        return result

    fills = execs[execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])].copy()
    if fills.empty:
        return result

    fills["transact_time"] = pd.to_datetime(fills["transact_time"], utc=True)
    fills["price"]         = fills["price"].astype(float)
    fills["fill_quantity"] = fills["fill_quantity"].astype(int)

    total_qty = int(fills["fill_quantity"].sum())
    result["qty_total"] = total_qty

    # Detect price move: find first fill where price jumped > $0.10 from median of prior fills
    price_move_idx  = None
    price_move_time = None
    price_before    = None
    price_after     = None

    for i in range(1, len(fills)):
        prior_prices = fills["price"].iloc[:i]
        median_prior = float(prior_prices.median())
        current_price = float(fills["price"].iloc[i])
        jump = abs(current_price - median_prior)
        if jump >= 0.10:
            price_move_idx  = i
            price_move_time = fills["transact_time"].iloc[i]
            price_before    = round(median_prior, 4)
            price_after     = round(current_price, 4)
            result["price_jump"] = round(jump, 4)
            break

    if price_move_time is None:
        result["detail"] = (
            "No significant price move detected in execution data "
            "(no fill with price jump >= $0.10). Timing analysis inconclusive."
        )
        return result

    result["price_move_time"] = price_move_time
    result["price_before"]    = price_before
    result["price_after"]     = price_after

    # Count qty accumulated in the timing window before the price move
    window_start = price_move_time - pd.Timedelta(minutes=timing_window_minutes)
    in_window = fills[
        (fills["transact_time"] >= window_start) &
        (fills["transact_time"] < price_move_time)
    ]
    before_window = fills[fills["transact_time"] < window_start]

    qty_in_window    = int(in_window["fill_quantity"].sum())
    qty_before_window = int(before_window["fill_quantity"].sum())
    pct_in_window    = qty_in_window / total_qty * 100 if total_qty > 0 else 0.0

    result["qty_in_window"]    = qty_in_window
    result["pct_in_window"]    = round(pct_in_window, 1)
    result["within_threshold"] = pct_in_window > 50.0

    window_str = f"{timing_window_minutes}-minute"
    move_str   = price_move_time.strftime("%Y-%m-%d %H:%M:%S UTC")

    if qty_before_window > 0 and qty_in_window == 0:
        result["detail"] = (
            f"Price move detected at {move_str} (${price_before:.2f} -> ${price_after:.2f}, "
            f"jump of ${result['price_jump']:.2f}). "
            f"All {total_qty:,} contracts were accumulated more than {timing_window_minutes} minutes "
            f"before the price move. No activity within the {window_str} insider trading window."
        )
    elif qty_in_window == 0:
        result["detail"] = (
            f"Price move detected at {move_str} (${price_before:.2f} -> ${price_after:.2f}). "
            f"No contracts accumulated within the {window_str} window before the price move."
        )
    else:
        result["detail"] = (
            f"Price move detected at {move_str} (${price_before:.2f} -> ${price_after:.2f}, "
            f"jump of ${result['price_jump']:.2f}). "
            f"{qty_in_window:,} of {total_qty:,} contracts ({pct_in_window:.1f}%) were accumulated "
            f"within the {window_str} window before the price move; "
            f"{qty_before_window:,} contracts were accumulated outside the window."
        )

    return result


def adjudicate_insider_price_move(account: str, symbol: str, description: str,
                                   conn=None) -> str:
    """
    Lead with raw description, then append disposition.
    Checks profit threshold and timing window for all product categories
    using thresholds defined in INSIDER_THRESHOLDS.
    """
    desc = description.strip()

    # Inject instrument name after "YES price" or "NO price"
    if symbol:
        for side in ("YES price", "NO price"):
            if side in desc:
                desc = desc.replace(side, f"{side} in instrument {symbol}", 1)
                break

    # Market maker check
    if account in ALL_MM_ACCOUNTS:
        return (
            f"{desc} "
            f"{account} is a known market maker and is acting in accordance with "
            f"expectations for a market maker. Occasional directional profits are "
            f"expected. No concerns."
        )

    # Look up product category and applicable thresholds
    profit_threshold, timing_days, timing_minutes, category_label = (
        get_insider_threshold(conn, symbol)
    )

    # Parse profit from description: "...over the past 14 days of 1234.56"
    profit = None
    m = re.search(r'of\s+([\d,]+\.?\d*)\s*(?:\(|$)', desc)
    if m:
        try:
            profit = float(m.group(1).replace(",", ""))
        except ValueError:
            profit = None

    profit_str     = f"${profit:,.2f}" if profit is not None else "an amount"
    threshold_str  = f"${profit_threshold:,.0f}"
    category_str   = category_label.title()

    # ── Check 1: profit below threshold — dismiss immediately ──
    if profit is not None and profit < profit_threshold:
        return (
            f"{desc} "
            f"This is a {category_str} product. The total profit of {profit_str} "
            f"does not exceed the {threshold_str} profit threshold applicable to "
            f"{category_str} instruments. "
            f"This alert does not meet the criteria for further review and can be dismissed."
        )

    # ── Check 2: profit at or above threshold — run timing analysis ──
    # Determine timing window
    if timing_days is not None:
        window_minutes = timing_days * 24 * 60
        window_label   = f"{timing_days}-day"
    elif timing_minutes is not None:
        window_minutes = timing_minutes
        window_label   = f"{timing_minutes}-minute"
    else:
        # Sports and other categories where timing requires analyst judgment
        return (
            f"{desc} "
            f"This is a {category_str} product. The total profit of {profit_str} "
            f"meets or exceeds the {threshold_str} profit threshold. "
            f"Analyst should review the timing of position accumulation relative to the "
            f"applicable insider trading window for {category_str} products."
        )

    timing = {}
    if conn:
        execs  = fetch_insider_executions(conn, account, symbol)
        timing = analyze_insider_timing(execs, timing_window_minutes=window_minutes)

    if timing.get("price_move_time") is not None:
        pct_in  = timing["pct_in_window"]
        qty_in  = timing["qty_in_window"]
        qty_tot = timing["qty_total"]
        detail  = timing["detail"]

        if not timing.get("within_threshold"):
            return (
                f"{desc} "
                f"This is a {category_str} product. The total profit of {profit_str} "
                f"meets or exceeds the {threshold_str} profit threshold. "
                f"However, timing analysis indicates that the account's position was "
                f"accumulated predominantly outside the {window_label} insider trading "
                f"window applicable to {category_str} products. "
                f"{detail} "
                f"As the position was not accumulated within the timing threshold, "
                f"this alert does not meet the criteria for further review and can be dismissed."
            )
        else:
            return (
                f"{desc} "
                f"This is a {category_str} product. The total profit of {profit_str} "
                f"meets or exceeds the {threshold_str} profit threshold. "
                f"Timing analysis indicates that {pct_in:.1f}% of the account's position "
                f"({qty_in:,} of {qty_tot:,} contracts) was accumulated within the "
                f"{window_label} insider trading window before the price move. "
                f"{detail} "
                f"This alert meets both the profit and timing thresholds and requires analyst review."
            )
    else:
        timing_detail = timing.get("detail", "Timing analysis could not be completed.")
        return (
            f"{desc} "
            f"This is a {category_str} product. The total profit of {profit_str} "
            f"meets or exceeds the {threshold_str} profit threshold. "
            f"{timing_detail} "
            f"Analyst should review the timing of position accumulation relative to the "
            f"{window_label} insider trading window for {category_str} products."
        )


# =============================================================================
# CONSOLIDATION
# =============================================================================

def consolidate_alerts(df: pd.DataFrame, conn, run_date: date) -> list[dict]:
    """
    Group alerts by (Alert Type, Account, Symbol) and adjudicate each group.
    Returns list of output row dicts.
    """
    # We need to enrich UOA-Size rows with instrument from DB before grouping
    log.info("  Enriching UOA-Size instruments from DB...")
    uoa_size_mask = df["Alert Type"] == "Unusual Order Activity - Size"
    for idx, row in df[uoa_size_mask].iterrows():
        if not str(row.get("Symbol", "")).strip():
            # Try first alert ID in the group
            alert_ids = [a.strip() for a in str(row["Alert ID"]).split("\n") if a.strip()]
            for aid in alert_ids[:1]:
                db_desc = fetch_uoa_size_description(conn, aid)
                if db_desc:
                    sym = extract_symbol_from_description(db_desc)
                    if sym:
                        df.at[idx, "Symbol"] = sym
                        break

    # Build grouping key: (Alert Type, Account, Symbol)
    # Account comes from Accounts column
    df["_account"] = df["Accounts"].apply(lambda x: extract_account(str(x)))
    df["_symbol"]  = df["Symbol"].apply(lambda x: str(x).strip())

    results = []

    # Group by alert type + account + symbol
    group_keys = ["Alert Type", "_account", "_symbol"]
    for keys, group in df.groupby(group_keys, sort=False):
        alert_type, account, symbol = keys

        # Collect all alert IDs in this group
        all_ids = []
        for raw_ids in group["Alert ID"]:
            for aid in str(raw_ids).split("\n"):
                aid = aid.strip()
                if aid:
                    all_ids.append(aid)

        # If ANY row in this group already has a note in EP3, keep it — don't overwrite
        existing_notes = [
            str(n).strip()
            for n in group["Notes From Admin Tool"]
            if str(n).strip()
        ]
        if existing_notes:
            results.append({
                "Alert Type":            alert_type,
                "Alert ID":              "\n".join(all_ids),
                "Notes From Admin Tool": existing_notes[0],
                "First Reviewer":        str(group.iloc[0].get("First Reviewer", "")).strip(),
                "Second Reviewer":       str(group.iloc[0].get("Second Reviewer", "")).strip(),
            })
            continue

        # Fetch DB description for cross trading and insider trading
        db_description = ""
        if alert_type in ("Cross Trading", "Insider Trading - Price Move",
                          "Manipulative Activity - Day to Day - Price Fluctuation"):
            db_row = fetch_db_alert(conn, all_ids[0]) if all_ids else None
            db_description = (db_row or {}).get("description", "") or ""

        # --- Adjudicate ---
        if alert_type == "Unusual Order Activity - Frequency":
            note = adjudicate_uoa_frequency(account, symbol)

        elif alert_type == "Unusual Order Activity - Size":
            note = adjudicate_uoa_size(account, symbol)

        elif alert_type == "Manipulative Activity - Day to Day - Price Fluctuation":
            # Use DB description for the note if it's a same-day note
            note = adjudicate_day_to_day(account, symbol, run_date)
            # If same-day and not weather, reconstruct the note from DB description
            if note and not symbol.upper().startswith(WEATHER_PREFIXES) and db_description:
                # Build note from the actual DB description content
                note = _build_day_to_day_note_from_db(account, symbol, db_description)

        elif alert_type == "Cross Trading":
            accounts_raw = str(group.iloc[0].get("Accounts", ""))
            first_aid    = all_ids[0] if all_ids else ""
            note = adjudicate_cross_trading(
                accounts_raw, symbol, db_description, conn=conn, alert_id=first_aid
            )

        elif alert_type == "Insider Trading - Price Move":
            # For grouped alerts, try each ID until we find a description with a profit amount
            ipm_description = db_description
            if not re.search(r'of\s+[\d,]+\.?\d*\s*(?:\(|$)', ipm_description):
                for aid in all_ids[1:]:
                    alt_row = fetch_db_alert(conn, aid) if conn and aid else None
                    alt_desc = (alt_row or {}).get("description", "") or ""
                    if alt_desc and re.search(r'of\s+[\d,]+\.?\d*\s*(?:\(|$)', alt_desc):
                        ipm_description = alt_desc
                        break
            note = adjudicate_insider_price_move(account, symbol, ipm_description, conn=conn)
            # Prepend account and symbol for clarity
            if note:
                note = f"Account: {account} | Symbol: {symbol}. {note}"

        elif alert_type in ("Spoofing", "Real Time Manipulative Activity"):
            # Query DB for each alert ID, classify, use first result
            # Pass EP3 data as fallback in case alert is not in the DB
            ep3_desc        = db_description or str(group.iloc[0].get("description", ""))
            ep3_alert_time  = str(group.iloc[0].get("alert_time", "now()"))
            note = ""
            for aid in all_ids[:1]:
                parsed, execs = fetch_spoofing_executions(
                    conn, aid,
                    fallback_desc=ep3_desc,
                    fallback_account=account,
                    fallback_symbol=symbol,
                    fallback_alert_time=ep3_alert_time,
                )
                if parsed:
                    _, note = classify_spoofing(parsed, execs)
                    break
            if not note:
                note = f"Account: {account} | Symbol: {symbol}. Alert requires analyst review — spoofing classification could not be completed."

        else:
            note = ""

        results.append({
            "Alert Type":           alert_type,
            "Alert ID":             "\n".join(all_ids),
            "Notes From Admin Tool": note,
            "First Reviewer":       "",
            "Second Reviewer":      "",
        })

    return results


def _build_day_to_day_note_from_db(account: str, symbol: str, db_description: str) -> str:
    """
    Reconstruct a Day to Day price fluctuation note from DB description.
    DB description format: "Account X represented Y% of the daily trade volume in
    symbol Z when an unusual change in price ranging from A to B was detected."
    We replicate the output style exactly.
    """
    if not db_description:
        return (
            f"{account} represented [X]% of the daily trade volume in instrument {symbol} "
            f"when an unusual change in price was detected. Price fluctuations are "
            f"anticipated in instruments prior to expiration and near the strike price. "
            f"This activity is not considered to be problematic."
        )

    # Extract percentage
    pct_m = re.search(r'([\d.]+)%\s+of the daily', db_description, re.IGNORECASE)
    pct   = f"{pct_m.group(1)}%" if pct_m else "[X]%"

    # Extract price range
    rng_m = re.search(r'ranging from ([\d.]+) to ([\d.]+)', db_description, re.IGNORECASE)
    rng   = f"ranging from {rng_m.group(1)} to {rng_m.group(2)}" if rng_m else "ranging from [A] to [B]"

    return (
        f"{account} represented {pct} of the daily trade volume in instrument {symbol} "
        f"when an unusual change in price {rng} was detected. Price fluctuations are "
        f"anticipated in instruments prior to expiration and near the strike price. "
        f"This activity is not considered to be problematic."
    )


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

# Colors
NAVY   = "1F3864"
WHITE  = "FFFFFF"
GREY   = "F2F2F2"
BORDER = "D3D3D3"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border():
    s = Side(style="thin", color=BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


COL_WIDTHS = {
    "Alert Type":            28,
    "Alert ID":              22,
    "Notes From Admin Tool": 80,
    "First Reviewer":        16,
    "Second Reviewer":       16,
}


def write_excel(results: list[dict], output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alert Adjudications"

    headers = ["Alert Type", "Alert ID", "Notes From Admin Tool",
               "First Reviewer", "Second Reviewer"]

    # Header row
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = _fill(NAVY)
        cell.font      = _font(bold=True, color=WHITE, size=11)
        cell.border    = _border()
        cell.alignment = _align(h="center", v="center", wrap=True)

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, row in enumerate(results, start=2):
        is_blank_note = not str(row.get("Notes From Admin Tool", "")).strip()
        row_fill = _fill("FFF2CC") if is_blank_note else _fill(WHITE)

        for col_idx, h in enumerate(headers, start=1):
            val  = row.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.font      = _font(size=10)
            cell.border    = _border()
            cell.alignment = _align()

        # Auto row height based on content
        note      = str(row.get("Notes From Admin Tool", ""))
        ids       = str(row.get("Alert ID", ""))
        max_lines = max(
            ids.count("\n") + 1,
            len(note) // 80 + 1,
        )
        ws.row_dimensions[row_idx].height = max(15 * max_lines, 30)

    # Column widths
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(h, 18)

    ws.freeze_panes = "A2"

    wb.save(output_path)
    log.info(f"  Saved: {output_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 65)
    print("  ForecastEx Unified Alert Adjudication Processor")
    print(f"  Run time : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65)

    # Find inbox file
    log.info("Loading inbox file...")
    inbox_file = find_inbox_file()
    df = load_inbox(inbox_file)

    if df.empty:
        log.info("No alerts to process. Exiting.")
        return

    # Determine run date (today) for expiry checks
    run_date = date.today()
    log.info(f"  Run date (for expiry checks): {run_date}")

    # Connect to DB
    log.info("Connecting to database...")
    conn = get_db_connection()
    if conn:
        log.info("  Connected.")
    else:
        log.warning("  DB unavailable — spoofing and UOA-Size instrument enrichment disabled.")

    # Determine output timestamp from most recent alert_time in file
    # Convert to US Eastern time. Fall back to now if no alert_time column or all blank.
    from zoneinfo import ZoneInfo
    eastern = ZoneInfo("America/New_York")
    output_ts = datetime.now(tz=eastern).strftime("%Y-%m-%d %H-%M ET")
    if "alert_time" in df.columns:
        try:
            times = pd.to_datetime(df["alert_time"], errors="coerce", utc=True).dropna()
            if not times.empty:
                latest = times.max().astimezone(eastern)
                output_ts = latest.strftime("%Y-%m-%d %H-%M ET")
        except Exception:
            pass

    # Consolidate and adjudicate
    log.info("Processing alerts...")
    results = consolidate_alerts(df, conn, run_date)
    log.info(f"  Produced {len(results)} output rows from {len(df)} input rows.")

    if conn:
        conn.close()

    # Summary
    from collections import Counter
    type_counts = Counter(r["Alert Type"] for r in results)
    blank_count = sum(1 for r in results if not str(r.get("Notes From Admin Tool", "")).strip())
    print("\n  Alert Type Summary:")
    for atype, cnt in sorted(type_counts.items()):
        print(f"    {atype:<50} {cnt}")
    print(f"\n  Rows with blank notes (analyst review needed): {blank_count}")

    # Write output
    output_filename = f"Alert Adjudication {output_ts}.xlsx"
    output_path = REPORT_DIR / output_filename
    log.info(f"Writing report to {output_path}...")
    write_excel(results, output_path)

    # Rename inbox file so it won't be picked up on next run
    mark_inbox_complete(inbox_file)

    print(f"\n  Done. Report saved to:")
    print(f"  {output_path}")
    input("\n  Press Enter to close...")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        print("\n" + "=" * 65)
        print("  FATAL ERROR:")
        print("=" * 65)
        traceback.print_exc()
        input("\n  Press Enter to close...")
