"""
=============================================================================
FORECASTEX SPOOFING ALERT MONITOR
=============================================================================
Polls public.alerts every 30 minutes for new Spoofing alerts.
For each new alert, fetches execution data, classifies the activity, and
writes a standalone Excel report to the analyst's Documents folder.

USAGE:
    python spoofing_alert_monitor.py

    Runs indefinitely, polling every 30 minutes. Press Ctrl+C to stop.
    Safe to run simultaneously across analyst machines — state is stored
    per-machine in the output folder.

OUTPUT FOLDER:
    ~/Documents/ForecastEx Compliance/Spoofing/

OUTPUT FILES (one per alert):
    Spoofing_{AlertID}_{Account}_{Symbol}.xlsx

EXCEL REPORT SHEETS:
    1. Summary          — alert metadata, classification banner, adjudication verbiage
    2. Activity Chart   — last price + accused bids + best market bids over time
    3. Execution Detail — full execution log, color-coded by side

CLASSIFICATIONS (in priority order):
    MARKET_MAKER            Known MM/LP account — auto-dismiss
    WEATHER_VOLATILITY      DH/UH weather forecast contracts near expiry — auto-dismiss
    RESTING_BIDS_LONG_DUR   Both sides resting >= 4 hours before fill — auto-dismiss
    NON_COMPETITIVE_BIDS    Dominant bids >= $0.05 from fill-side prices — auto-dismiss
    GTC_RESTING             GTT/GTC dominant orders, never filled, priced away — auto-dismiss
    NEEDS_REVIEW            Does not match any known dismissal pattern
=============================================================================
"""

import os
import re
import sys
import json
import time
import logging
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path

import psycopg2
import pandas as pd
from sqlalchemy import create_engine
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel


# =============================================================================
# CONFIGURATION
# =============================================================================

DB_CONFIG = {
    "host":     "web-prod-db.cla246ai8gve.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "database": "postgres",
    "user":     "postgres_read_only",
    "password": os.environ.get("FXDB_PASSWORD", "pgreadonlypw"),
}

# Known market maker / LP accounts — always auto-dismiss
# FCX001 = Susquehanna (market maker)
# UIBGIECTC, UIBGIECTMC = Liquidity providers
# RRA001 = Rockridge (market maker)
MARKET_MAKER_ACCOUNTS = {"FCX001", "UIBGIECTC", "UIBGIECTMC", "RRA001"}

# Price distance thresholds (dollars on a $0.01–$0.99 binary contract)
GTC_PRICE_DISTANCE_THRESHOLD    = 0.05   # GTT/GTC resting orders
NON_COMPETITIVE_PRICE_THRESHOLD = 0.05   # Non-competitive dominant bids

# Resting duration threshold — both sides must have rested this long before any fill
RESTING_DURATION_HOURS = 4

# Weather contract symbol prefixes (DH = daily high, UH = urban high, all city codes)
WEATHER_SYMBOL_PREFIXES = ("DH", "UH")

# How long to sleep between poll cycles
POLL_INTERVAL_SECONDS = 30 * 60   # 30 minutes

# Output folder — created automatically on first run
OUTPUT_DIR  = Path.home() / "Documents" / "ForecastEx Compliance" / "Spoofing"
REVIEWS_DIR = OUTPUT_DIR / "Reviews"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
REVIEWS_DIR.mkdir(parents=True, exist_ok=True)

# State file tracks processed alert IDs so re-runs don't duplicate reports
STATE_FILE = OUTPUT_DIR / "spoofing_monitor_state.json"


# =============================================================================
# LOGGING  (file + console)
# =============================================================================

# Force UTF-8 on Windows console (Python 3.7+)
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(OUTPUT_DIR / "spoofing_monitor.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# =============================================================================
# STYLE CONSTANTS
# =============================================================================

# Hex colors
HEX = {
    "navy":        "1F3864",
    "navy2":       "2E4D7B",
    "white":       "FFFFFF",
    "black":       "000000",
    "border":      "B8CCE4",
    "label_bg":    "EEF4FB",
    "section_bg":  "D6E4F0",
    "alt_row":     "F5F9FF",
    "adj_bg":      "FFFBE6",
    "adj_border":  "E0A800",
    # Classification row colors
    "cls_mm":      "C6EFCE",   # green  — MARKET_MAKER
    "cls_gtc":     "FFEB9C",   # yellow — GTC_RESTING
    "cls_ncb":     "BDD7EE",   # blue   — NON_COMPETITIVE_BIDS
    "cls_wv":      "EDEDED",   # grey   — WEATHER_VOLATILITY
    "cls_rbl":     "D9EAD3",   # mint   — RESTING_BIDS_LONG_DUR
    "cls_nr":      "FFC7CE",   # red    — NEEDS_REVIEW
    # Chart series colors
    "price":       "1F4E79",
    "acct_dom":    "C00000",
    "acct_fil":    "E97132",
    "best_dom":    "375623",
    "best_fil":    "70AD47",
}

CLASS_COLOR = {
    "MARKET_MAKER":          HEX["cls_mm"],
    "GTC_RESTING":           HEX["cls_gtc"],
    "NON_COMPETITIVE_BIDS":  HEX["cls_ncb"],
    "WEATHER_VOLATILITY":    HEX["cls_wv"],
    "RESTING_BIDS_LONG_DUR": HEX["cls_rbl"],
    "NEEDS_REVIEW":          HEX["cls_nr"],
}

CLASS_LABEL = {
    "MARKET_MAKER":          "MARKET MAKER — AUTO-DISMISSED",
    "GTC_RESTING":           "GTC/GTT RESTING ORDERS — AUTO-DISMISSED",
    "NON_COMPETITIVE_BIDS":  "NON-COMPETITIVE BIDS — AUTO-DISMISSED",
    "WEATHER_VOLATILITY":    "WEATHER VOLATILITY — AUTO-DISMISSED",
    "RESTING_BIDS_LONG_DUR": "RESTING BIDS (LONG DURATION) — AUTO-DISMISSED",
    "NEEDS_REVIEW":          "⚠  NEEDS ANALYST REVIEW",
}


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11, italic=False, name="Calibri") -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def _thin_border(color=None) -> Border:
    c = color or HEX["border"]
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _medium_border(color=None) -> Border:
    c = color or HEX["adj_border"]
    s = Side(style="medium", color=c)
    return Border(left=s, right=s, top=s, bottom=s)


# =============================================================================
# DATABASE
# =============================================================================

def get_engine():
    url = (
        f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
        f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
    )
    return create_engine(url)


def fetch_new_spoofing_alerts(engine, processed_ids: set) -> pd.DataFrame:
    """Return unprocessed ALERT_STATUS_NEW Spoofing alerts from the last 14 days."""
    df = pd.read_sql("""
        SELECT id, alert_id, alert_name, status, symbol,
               accounts, firm, description, alert_time, created_at
        FROM public.alerts
        WHERE alert_name = 'Spoofing'
          AND status     = 'ALERT_STATUS_NEW'
          AND created_at >= NOW() - INTERVAL '14 days'
        ORDER BY created_at DESC
        LIMIT 100
    """, engine)
    if processed_ids:
        df = df[~df["alert_id"].isin(processed_ids)]
    log.info(f"Found {len(df)} new spoofing alert(s) to process")
    return df


def fetch_executions(engine, account: str, instrument_id: str) -> pd.DataFrame:
    """Fetch all executions for a given account + instrument (exact, then fuzzy fallback)."""
    cols = """
        execution_id, order_id, instrument_id, symbol_subtype, execution_type,
        order_status, order_quantity, fill_quantity, cum_quantity, leaves_quantity,
        price, aggressor, account, transact_time, trading_date, time_in_force
    """
    for exact in (True, False):
        try:
            if exact:
                df = pd.read_sql(
                    f"SELECT {cols} FROM public.executions "
                    f"WHERE account = %(a)s AND instrument_id = %(i)s "
                    f"ORDER BY transact_time ASC",
                    engine, params={"a": account, "i": instrument_id}
                )
            else:
                df = pd.read_sql(
                    f"SELECT {cols} FROM public.executions "
                    f"WHERE account = %(a)s AND instrument_id ILIKE %(p)s "
                    f"ORDER BY transact_time ASC",
                    engine, params={"a": account, "p": f"%{instrument_id}%"}
                )
            if not df.empty:
                log.info(f"    executions ({'exact' if exact else 'fuzzy'}): {len(df)} rows")
                return df
        except Exception as e:
            log.warning(f"    executions query ({'exact' if exact else 'fuzzy'}) failed: {e}")
    log.info("    executions: 0 rows")
    return pd.DataFrame()


def fetch_market_executions(engine, instrument_id: str,
                            start_time, end_time) -> pd.DataFrame:
    """
    Fetch all market-wide executions for an instrument within a time window.
    Used to derive last-traded price and best bids for the chart.
    """
    try:
        df = pd.read_sql("""
            SELECT execution_id, account, symbol_subtype, execution_type,
                   order_quantity, price, transact_time, time_in_force
            FROM public.executions
            WHERE instrument_id = %(i)s
              AND transact_time BETWEEN %(s)s AND %(e)s
            ORDER BY transact_time ASC
        """, engine, params={"i": instrument_id, "s": start_time, "e": end_time})
        log.info(f"    market executions: {len(df)} rows")
        return df
    except Exception as ex:
        log.warning(f"    market executions query failed: {ex}")
        return pd.DataFrame()


# =============================================================================
# POSITION DETAIL FETCHER
# =============================================================================

def fetch_position_detail(engine, account: str, instrument_id: str) -> dict:
    """
    Fetch per-side P&L and cost basis from position_ledgers for the accused account.
    Prioritises RESOLUTION rows for accurate final cost basis.
    Returns a dict keyed by side (YES/NO) with pnl, qty, avg_cost, implied_pnl_per.
    """
    result = {}
    try:
        df = pd.read_sql("""
            SELECT DISTINCT ON (symbol_subtype)
                symbol_subtype, after_realized, after_qty_bought,
                after_qty_sold, before_cost, entry_type,
                CASE WHEN entry_type = 'RESOLUTION' THEN 0 ELSE 1 END AS sort_priority
            FROM position_ledgers
            WHERE account = %(a)s
              AND instrument_id = %(i)s
            ORDER BY symbol_subtype, sort_priority ASC, transaction_time DESC
        """, engine, params={"a": account, "i": instrument_id})
        for _, row in df.iterrows():
            side     = row["symbol_subtype"]
            received = float(row["after_realized"] or 0)
            bought   = int(row["after_qty_bought"] or 0)
            sold     = int(row["after_qty_sold"] or 0)
            cost     = abs(float(row["before_cost"] or 0))
            net_pnl  = received - cost   # true net: cash received minus cost paid
            qty      = bought if bought > 0 else sold
            avg_cost = round(cost / qty, 4) if qty > 0 else None
            pnl_per  = round(net_pnl / qty, 4) if qty > 0 else None
            result[side] = {
                "net_pnl":         round(net_pnl, 2),
                "received":        round(received, 2),
                "cost":            round(cost, 2),
                "qty":             qty,
                "avg_cost":        avg_cost,
                "implied_pnl_per": pnl_per,
                "resolved":        row["entry_type"] == "RESOLUTION",
            }
    except Exception as e:
        log.warning(f"  fetch_position_detail error: {e}")
    return result


def _format_position_note(account: str, sides: dict) -> tuple[str, list]:
    """
    Build a position context string and list of red flags from per-side position data.
    Uses true net P&L = cash received - cost paid.
    Returns (note_str, flag_list).
    """
    parts     = []
    flags     = []
    total_net = sum(s.get("net_pnl", 0) for s in sides.values())

    for side in ("YES", "NO"):
        s = sides.get(side)
        if not s:
            continue
        qty  = s.get("qty", 0)
        ac   = s.get("avg_cost")
        net  = s.get("net_pnl", 0.0)
        recv = s.get("received", 0.0)
        cost = s.get("cost", 0.0)
        ipp  = s.get("implied_pnl_per")
        res  = s.get("resolved", False)
        if ac is not None and qty > 0:
            pnl_per_str = (
                f"~${ipp:.2f}/contract profit" if ipp is not None and ipp >= 0
                else f"~${abs(ipp):.2f}/contract loss" if ipp is not None
                else ""
            )
            parts.append(
                f"{side}: {qty:,} contracts @ avg ${ac:.2f}, "
                f"net {'+' if net >= 0 else ''}{net:,.2f} ({pnl_per_str})"
                + (" [RESOLVED]" if res else "")
            )
            if ipp is not None and ipp > 0.60:
                flags.append(
                    f"{account} {side}-side profit of ${ipp:.2f}/contract is unusually high "
                    f"(avg cost ${ac:.2f} — contracts acquired at steep discount to resolution value)"
                )

    total_str = f"${total_net:,.2f}" if total_net >= 0 else f"-${abs(total_net):,.2f}"
    note = (
        f"Position context for {account}: total net P&L {total_str}. "
        f"Side breakdown — {'; '.join(parts)}."
        if parts else ""
    )
    return note, flags


# =============================================================================
# DESCRIPTION PARSER
# =============================================================================

def parse_description(description: str) -> dict:
    """Extract structured fields from the alert description text."""
    result = {
        "account_path": None, "account_id": None, "firm": None,
        "instrument": None,
        "imbalance_start": None, "imbalance_end": None,
        "dominant_side": None, "dominant_qty": None,
        "minor_side": None,    "minor_qty": None,
        "multiplier": None,
        "fill_side": None,     "fill_qty": None,
        "no_fill_side": None,
        "has_criteria3_period": False, "criteria3_start": None,
        "raw_description": description,
    }
    if not description:
        return result

    def _parse_ts(ts_str, tz_str):
        try:
            th = int(tz_str[:3]);  tm = int(tz_str[0] + tz_str[3:])
            tz = timezone(timedelta(hours=th, minutes=tm))
            return datetime.strptime(ts_str[:26], "%Y-%m-%d %H:%M:%S.%f").replace(tzinfo=tz)
        except Exception:
            return ts_str

    m = re.search(r"Account (firms/([^/]+)/accounts/(\S+)) is suspected", description)
    if m:
        result["account_path"] = m.group(1)
        result["firm"]         = m.group(2)
        result["account_id"]   = m.group(3).rstrip(".")

    m = re.search(r"on instrument (\S+)\.", description)
    if m:
        result["instrument"] = m.group(1)

    m = re.search(r"At (\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d+) ([+-]\d{4}) \w+,", description)
    if m:
        result["imbalance_start"] = _parse_ts(m.group(1), m.group(2))

    m = re.search(
        r"and (\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d+) ([+-]\d{4}) \w+,\s+the account filled",
        description
    )
    if m:
        result["imbalance_end"] = _parse_ts(m.group(1), m.group(2))

    m = re.search(
        r"had ([\d.]+) times more active order quantity on the (\w+) side \((\d+)\)"
        r" than the (\w+) side \((\d+)\)",
        description
    )
    if m:
        result["multiplier"]    = float(m.group(1))
        result["dominant_side"] = m.group(2)
        result["dominant_qty"]  = int(m.group(3))
        result["minor_side"]    = m.group(4)
        result["minor_qty"]     = int(m.group(5))

    m = re.search(
        r"filled exclusively on the (\w+) side \((\d+)\) with no fills on the (\w+) side \((\d+)\)",
        description
    )
    if m:
        result["fill_side"]    = m.group(1)
        result["fill_qty"]     = int(m.group(2))
        result["no_fill_side"] = m.group(3)

    if "additional Criteria 3 check period" in description:
        result["has_criteria3_period"] = True
        m = re.search(r"began at (\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d+)", description)
        if m:
            result["criteria3_start"] = m.group(1)

    return result


# =============================================================================
# CLASSIFICATION
# =============================================================================

def _dom_new_prices(dom_execs: pd.DataFrame) -> pd.Series:
    return dom_execs[dom_execs["execution_type"] == "NEW"]["price"]

def _fill_prices(execs: pd.DataFrame, fill_side: str) -> pd.Series:
    return execs[
        (execs["symbol_subtype"] == fill_side) &
        (execs["execution_type"].isin(["FILL", "PARTIAL_FILL"]))
    ]["price"]

def _resting_hours(execs: pd.DataFrame, fill_time) -> float:
    new_ts = execs[execs["execution_type"] == "NEW"]["transact_time"]
    if new_ts.empty or fill_time is None:
        return 0.0
    try:
        delta = pd.to_datetime(fill_time, utc=True) - pd.to_datetime(new_ts.min(), utc=True)
        return max(delta.total_seconds() / 3600, 0.0)
    except Exception:
        return 0.0

def _hstr(hours: float) -> str:
    return f"{int(hours // 24)} days" if hours >= 48 else f"{hours:.1f} hours"


def classify_alert(parsed: dict, execs: pd.DataFrame) -> tuple[str, str]:
    """
    Returns (classification, adjudication_note).
    Priority: MARKET_MAKER > WEATHER_VOLATILITY > RESTING_BIDS_LONG_DUR
              > NON_COMPETITIVE_BIDS > GTC_RESTING > NEEDS_REVIEW
    """
    acct     = parsed.get("account_id", "")
    symbol   = parsed.get("instrument", "")
    dom_side = parsed.get("dominant_side", "")
    fil_side = parsed.get("fill_side", "")
    dom_qty  = parsed.get("dominant_qty", 0) or 0
    min_qty  = parsed.get("minor_qty", 0) or 0
    min_side = parsed.get("minor_side", "")
    fil_qty  = parsed.get("fill_qty", 0) or 0
    mult     = parsed.get("multiplier", 0)

    # 1 — MARKET MAKER
    if acct in MARKET_MAKER_ACCOUNTS:
        return "MARKET_MAKER", (
            f"Account {acct} is a known market maker. "
            f"Activity consistent with known market making algo providing two-sided liquidity. "
            f"No concerns."
        )

    # 2 — WEATHER VOLATILITY
    if (symbol or "").upper().startswith(WEATHER_SYMBOL_PREFIXES):
        return "WEATHER_VOLATILITY", (
            f"Account {acct} activity on weather forecast contract {symbol} near maturity. "
            f"Price and volume volatility are expected to rise as contracts approach expiration, "
            f"and observed imbalances ({dom_qty:,} {dom_side} vs {min_qty:,} {min_side}) "
            f"are most likely influenced by weather forecast updates and/or liquidity. "
            f"No indicators of disorderly trading were identified. "
            f"Exchange does not find this activity problematic. No further action required."
        )

    if not execs.empty and dom_side and fil_side:
        dom_execs = execs[execs["symbol_subtype"] == dom_side].copy()
        fil_execs = execs[execs["symbol_subtype"] == fil_side].copy()
        dom_px    = _dom_new_prices(dom_execs)
        fil_px    = _fill_prices(execs, fil_side)
        avg_dom   = dom_px.mean() if not dom_px.empty else None
        avg_fil   = fil_px.mean() if not fil_px.empty else None

        # 3 — RESTING BIDS LONG DURATION
        fil_events = fil_execs[fil_execs["execution_type"].isin(["FILL", "PARTIAL_FILL"])]
        if not fil_events.empty:
            first_fill = fil_events["transact_time"].min()
            dom_h = _resting_hours(dom_execs, first_fill)
            fil_h = _resting_hours(fil_execs,  first_fill)
            if dom_h >= RESTING_DURATION_HOURS and fil_h >= RESTING_DURATION_HOURS:
                return "RESTING_BIDS_LONG_DUR", (
                    f"At time of flagged activity, account {acct} had {dom_qty:,} {dom_side} "
                    f"orders vs {min_qty:,} {min_side} orders ({mult}x imbalance) in {symbol}. "
                    f"Dominant-side ({dom_side}) bids had rested in the market unmodified for "
                    f"{_hstr(dom_h)} before the flagged fill-side execution. "
                    f"Fill-side ({fil_side}) bids had rested unmodified for {_hstr(fil_h)}. "
                    f"Leaving bids open on both sides for an extended period without modification "
                    f"is not indicative of spoofing behavior. "
                    f"Exchange does not find this activity problematic."
                )

        # 4 — NON-COMPETITIVE BIDS
        if avg_dom is not None and avg_fil is not None:
            gap        = abs(avg_dom - avg_fil)
            fill_ratio = fil_qty / dom_qty if dom_qty > 0 else 1.0
            if gap >= NON_COMPETITIVE_PRICE_THRESHOLD and fill_ratio <= 0.20:
                return "NON_COMPETITIVE_BIDS", (
                    f"At time of flagged activity, account {acct} had {dom_qty:,} {dom_side} "
                    f"orders vs {min_qty:,} {min_side} orders ({mult}x imbalance) in {symbol}. "
                    f"Dominant-side ({dom_side}) bids were priced an average of ${gap:.2f} "
                    f"away from prevailing {fil_side}-side execution prices "
                    f"(dominant avg: ${avg_dom:.2f}, fill-side avg: ${avg_fil:.2f}). "
                    f"Bids priced this far from the market are not competitive and not consistent "
                    f"with spoofing, which requires a credible threat of execution. "
                    f"No other signs of potential manipulation. "
                    f"Exchange does not find this activity problematic."
                )

        # 5 — GTC RESTING
        if not dom_execs.empty:
            tif_vals   = dom_execs["time_in_force"].unique()
            is_gtc     = all(t in ("GOOD_TILL_TIME", "GOOD_TILL_CANCEL") for t in tif_vals)
            never_fill = not any(
                t in ("FILL", "PARTIAL_FILL") for t in dom_execs["execution_type"].unique()
            )
            gap_gtc   = abs(avg_dom - avg_fil) if avg_dom is not None and avg_fil is not None else None
            price_far = gap_gtc is not None and gap_gtc >= GTC_PRICE_DISTANCE_THRESHOLD

            if is_gtc and never_fill and price_far:
                new_ts = dom_execs[dom_execs["execution_type"] == "NEW"]["transact_time"]
                exp_ts = dom_execs[
                    dom_execs["execution_type"].isin(["EXPIRED", "CANCELED", "DONE_FOR_DAY"])
                ]["transact_time"]
                t_sit = None
                if not new_ts.empty and not exp_ts.empty:
                    t_sit = pd.to_datetime(exp_ts.max(), utc=True) - pd.to_datetime(new_ts.min(), utc=True)
                t_str   = str(t_sit).split(".")[0] if t_sit else "an extended period"
                gap_str = f"${gap_gtc:.2f}" if gap_gtc else "significantly"
                return "GTC_RESTING", (
                    f"Account {acct} had unexecuted GTT/GTC orders totaling {dom_qty:,} "
                    f"{dom_side} contracts in {symbol} that sat on the market for {t_str} "
                    f"and were priced {gap_str} away from prevailing market prices at the time "
                    f"of execution. "
                    f"Inactive orders resting away from the market are not indicative of spoofing. "
                    f"No concerns."
                )

    # 6 — NEEDS REVIEW
    return "NEEDS_REVIEW", (
        f"Alert requires analyst review. "
        f"Account {acct} had {dom_qty:,} {dom_side} orders vs {min_qty:,} {min_side} "
        f"orders ({mult}x imbalance). "
        f"Fills occurred on {fil_side} side. "
        f"Execution pattern does not match known auto-dismiss scenarios."
    )


# =============================================================================
# CHART DATA BUILDER  (vectorized — handles 30k+ rows efficiently)
# =============================================================================

def build_chart_data(acct_execs: pd.DataFrame, mkt_execs: pd.DataFrame,
                     dom_side: str, fil_side: str) -> pd.DataFrame:
    """
    Returns a time-bucketed DataFrame for the activity chart.
    Fully vectorized with pandas resample — O(n log n), safe on large datasets.

    Columns:
        time_label    — HH:MM string for X axis
        last_price    — last fill price per bucket, forward-filled
        acct_dom_bid  — last accused-account NEW order price on dominant side
        acct_fil_bid  — last accused-account NEW order price on fill side
        best_dom_bid  — highest market NEW order price on dominant side
        best_fil_bid  — highest market NEW order price on fill side

    Bucket width auto-scaled by window length:
        window <= 2h  -> 1-minute buckets
        window <= 12h -> 5-minute buckets
        else          -> 30-minute buckets
    """
    frames = [
        pd.to_datetime(df["transact_time"], utc=True).dropna()
        for df in [acct_execs, mkt_execs]
        if not df.empty and "transact_time" in df.columns
    ]
    if not frames:
        return pd.DataFrame()

    all_ts   = pd.concat(frames)
    t_min    = all_ts.min()
    t_max    = all_ts.max()
    window_h = (t_max - t_min).total_seconds() / 3600
    freq     = "1min" if window_h <= 2 else ("5min" if window_h <= 12 else "30min")

    def _prep(df: pd.DataFrame) -> pd.DataFrame:
        """Attach a bucket column to a copy of df."""
        if df.empty:
            return df
        out = df.copy()
        out["_ts"] = pd.to_datetime(out["transact_time"], utc=True)
        out["_bucket"] = out["_ts"].dt.floor(freq)
        return out

    # ---- Last traded price per bucket (from market fills), forward-filled ----
    price_series = pd.Series(dtype=float)
    if not mkt_execs.empty:
        mf = _prep(mkt_execs)
        fills = mf[mf["execution_type"].isin(["FILL", "PARTIAL_FILL"])]
        if not fills.empty:
            # Last fill price in each bucket
            price_series = (
                fills.groupby("_bucket")["price"]
                .last()
                .astype(float)
            )

    # ---- Accused account: last NEW price per bucket per side ----
    acct_dom_series = acct_fil_series = pd.Series(dtype=float)
    if not acct_execs.empty:
        af = _prep(acct_execs)
        new_orders = af[af["execution_type"] == "NEW"]
        for side, target in [(dom_side, "dom"), (fil_side, "fil")]:
            s = new_orders[new_orders["symbol_subtype"] == side]
            if not s.empty:
                ser = s.groupby("_bucket")["price"].last().astype(float)
                if target == "dom":
                    acct_dom_series = ser
                else:
                    acct_fil_series = ser

    # ---- Market: best (highest) NEW bid per bucket per side ----
    best_dom_series = best_fil_series = pd.Series(dtype=float)
    if not mkt_execs.empty:
        mf = _prep(mkt_execs)
        new_orders = mf[mf["execution_type"] == "NEW"]
        for side, target in [(dom_side, "dom"), (fil_side, "fil")]:
            s = new_orders[new_orders["symbol_subtype"] == side]
            if not s.empty:
                ser = s.groupby("_bucket")["price"].max().astype(float)
                if target == "dom":
                    best_dom_series = ser
                else:
                    best_fil_series = ser

    # ---- Combine into a single DataFrame on a complete bucket index ----
    idx = pd.date_range(start=t_min.floor(freq), end=t_max.floor(freq), freq=freq, tz="UTC")
    result = pd.DataFrame(index=idx)
    result["last_price"]   = price_series.reindex(idx)
    result["acct_dom_bid"] = acct_dom_series.reindex(idx)
    result["acct_fil_bid"] = acct_fil_series.reindex(idx)
    result["best_dom_bid"] = best_dom_series.reindex(idx)
    result["best_fil_bid"] = best_fil_series.reindex(idx)

    # Forward-fill last_price so the chart shows a continuous price line
    result["last_price"] = result["last_price"].ffill()

    # Drop rows where everything is null (no data at all in that bucket)
    result = result.dropna(how="all")

    result.insert(0, "time_label", result.index.strftime("%H:%M"))
    result = result.reset_index(drop=True)
    return result


# =============================================================================
# EXCEL REPORT
# =============================================================================

def _safe_name(s: str) -> str:
    """Strip characters illegal in Windows filenames."""
    return re.sub(r'[\\/:*?"<>|]', "_", s or "")


def get_output_path(alert_id: str, account: str, symbol: str) -> Path:
    fname = _safe_name(f"Spoofing_{alert_id}_{account}_{symbol}.xlsx")
    return REVIEWS_DIR / fname


def write_alert_report(alert_id, alert_row, parsed, classification, adj_note,
                       acct_execs, mkt_execs):
    """Produce a three-sheet Excel workbook for a single alert."""
    wb = Workbook()
    wb.remove(wb.active)   # start clean

    _sheet_summary(wb, alert_id, alert_row, parsed, classification, adj_note)
    _sheet_chart(wb, acct_execs, mkt_execs, parsed, classification)
    _sheet_executions(wb, acct_execs, parsed)

    path = get_output_path(
        alert_id,
        parsed.get("account_id", ""),
        parsed.get("instrument", "") or alert_row.get("symbol", "")
    )
    wb.save(path)
    log.info(f"  Report saved: {path}")
    return path


# ──────────────────────────────────────────────
# Sheet 1: Summary
# ──────────────────────────────────────────────

def _sheet_summary(wb, alert_id, alert_row, parsed, classification, adj_note):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 1
    ws.column_dimensions["D"].width = 1

    r = 1

    # ── Title bar ──
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(r, 1, "ForecastEx Compliance  |  Spoofing Alert Report")
    c.fill = _fill(HEX["navy"]); c.font = _font(bold=True, color="FFFFFF", size=13)
    c.alignment = _align("center"); ws.row_dimensions[r].height = 26
    r += 1

    # ── Classification banner ──
    color = CLASS_COLOR.get(classification, HEX["cls_nr"])
    label = CLASS_LABEL.get(classification, classification)
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(r, 1, label)
    c.fill = _fill(color); c.font = _font(bold=True, size=12)
    c.alignment = _align("center"); ws.row_dimensions[r].height = 22
    r += 2

    # ── Metadata grid ──
    def meta(label_text, value):
        nonlocal r
        lc = ws.cell(r, 1, label_text)
        lc.fill = _fill(HEX["label_bg"]); lc.font = _font(bold=True, size=10)
        lc.border = _thin_border(); lc.alignment = _align()
        ws.merge_cells(f"B{r}:D{r}")
        vc = ws.cell(r, 2, str(value) if value is not None else "—")
        vc.font = _font(size=10); vc.border = _thin_border()
        vc.alignment = _align(wrap=True)
        ws.row_dimensions[r].height = 15
        r += 1

    def _fmt_ts(val):
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d %H:%M:%S %Z")
        return str(val) if val else "—"

    meta("Alert ID",              alert_id)
    meta("Alert Time",            _fmt_ts(alert_row.get("alert_time")))
    meta("Symbol / Instrument",   parsed.get("instrument") or alert_row.get("symbol", ""))
    meta("Account",               parsed.get("account_id", ""))
    meta("Firm",                  parsed.get("firm", "") or alert_row.get("firm", ""))
    meta("Classification",        classification)
    meta("Imbalance Multiplier",  f"{parsed.get('multiplier', '')}x")
    meta("Dominant Side",         parsed.get("dominant_side", ""))
    meta("Dominant Qty",          f"{(parsed.get('dominant_qty') or 0):,}")
    meta("Fill Side",             parsed.get("fill_side", ""))
    meta("Fill Qty",              f"{(parsed.get('fill_qty') or 0):,}")
    meta("Imbalance Start",       _fmt_ts(parsed.get("imbalance_start")))
    meta("Imbalance End",         _fmt_ts(parsed.get("imbalance_end")))
    meta("Criteria 3 Period",     "Yes" if parsed.get("has_criteria3_period") else "No")
    r += 1

    # ── Adjudication note (prominent gold box) ──
    ws.merge_cells(f"A{r}:B{r}")
    hc = ws.cell(r, 1, "SUGGESTED ADJUDICATION VERBIAGE  — copy and paste if you agree")
    hc.fill = _fill(HEX["navy2"]); hc.font = _font(bold=True, color="FFFFFF", size=11)
    hc.alignment = _align("center"); ws.row_dimensions[r].height = 20
    r += 1

    note_start = r
    note_rows  = 8
    ws.merge_cells(f"A{r}:B{r + note_rows - 1}")
    nc = ws.cell(r, 1, adj_note)
    nc.fill = _fill(HEX["adj_bg"])
    nc.font = _font(size=11)
    nc.alignment = _align(wrap=True, v="top")
    nc.border = _medium_border()
    for nr in range(r, r + note_rows):
        ws.row_dimensions[nr].height = 20
    r += note_rows + 2

    # ── Raw description ──
    ws.merge_cells(f"A{r}:B{r}")
    rc = ws.cell(r, 1, "RAW ALERT DESCRIPTION")
    rc.fill = _fill(HEX["section_bg"]); rc.font = _font(bold=True, size=10)
    rc.alignment = _align(); ws.row_dimensions[r].height = 16
    r += 1

    desc_rows = 6
    ws.merge_cells(f"A{r}:B{r + desc_rows - 1}")
    dc = ws.cell(r, 1, parsed.get("raw_description", ""))
    dc.fill = _fill("FAFAFA"); dc.font = _font(size=9, color="555555")
    dc.alignment = _align(wrap=True, v="top")
    for dr in range(r, r + desc_rows):
        ws.row_dimensions[dr].height = 18


# ──────────────────────────────────────────────
# Sheet 2: Activity Chart
# ──────────────────────────────────────────────

def _sheet_chart(wb, acct_execs, mkt_execs, parsed, classification):
    ws = wb.create_sheet("Activity Chart", 1)
    ws.sheet_view.showGridLines = False

    dom_side = parsed.get("dominant_side", "YES")
    fil_side = parsed.get("fill_side", "NO")
    account  = parsed.get("account_id", "")
    symbol   = parsed.get("instrument", "")

    # Title
    ws.merge_cells("A1:N1")
    tc = ws.cell(1, 1, f"Trading Activity  —  {account}  /  {symbol}")
    tc.fill = _fill(HEX["navy"]); tc.font = _font(bold=True, color="FFFFFF", size=13)
    tc.alignment = _align("center"); ws.row_dimensions[1].height = 24

    if acct_execs.empty:
        ws["A3"] = "No execution data — chart cannot be generated."
        ws["A3"].font = _font(italic=True, color="888888")
        return

    chart_df = build_chart_data(acct_execs, mkt_execs, dom_side, fil_side)

    if chart_df.empty or len(chart_df) < 2:
        ws["A3"] = "Insufficient data points to render chart."
        ws["A3"].font = _font(italic=True, color="888888")
        return

    # ── Write data table (chart source) starting at row 42 ──
    DATA_ROW = 42
    col_defs = [
        ("Time",                       "time_label",   14),
        ("Last Price ($)",             "last_price",   16),
        (f"Acct {dom_side} Bid ($)",   "acct_dom_bid", 18),
        (f"Acct {fil_side} Bid ($)",   "acct_fil_bid", 18),
        (f"Best {dom_side} Bid — Mkt", "best_dom_bid", 22),
        (f"Best {fil_side} Bid — Mkt", "best_fil_bid", 22),
    ]
    n_cols = len(col_defs)

    # Section label
    ws.merge_cells(f"A{DATA_ROW - 1}:{get_column_letter(n_cols)}{DATA_ROW - 1}")
    sl = ws.cell(DATA_ROW - 1, 1, "Underlying Chart Data")
    sl.fill = _fill(HEX["section_bg"]); sl.font = _font(bold=True, size=10)
    sl.alignment = _align()

    # Column headers
    for ci, (hdr, _, width) in enumerate(col_defs, 1):
        c = ws.cell(DATA_ROW, ci, hdr)
        c.fill = _fill(HEX["navy2"]); c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _align("center"); c.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Data rows
    for ri, (_, row_data) in enumerate(chart_df.iterrows(), DATA_ROW + 1):
        bg = HEX["alt_row"] if (ri - DATA_ROW) % 2 == 0 else "FFFFFF"
        for ci, (_, key, _) in enumerate(col_defs, 1):
            val = row_data.get(key)
            if isinstance(val, float):
                val = round(val, 4)
            c = ws.cell(ri, ci, val)
            c.fill = _fill(bg); c.font = _font(size=9)
            c.border = _thin_border()
            c.alignment = _align("center" if ci > 1 else "left")
        ws.row_dimensions[ri].height = 13

    n_data  = len(chart_df)
    dr_s    = DATA_ROW + 1
    dr_e    = DATA_ROW + n_data

    # ── Build chart ──
    chart = LineChart()
    chart.title  = f"{account} — {symbol}"
    chart.style  = 10
    chart.width  = 28
    chart.height = 15
    chart.y_axis.title  = "Price ($)"
    chart.x_axis.title  = "Time"
    chart.y_axis.numFmt = "0.00"
    chart.legend.position = "b"

    series_defs = [
        # (col_index, title, hex_color, line_width_emu, dashed)
        (2, "Last Price",                    HEX["price"],    38100, False),
        (3, f"Acct {dom_side} Bid",          HEX["acct_dom"], 25400, True),
        (4, f"Acct {fil_side} Bid",          HEX["acct_fil"], 25400, True),
        (5, f"Best {dom_side} (Market)",     HEX["best_dom"], 19050, True),
        (6, f"Best {fil_side} (Market)",     HEX["best_fil"], 19050, True),
    ]

    for col_i, title, color, width_emu, dashed in series_defs:
        vals = Reference(ws, min_col=col_i, max_col=col_i, min_row=dr_s, max_row=dr_e)
        s = chart.series.append(vals)
        # openpyxl Series object returned from append
    # Rebuild cleanly using proper API
    chart.series.clear()
    from openpyxl.chart import Series as ChartSeries
    for col_i, title, color, width_emu, dashed in series_defs:
        vals = Reference(ws, min_col=col_i, max_col=col_i, min_row=dr_s, max_row=dr_e)
        s = ChartSeries(vals, title=title)
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width = width_emu
        if dashed:
            s.graphicalProperties.line.dashDot = "dash"
        s.smooth = False
        chart.series.append(s)

    cats = Reference(ws, min_col=1, max_col=1, min_row=dr_s, max_row=dr_e)
    chart.set_categories(cats)

    ws.add_chart(chart, "A3")

    # ── Classification annotation below chart ──
    ann_row = 40
    imb_s = parsed.get("imbalance_start")
    imb_e = parsed.get("imbalance_end")
    window_str = ""
    if imb_s and imb_e:
        try:
            s_str = imb_s.strftime("%H:%M:%S") if hasattr(imb_s, "strftime") else str(imb_s)
            e_str = imb_e.strftime("%H:%M:%S") if hasattr(imb_e, "strftime") else str(imb_e)
            window_str = f"    |    Alert window: {s_str} → {e_str}"
        except Exception:
            pass

    ws.merge_cells(f"A{ann_row}:{get_column_letter(n_cols)}{ann_row}")
    ac = ws.cell(ann_row, 1,
        f"Classification: {CLASS_LABEL.get(classification, classification)}{window_str}"
    )
    ac.fill = _fill(CLASS_COLOR.get(classification, HEX["cls_nr"]))
    ac.font = _font(bold=True, size=10)
    ac.alignment = _align("center")
    ws.row_dimensions[ann_row].height = 18

    # ── Color legend ──
    leg_row = ann_row + 2
    ws.cell(leg_row, 1, "Series legend:").font = _font(bold=True, size=9)
    legend_items = [
        ("Last Price",            HEX["price"]),
        (f"Acct {dom_side} Bid",  HEX["acct_dom"]),
        (f"Acct {fil_side} Bid",  HEX["acct_fil"]),
        (f"Best {dom_side} Mkt",  HEX["best_dom"]),
        (f"Best {fil_side} Mkt",  HEX["best_fil"]),
    ]
    for ci, (lbl, clr) in enumerate(legend_items, 2):
        lc = ws.cell(leg_row, ci, lbl)
        lc.fill = _fill(clr); lc.font = _font(size=9, color="FFFFFF")
        lc.alignment = _align("center"); lc.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = max(
            ws.column_dimensions[get_column_letter(ci)].width, len(lbl) + 3
        )


# ──────────────────────────────────────────────
# Sheet 3: Execution Detail
# ──────────────────────────────────────────────

def _sheet_executions(wb, execs: pd.DataFrame, parsed: dict):
    ws = wb.create_sheet("Execution Detail", 2)
    ws.sheet_view.showGridLines = False

    dom_side = parsed.get("dominant_side", "")
    fil_side = parsed.get("fill_side", "")

    # Title
    title_text = (
        f"Execution Detail  —  "
        f"{parsed.get('account_id','')}  /  {parsed.get('instrument','')}"
    )
    col_count = 12
    ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
    tc = ws.cell(1, 1, title_text)
    tc.fill = _fill(HEX["navy"]); tc.font = _font(bold=True, color="FFFFFF", size=12)
    tc.alignment = _align("center"); ws.row_dimensions[1].height = 22

    if execs.empty:
        ws["A3"] = "No execution data available for this alert."
        ws["A3"].font = _font(italic=True, color="888888")
        return

    headers = [
        ("Time (UTC)",   "transact_time",  22),
        ("Side",         "symbol_subtype",   8),
        ("Exec Type",    "execution_type",  18),
        ("Status",       "order_status",    16),
        ("TIF",          "time_in_force",   20),
        ("Order Qty",    "order_quantity",  12),
        ("Fill Qty",     "fill_quantity",   10),
        ("Cum Qty",      "cum_quantity",    10),
        ("Leaves Qty",   "leaves_quantity", 12),
        ("Price ($)",    "price",           10),
        ("Aggressor",    "aggressor",       10),
        ("Order ID",     "order_id",        28),
    ]

    # Column headers
    for ci, (hdr, _, width) in enumerate(headers, 1):
        c = ws.cell(2, ci, hdr)
        c.fill = _fill(HEX["navy2"]); c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _align("center"); c.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 18

    FILLS = {"FILL", "PARTIAL_FILL"}

    for ri, (_, row_data) in enumerate(execs.iterrows(), 3):
        side      = str(row_data.get("symbol_subtype", ""))
        exec_type = str(row_data.get("execution_type", ""))
        is_fill   = exec_type in FILLS

        if side == dom_side:
            bg = "FFF0A0" if is_fill else "FFF9E6"
        elif side == fil_side:
            bg = "C5E8FF" if is_fill else "E8F5FF"
        else:
            bg = "FFFFFF"

        for ci, (_, col_key, _) in enumerate(headers, 1):
            val = row_data.get(col_key, "")
            if col_key == "transact_time" and hasattr(val, "strftime"):
                val = val.strftime("%Y-%m-%d %H:%M:%S")
            elif col_key == "price" and val is not None:
                try:
                    val = round(float(val), 4)
                except Exception:
                    pass
            elif col_key == "aggressor" and val is not None:
                val = "Yes" if val else "No"
            c = ws.cell(ri, ci, val if val is not None else "")
            c.fill = _fill(bg); c.font = _font(size=10)
            c.border = _thin_border()
            c.alignment = _align("left" if ci == 1 else "center")
        ws.row_dimensions[ri].height = 15

    ws.freeze_panes = "A3"

    # Color key
    leg_row = len(execs) + 5
    ws.cell(leg_row, 1, "Color key:").font = _font(bold=True, size=9)
    for ci, (lbl, bg) in enumerate([
        (f"{dom_side} order (dominant side)", "FFF9E6"),
        (f"{dom_side} fill",                  "FFF0A0"),
        (f"{fil_side} order (fill side)",     "E8F5FF"),
        (f"{fil_side} fill",                  "C5E8FF"),
    ], 2):
        c = ws.cell(leg_row, ci, lbl)
        c.fill = _fill(bg); c.font = _font(size=9)
        c.border = _thin_border(); c.alignment = _align("center")


# =============================================================================
# STATE MANAGEMENT
# =============================================================================

LOOKBACK_DAYS = 14  # Must match the alert query lookback window


def load_state() -> dict:
    if STATE_FILE.exists():
        with open(STATE_FILE) as f:
            return json.load(f)
    return {"processed_alert_ids": [], "last_run": None}


def prune_state(processed_ids: set) -> set:
    """
    Remove alert IDs that are older than the lookback window.
    Queries the alerts table to find which IDs are still within the window.
    IDs not found in the DB (too old) are dropped from the state file.
    This prevents the state file from growing indefinitely and ensures
    alerts reprocessed after a fix get picked up correctly.
    """
    if not processed_ids:
        return processed_ids
    try:
        engine = get_engine()
        with engine.connect() as conn:
            id_list = list(processed_ids)
            result = conn.execute(
                f"""
                SELECT alert_id FROM alerts
                WHERE alert_id = ANY(ARRAY{id_list!r}::text[])
                  AND created_at >= NOW() - INTERVAL '{LOOKBACK_DAYS} days'
                """
            )
            still_valid = {row[0] for row in result}
            pruned = processed_ids - still_valid
            if pruned:
                log.info(f"  Pruned {len(pruned)} expired alert ID(s) from state file.")
            return still_valid
    except Exception as e:
        log.warning(f"  State pruning skipped: {e}")
        return processed_ids


def save_state(state: dict):
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2, default=str)


# =============================================================================
# SINGLE POLL CYCLE
# =============================================================================

def run_once(engine, processed_ids: set) -> set:
    alerts_df = fetch_new_spoofing_alerts(engine, processed_ids)
    if alerts_df.empty:
        return processed_ids

    for _, alert_row in alerts_df.iterrows():
        alert_id = alert_row["alert_id"]
        symbol   = alert_row.get("symbol", "")
        log.info(f"Processing {alert_id} -- {symbol}")

        try:
            parsed = parse_description(alert_row.get("description", ""))

            # Resolve account
            account_id = parsed.get("account_id")
            if not account_id:
                raw = alert_row.get("accounts", [])
                if raw and len(raw) > 0:
                    account_id = str(raw[0]).strip("{}")
                    parsed["account_id"] = account_id

            # Fetch accused-account executions
            acct_execs = pd.DataFrame()
            if account_id and symbol:
                try:
                    acct_execs = fetch_executions(engine, account_id, symbol)
                except Exception as e:
                    log.warning(f"  acct executions failed: {e}")

            # Fetch market-wide executions for chart (alert window ± 30 min)
            mkt_execs = pd.DataFrame()
            imb_s = parsed.get("imbalance_start")
            imb_e = parsed.get("imbalance_end")
            if symbol and imb_s and imb_e:
                try:
                    chart_start = pd.to_datetime(imb_s, utc=True) - timedelta(minutes=30)
                    chart_end   = pd.to_datetime(imb_e, utc=True) + timedelta(minutes=30)
                    mkt_execs   = fetch_market_executions(engine, symbol, chart_start, chart_end)
                except Exception as e:
                    log.warning(f"  market executions failed: {e}")

            # Fetch position detail for P&L context
            position_detail = {}
            if account_id and symbol:
                try:
                    position_detail = fetch_position_detail(engine, account_id, symbol)
                except Exception as e:
                    log.warning(f"  position detail failed: {e}")

            classification, adj_note = classify_alert(parsed, acct_execs)
            log.info(f"  -> {classification}")

            # Append position context to adjudication note
            if position_detail:
                pos_note, pos_flags = _format_position_note(account_id, position_detail)
                if pos_note:
                    adj_note = adj_note.rstrip() + " " + pos_note
                if pos_flags:
                    flag_str = "; ".join(pos_flags)
                    adj_note = adj_note.rstrip() + (
                        f" NOTE: The following position factors warrant attention: {flag_str}."
                    )

            write_alert_report(
                alert_id, alert_row, parsed,
                classification, adj_note,
                acct_execs, mkt_execs,
            )

            processed_ids.add(alert_id)

        except Exception as e:
            log.error(f"  Error on {alert_id}: {e}\n{traceback.format_exc()}")

    return processed_ids


# =============================================================================
# MAIN — continuous polling loop
# =============================================================================

def main():
    log.info("=" * 65)
    log.info("ForecastEx Spoofing Alert Monitor")
    log.info(f"Output  : {OUTPUT_DIR}")
    log.info(f"Interval: every {POLL_INTERVAL_SECONDS // 60} minutes")
    log.info("Press Ctrl+C to stop.")
    log.info("=" * 65)

    state         = load_state()
    processed_ids = set(state.get("processed_alert_ids", []))

    try:
        engine = get_engine()
        log.info("Database connection established.")
    except Exception as e:
        log.error(f"Failed to connect to database: {e}")
        return

    while True:
        log.info(f"--- Poll: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---")
        try:
            processed_ids = run_once(engine, processed_ids)
        except Exception as e:
            log.error(f"Poll cycle error: {e}\n{traceback.format_exc()}")

        processed_ids = prune_state(processed_ids)
        state["processed_alert_ids"] = list(processed_ids)
        state["last_run"] = datetime.now().isoformat()
        save_state(state)

        log.info(f"Sleeping {POLL_INTERVAL_SECONDS // 60} min...")
        try:
            time.sleep(POLL_INTERVAL_SECONDS)
        except KeyboardInterrupt:
            log.info("Stopped by user.")
            break



if __name__ == "__main__":
    try:
        main()
    except Exception as _e:
        import traceback as _tb
        _crash_path = OUTPUT_DIR / "spoofing_monitor_crash.txt"
        try:
            _crash_path.write_text(
                datetime.now().isoformat() + "\n" + _tb.format_exc(),
                encoding="utf-8"
            )
            print("CRASH -- details written to: " + str(_crash_path))
        except Exception:
            print("CRASH: " + str(_e))
    finally:
        input("\nPress Enter to close...")
